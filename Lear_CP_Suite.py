#ODM


from tkinter import *
import tkinter as tk
import pyautogui
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from tkinter import filedialog
import webbrowser
#XML Extraction----------------------------------------------------------------------------------------------------------------------------------------

import xml.etree.cElementTree as ET
import shutil
import string
import PyPDF2 
from tkinter.ttk import Progressbar
import tkinter.ttk as ttk

from tkinter import *
import tkinter as tk
import pyautogui
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, Font
from tkinter import filedialog
import string
import tkinter.ttk as ttk
import time
import dictdiffer
from tkinter.filedialog import askdirectory
from openpyxl.utils.cell import get_column_letter
import tkinter as tk
from tkinter import *
from bs4 import BeautifulSoup
import os
from openpyxl import workbook
import os,pyautogui,time
from PIL import Image
import time,pyautogui,os
import os
from openpyxl import Workbook
import time
from openpyxl import load_workbook
import os,pyautogui,time
from PIL import Image
import time,pyautogui,os
from openpyxl.styles import PatternFill
import openpyxl
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
#XML Extraction----------------------------------------------------------------------------------------------------------------------------------------
from xml.etree import ElementTree
import xml.etree.cElementTree as ET
from openpyxl import workbook
import os,pyautogui,time
from PIL import Image
import time,pyautogui,os
import os
from openpyxl import Workbook
import time
from openpyxl import load_workbook
import os,pyautogui,time
import time,pyautogui,os
from openpyxl.styles import PatternFill
import openpyxl
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import shutil
import string
import tabula
import PyPDF2
from tabula import read_pdf
import tabula
import pandas as pd
pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 1000)
from openpyxl.utils import get_column_letter
import shutil
from tkinter.filedialog import askdirectory
from collections import Counter

#Text an cell Style-------------------------------------------------------------------------------------------------------------------
fontgreen = Font(name='Calibri',
size=11,
bold=True,
italic=False,
vertAlign=None,
underline='none',
strike=False,
color='00003300')

fontbold = Font(name='Calibri',
size=11,
bold=True,
italic=False,
vertAlign=None,
underline='none',
strike=False,
color='000000')

fontboldbig = Font(name='Calibri',
size=16,
bold=True,
italic=False,
vertAlign=None,
underline='none',
strike=False,
color='000000')

my_green = openpyxl.styles.colors.Color(rgb='a8f0bc')
my_fillgreen = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_green)

my_darkgreen = openpyxl.styles.colors.Color(rgb='00003300')
my_filldarkgreen = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_darkgreen)



fontwhite = Font(name='Calibri',
size=11,
bold=True,
italic=False,
vertAlign=None,
underline='none',
strike=False,
color='ffffff')

my_darkgray = openpyxl.styles.colors.Color(rgb='333333')
my_filldarkgray = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_darkgray)

my_lightgray = openpyxl.styles.colors.Color(rgb='C5C5C5')
my_filllightgray = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_lightgray)

fontred = Font(name='Calibri',
size=11,
bold=True,
italic=False,
vertAlign=None,
underline='none',
strike=False,
color='82000b')


fontorange = Font(name='Calibri',
size=11,
bold=True,
italic=False,
vertAlign=None,
underline='none',
strike=False,
color='543101')

fontbolnormal = Font(name='Calibri',
size=11,
bold=True,
italic=False,
vertAlign=None,
underline='none',
strike=False,
color='000000')
    

my_red = openpyxl.styles.colors.Color(rgb='f29da4')
my_fillred = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)

my_darkred = openpyxl.styles.colors.Color(rgb='82000b')
my_filldarkred = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_darkred)

my_orange = openpyxl.styles.colors.Color(rgb='ffac38')
my_fillorange = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_orange)

my_yellow = openpyxl.styles.colors.Color(rgb='fff3c2')
my_fillyellow = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_yellow)

alignment = Alignment(horizontal='center',
                      vertical='center',
                      text_rotation=0,
                      wrap_text=False,
                      shrink_to_fit=True,
                      indent=0)

cell_border = Border(top = Side(border_style='thin', color='FF000000'),    
                            right = Side(border_style='thin', color='FF000000'), 
                            bottom = Side(border_style='thin', color='FF000000'),
                            left = Side(border_style='thin', color='FF000000'))


def Button_XML_Circuitry():
    pyautogui.alert(text='Select Tesla XML File', title='Select File', button='OK')
    directorio1 = filedialog.askopenfilename()
    directorio=os.path.split(directorio1)[0]
    names=os.path.split(directorio1)[1]
    os.chdir(directorio)

    my_pinkf = openpyxl.styles.colors.Color(rgb='FF2975')
    my_fill6 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_pinkf)

    font7 = Font(name='Calibri',
    size=11,
    bold=True,
    italic=True,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FFFFFF')


    #Parser del archivo XML--------------------------------------------------------------------------------------------------------------------------
    tree = ET.parse(names)
    root = tree.getroot()

    lista=[]
    for child in root.findall("designmgr"):
        lista.append(child)
        '''print(child.tag, child.attrib)'''
        
    designmgr=lista[0]
    lista2=[]
    for child in designmgr.findall("harnessdesign"):
        lista2.append(child)

    harnessdesign=lista2[0]
    atributoharnessdesign=str(harnessdesign.attrib.get('name'))
    #Creacion de hoja excel--------------------------------------------------------------------------------------------------------------------------
    book = Workbook()
    book.save("Circuitry extraction "+atributoharnessdesign+".xlsx")
    book.create_sheet('Circuitry '+ atributoharnessdesign)
    HC=book['Circuitry '+ atributoharnessdesign]
    labelrow=[" ","Name","Color","Spec","Mat","Option","Length","Multicore","From","From Cav","Plat","Term PN","Term PN S","Seal PN","Seal PN S","To","To Cav","Plat","Term PN","Term PN S", "Seal PN", "Seal PN S"]
    fila=1
    longitudlabelrow=len(labelrow)
    startolabelrow=0
    columnalabelrow=1
    while(startolabelrow<longitudlabelrow):
        HC.cell(row=fila, column=columnalabelrow).value=str(labelrow[startolabelrow])
        HC.cell(row=fila, column=columnalabelrow).fill=my_fill6
        HC.cell(row=fila, column=columnalabelrow).font = font7
        startolabelrow=startolabelrow+1
        columnalabelrow=columnalabelrow+1
    std=book.get_sheet_by_name('Sheet')
    book.remove_sheet(std)
    book.save("Circuitry extraction "+atributoharnessdesign+".xlsx")
    fila=fila+1

    lista3=[]
    for child in harnessdesign.findall("harnessdiagram"):
        lista3.append(child)

    harnessdiagram=lista3[0]
    lista4=[]
    for child in harnessdiagram.findall("harnessdiagramcontent"):
        lista4.append(child)
        
    harnessdiagramcontent=lista4[0]

    lista5=[]
    for child in harnessdiagramcontent.findall("tablegroup"):
        lista5.append(child)

    longitud=len(lista5)
    starto=0

    while(starto<longitud):
        tablegroup=lista5[starto]
        lista6=[]
        for child in tablegroup.findall("columnstyle"):
            lista6.append(child)
        starto=starto+1
        testo=lista6[1]
        atributo=str(testo.attrib.get('columnname'))
        if(atributo=="WIRE_NAME"):
            lista7=[]
            for child in tablegroup.findall("tablefamily"):
                lista7.append(child)
            tablefamily=lista7[0]
            lista8=[]
            for child in tablefamily.findall("table"):
                lista8.append(child)
            table=lista8
            longitud2=len(table)
            startotable=0
            while(startotable<longitud2):
                tablecurrent=table[startotable]
                lista9=[]
                for child in tablecurrent.findall("tabledatacache"):
                    lista9.append(child)
                startotable=startotable+1
                tabledatacache=lista9[0]
                lista10=[]
                for child in tabledatacache.findall("datavalues"):
                    lista10.append(child)
                datavalues=lista10[0]
                lista11=[]
                for child in datavalues.findall("datarow"):
                    lista11.append(child)
                longituddatavalues=len(lista11)
                startodatavalues=0
                while(startodatavalues<longituddatavalues):
                    datarow=lista11[startodatavalues]
                    lista12=[]
                    for child in datarow.findall("cellval"):
                        lista12.append(child)
                    startodatavalues=startodatavalues+1
                    longitudstartodatavalues=len(lista12)
                    startocellval=0
                    columna=1
                    while(startocellval<longitudstartodatavalues):
                        lista13=[]
                        cval=lista12[startocellval]
                        for child in cval.findall("cval"):
                            lista13.append(child)
                        testo2=lista13[0]
                        atributo2=str(testo2.attrib.get('val'))
                        HC.cell(row=fila, column=columna).value=str(atributo2)
                        HC.cell(row=fila, column=columna).border = Border(top = Side(border_style='thin', color='FF000000'),    
                        right = Side(border_style='thin', color='FF000000'), 
                        bottom = Side(border_style='thin', color='FF000000'),
                        left = Side(border_style='thin', color='FF000000'))
                        
                        columna=columna+1
                        startocellval=startocellval+1
                    fila=fila+1
    variable=0                
    letras=list(string.ascii_uppercase)
    for col in HC.columns:
        print("columnas")
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(cell.value) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 5) * 1.2
        HC.column_dimensions[str(letras[variable])].width = adjusted_width
        variable=variable+1
    
    book.save("Circuitry extraction "+atributoharnessdesign+".xlsx")
    pyautogui.alert(text='Extraction Report generated', title='Completed', button='OK')

def Button_XML_BOM():
    pyautogui.alert(text='Select Tesla XML File', title='Select File', button='OK')
    directorio1 = filedialog.askopenfilename()
    directorio=os.path.split(directorio1)[0]
    #names=os.path.split(directorio1)[1]
    os.chdir(directorio)

    my_pinkf = openpyxl.styles.colors.Color(rgb='FF2975')
    my_fill6 = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_pinkf)

    
    my_grayf = openpyxl.styles.colors.Color(rgb='808080')
    my_fillgray = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_grayf)

    font7 = Font(name='Calibri',
    size=11,
    bold=True,
    italic=True,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FFFFFF')


    lista_archivos=[]
    for filename in os.listdir(directorio):
        print(filename)
        lista_archivos.append(filename)

    path=(directorio+"//BOMS")
    print(directorio)
    print(path)
    os.makedirs(path)

    for archivo in lista_archivos:
        
        tree = ET.parse(archivo)
        root = tree.getroot()

        lista=[]
        for child in root.findall("designmgr"):
            lista.append(child)
            '''print(child.tag, child.attrib)'''
            
        designmgr=lista[0]
        lista2=[]
        for child in designmgr.findall("harnessdesign"):
            print(child)
            lista2.append(child)
        
        #numero_derivativos=len(lista2)-1

        harnessdesign=lista2[0]
        atributoharnessdesign=str(harnessdesign.attrib.get('name'))
        #creacion de hojas de excel----------------------------------------------------------------------------------------------
        book = Workbook()
        book.save("BOM extraction "+atributoharnessdesign+ ".xlsx")
        book.create_sheet('BOM '+atributoharnessdesign)
        book.create_sheet('Detail')
        HC=book['BOM '+atributoharnessdesign]
        HC2=book['Detail']
        labelrow=["Index","Part Number","Description","Supplier Part Number", "Supplier Name", "Quantity/Length","Color Code", "Color Description", "Type Description", "Group Name"]
        fila=1
        longitudlabelrow=len(labelrow)
        startolabelrow=0
        columnalabelrow=1
        while(startolabelrow<longitudlabelrow):
            HC.cell(row=fila, column=columnalabelrow).value=str(labelrow[startolabelrow])
            HC.cell(row=fila, column=columnalabelrow).fill=my_fill6
            HC.cell(row=fila, column=columnalabelrow).font = font7
            startolabelrow=startolabelrow+1
            columnalabelrow=columnalabelrow+1
        std=book.get_sheet_by_name('Sheet')
        book.remove_sheet(std)
        book.save("BOM extraction "+atributoharnessdesign+ ".xlsx")
        fila=fila+1

        lista3=[]
        for child in harnessdesign.findall("harnessdiagram"):
            lista3.append(child)

        harnessdiagram=lista3[0]
        lista4=[]
        for child in harnessdiagram.findall("harnessdiagramcontent"):
            lista4.append(child)
            
        harnessdiagramcontent=lista4[0]

        lista5=[]
        for child in harnessdiagramcontent.findall("tablegroup"):
            lista5.append(child)

        lista_mini_bom=[]
        #Encontrar numero de derivativos--------------------------------------------------------------------------------------------------------------------

        longitud=len(lista5)
        starto=0

        while(starto<longitud):
            tablegroup=lista5[starto]
            lista6=[]
            for child in tablegroup.findall("columnstyle"):
                lista6.append(child)
            starto=starto+1
            testo=lista6[1]
            atributo=str(testo.attrib.get('columnname'))
            if(atributo=="BOM_ID"):
                lista7=[]
                for child in tablegroup.findall("tablefamily"):
                    lista7.append(child)
                tablefamily=lista7[0]
                lista8=[]
                for child in tablefamily.findall("table"):
                    lista8.append(child)
                table=lista8
                longitud2=len(table)
                startotable=0
                while(startotable<longitud2):
                    tablecurrent=table[startotable]
                    lista9=[]
                    for child in tablecurrent.findall("tabledatacache"):
                        lista9.append(child)
                    startotable=startotable+1
                    tabledatacache=lista9[0]
                    lista10=[]
                    for child in tabledatacache.findall("datavalues"):
                        lista10.append(child)
                    datavalues=lista10[0]
                    lista11=[]
                    for child in datavalues.findall("datarow"):
                        lista11.append(child)
                    longituddatavalues=len(lista11)
                    startodatavalues=0
                    while(startodatavalues<longituddatavalues):
                        datarow=lista11[startodatavalues]
                        lista12=[]
                        for child in datarow.findall("cellval"):
                            lista12.append(child)
                        startodatavalues=startodatavalues+1
                        longitudstartodatavalues=len(lista12)
                        startocellval=0
                        columna=1
                        while(startocellval<longitudstartodatavalues):
                            lista13=[]
                            cval=lista12[startocellval]
                            for child in cval.findall("cval"):
                                lista13.append(child)
                            testo2=lista13[0]
                            atributo2=str(testo2.attrib.get('val'))
                            HC.cell(row=fila, column=columna).value=str(atributo2)
                            HC.cell(row=fila, column=columna).border = Border(top = Side(border_style='thin', color='FF000000'),    
                            right = Side(border_style='thin', color='FF000000'), 
                            bottom = Side(border_style='thin', color='FF000000'),
                            left = Side(border_style='thin', color='FF000000'))
                            #Escritura de columnas en Detail-------------------------------------------------------------------------------
                            if(columna==2):
                                HC2.cell(row=fila+9, column=1).value=str(atributo2)
                                HC2.cell(row=fila+9, column=1).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                right = Side(border_style='thin', color='FF000000'), 
                                bottom = Side(border_style='thin', color='FF000000'),
                                left = Side(border_style='thin', color='FF000000'))
                            if(columna==3):
                                HC2.cell(row=fila+9, column=6).value=str(atributo2)
                                HC2.cell(row=fila+9, column=6).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                right = Side(border_style='thin', color='FF000000'), 
                                bottom = Side(border_style='thin', color='FF000000'),
                                left = Side(border_style='thin', color='FF000000'))
                            if(columna==4):
                                HC2.cell(row=fila+9, column=5).value=str(atributo2)
                                HC2.cell(row=fila+9, column=5).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                right = Side(border_style='thin', color='FF000000'), 
                                bottom = Side(border_style='thin', color='FF000000'),
                                left = Side(border_style='thin', color='FF000000'))
                            if(columna==6):
                                lista_mini_bom.append(str(atributo2))
                                
                            if(columna==9):
                                HC2.cell(row=fila+9, column=2).value=str(atributo2)
                                HC2.cell(row=fila+9, column=2).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                right = Side(border_style='thin', color='FF000000'), 
                                bottom = Side(border_style='thin', color='FF000000'),
                                left = Side(border_style='thin', color='FF000000'))
                            if(columna==10):
                                HC2.cell(row=fila+9, column=3).value=str(atributo2)
                                HC2.cell(row=fila+9, column=3).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                right = Side(border_style='thin', color='FF000000'), 
                                bottom = Side(border_style='thin', color='FF000000'),
                                left = Side(border_style='thin', color='FF000000'))
                                if(str(atributo2)=="Wire" or str(atributo2)=="Tube" or str(atributo2)=="Tape"):
                                    HC2.cell(row=fila+9, column=7).value="Per Length"
                                    HC2.cell(row=fila+9, column=7).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                    right = Side(border_style='thin', color='FF000000'), 
                                    bottom = Side(border_style='thin', color='FF000000'),
                                    left = Side(border_style='thin', color='FF000000'))
                                else:
                                    HC2.cell(row=fila+9, column=7).value="Each"
                                    HC2.cell(row=fila+9, column=7).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                    right = Side(border_style='thin', color='FF000000'), 
                                    bottom = Side(border_style='thin', color='FF000000'),
                                    left = Side(border_style='thin', color='FF000000'))
                                #COLUMNA DE MAT CODE--------------------------------------------------------------------------------------------------------------    
                                HC2.cell(row=fila+9, column=4).value="-"
                                HC2.cell(row=fila+9, column=4).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                right = Side(border_style='thin', color='FF000000'), 
                                bottom = Side(border_style='thin', color='FF000000'),
                                left = Side(border_style='thin', color='FF000000'))



                            columna=columna+1
                            startocellval=startocellval+1
                        fila=fila+1
                        
        variable=0                
        letras=list(string.ascii_uppercase)   

        for col in HC.columns:
            print("columnas")
            max_length = 0
            column = col[0].column # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 5) * 1.2
            HC.column_dimensions[str(letras[variable])].width = adjusted_width
            variable=variable+1

        #Llenado de hoja "Detail"-------------------------------------------------------------------------------------------------------------------------------
        #column 1
        HC2.cell(row=1, column=1).value="Charted BOM Report"
        HC2.cell(row=3, column=1).value="Family"
        HC2.cell(row=4, column=1).value="Internal Number"
        HC2.cell(row=5, column=1).value="Customer Number"
        HC2.cell(row=6, column=1).value="Model Range"
        HC2.cell(row=7, column=1).value="Customer Name"    

        #column 2
        HC2.cell(row=4, column=2).value=atributoharnessdesign
        HC2.cell(row=5, column=2).value=atributoharnessdesign
        HC2.cell(row=7, column=2).value="TESLA"  

        #column 3
        HC2.cell(row=1, column=3).value="Report Data Type"
        HC2.cell(row=4, column=3).value="Internal Issue"
        HC2.cell(row=5, column=3).value="Customer Issue"

        #column 4
        HC2.cell(row=1, column=4).value="Customer"

        #Row 9
        values=["CPN","Type","Group","Mat Code","SPN","Part Description","UOM"]
        columns_x=1
        for value in values:
            HC2.cell(row=9, column=columns_x).value=value
            HC2.cell(row=9, column=columns_x).border = Border(top = Side(border_style='thin', color='FF000000'),    
                            right = Side(border_style='thin', color='FF000000'), 
                            bottom = Side(border_style='thin', color='FF000000'),
                            left = Side(border_style='thin', color='FF000000'))
            HC2.cell(row=9, column=columns_x).fill=my_fillgray
            HC2.cell(row=9, column=columns_x).font = font7
            columns_x=columns_x+1

        columna3=8
        dcounter=1
        
        HC2.cell(row=9, column=columna3).value=str(atributoharnessdesign)
        HC2.cell(row=9, column=columna3).border = Border(top = Side(border_style='thin', color='FF000000'),    
                        right = Side(border_style='thin', color='FF000000'), 
                        bottom = Side(border_style='thin', color='FF000000'),
                        left = Side(border_style='thin', color='FF000000'))
        HC2.cell(row=9, column=columna3).fill=my_fillgray
        HC2.cell(row=9, column=columna3).font = font7

        HC2.cell(row=10, column=columna3).value="Qty"
        HC2.cell(row=10, column=columna3).border = Border(top = Side(border_style='thin', color='FF000000'),    
                        right = Side(border_style='thin', color='FF000000'), 
                        bottom = Side(border_style='thin', color='FF000000'),
                        left = Side(border_style='thin', color='FF000000'))
        HC2.cell(row=10, column=columna3).fill=my_fillgray
        HC2.cell(row=10, column=columna3).font = font7

        row_inicial=11
        for cantidad in lista_mini_bom:
            HC2.cell(row=row_inicial, column=columna3).value=float(cantidad)
            HC2.cell(row=row_inicial, column=columna3).border = Border(top = Side(border_style='thin', color='FF000000'),    
                        right = Side(border_style='thin', color='FF000000'), 
                        bottom = Side(border_style='thin', color='FF000000'),
                        left = Side(border_style='thin', color='FF000000'))
            HC2.cell(row=row_inicial, column=columna3+1).value=cantidad
            HC2.cell(row=row_inicial, column=columna3+1).border = Border(top = Side(border_style='thin', color='FF000000'),    
                        right = Side(border_style='thin', color='FF000000'), 
                        bottom = Side(border_style='thin', color='FF000000'),
                        left = Side(border_style='thin', color='FF000000'))
            row_inicial=row_inicial+1
            
        columna3=columna3+1
        dcounter=dcounter+1

        HC2.cell(row=9, column=columna3).value="TOTAL"
        HC2.cell(row=9, column=columna3).border = Border(top = Side(border_style='thin', color='FF000000'),    
                        right = Side(border_style='thin', color='FF000000'), 
                        bottom = Side(border_style='thin', color='FF000000'),
                        left = Side(border_style='thin', color='FF000000'))
        HC2.cell(row=9, column=columna3).fill=my_fillgray
        HC2.cell(row=9, column=columna3).font = font7

        #por mientras
        rowsitita=11
        for cantidad in lista_mini_bom:
            HC2.cell(row=rowsitita, column=columna3).value=float(cantidad)
            HC2.cell(row=rowsitita, column=columna3).border = Border(top = Side(border_style='thin', color='FF000000'),    
                    right = Side(border_style='thin', color='FF000000'), 
                    bottom = Side(border_style='thin', color='FF000000'),
                    left = Side(border_style='thin', color='FF000000'))
            rowsitita=rowsitita+1
        
            
            
        #derivativos=len(lista111)
        variable=0               
        letras=list(string.ascii_uppercase)
        for col in HC2.columns:
            print("columnas")
            max_length = 0
            column = col[0].column # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 5) * 1
            HC2.column_dimensions[str(letras[variable])].width = adjusted_width
            variable=variable+1
                            

        book.save("BOM extraction "+atributoharnessdesign+ ".xlsx")
        try:
            shutil.move(str(directorio)+"/"+"BOM extraction "+atributoharnessdesign+".xlsx",str(path))
        except:
            pass
        os.chdir(directorio)
    pyautogui.alert(text='Extraction Report generated', title='Completed', button='OK')
    
def Button_PDF_Multiple():
    paso=0
    bar['value'] = paso
    print(paso)
    root.update_idletasks() 
    my_grayf = openpyxl.styles.colors.Color(rgb='00000080')
    my_fillgray = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_grayf)
    
    font7 = Font(name='Calibri',
    size=8,
    bold=True,
    italic=True,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FFFFFF')
    #number_format = 'Accounting'
    
    pyautogui.alert(text='Select PDF Files', title='Select File', button='OK')
    directorio1 = filedialog.askopenfilename()
    directorio=os.path.split(directorio1)[0]
    #names=os.path.split(directorio1)[1]
    os.chdir(directorio)
    
    #print("ARCHIVOS!!!!!!!!!!!")
    lista_archivos=[]
    for nombrearchivo in os.listdir(directorio):
        #print(nombrearchivo)
        lista_archivos.append(nombrearchivo)
    
    #Creacion de excel---------------------------------------------------------------------
    #Creacion de hoja excel--------------------------------------------------------------------------------------------------------------------------
    book = Workbook()
    book.save("PO_Summary.xlsx")
    book.create_sheet('PO Summary')
    std=book.get_sheet_by_name('Sheet')
    book.remove_sheet(std)
    HC=book['PO Summary']
    labelrow=["File","Series","Module","Module","Family", "Cost","Cost","Cu Content (Kg)", "Cu Rate", "Current Cu Weight", "Cu Portion","Total Price","","", "Price w/o Cu","LTA","Price w/o Cu & LTA", "Cu Content", "Cu Portion", "Total Price"]
    lista_columnas=[4,8,9,10,2]
    columna=2
    fila=3
    for label in labelrow:
        HC.cell(row=fila, column=columna).value=str(label)
        if(len(label)>1):
            HC.cell(row=fila, column=columna).fill=my_fillgray
        HC.cell(row=fila, column=columna).font = font7
        columna=columna+1
    book.save("PO_Summary.xlsx")
    
    modulos=[]
    totales=[]
    contents=[]
    rates=[]
    filenames=[]
    
    print("PASO10")
    paso=10
    print(paso)
    bar['value'] = paso
    root.update_idletasks()
    print("LEN LISTA ARCHIVOS")
    
    valor=70/len(lista_archivos)
    for archivo in lista_archivos:
          
        file=str(archivo)
        paso25=file.find('Doc-')
        paso26=file.find('.pdf')
        file_name=file[paso25+4:paso26]
        
        #dfs = tabula.read_pdf(file, pages=5)
        #print(dfs)
        
        # creating a pdf file object 
        pdfFileObj = open(file, 'rb') 
          
        # creating a pdf reader object 
        pdfReader = PyPDF2.PdfFileReader(pdfFileObj) 
          
        # printing number of pages in pdf file 
        paginas=list(range(pdfReader.numPages ))
        print("PAGINAS!!!!!!")
        print(len(paginas))
        stringsito=""
          
        counter=0
        #creating a page objecT
        print("VALOR2!!!!!")
        valor2=(valor*0.5)/len(paginas)
        print(valor2)
        
        for pagina in paginas:
            pageObj = pdfReader.getPage(counter)
            #print(counter)
            #print(pageObj.extractText())
    
            str1=pageObj.extractText()
            stringsito=stringsito+str1
            counter=counter+1
            
            paso=paso+valor2
            bar['value'] = paso
            root.update_idletasks()
            print("PASO!!!!!!!!!!")
            print(paso)

        

        # closing the pdf file object 
        pdfFileObj.close()
        
        #Aumentar barra de progreso

        
        #test="qqwewItem becomes:ddwqffeItem becomes:fweqdqf"
        #print(test.split('Item becomes:'))
        separado=stringsito.split('.00\nA')
        separado=separado[1:]
        counter=0
        
        valor3=(valor*0.5)/len(separado)
        for item in separado:
            #paso1=item.find('.00')
            paso2=item.find('----Z')
            modulo="A"+str(item[:paso2])
            paso3=item.find('per1Piece')
            #item2=item[paso3:]
            paso4=item.find('USD')
            total=item[paso3+9:paso4]
            #total = total[:-4]
            total = total.replace(",", ".")
            modulos.append(str(modulo))
            totales.append(float(total[1:]))
            
            
            #sacar contents
            paso5=item.find('USDfor')
            if(paso5!=-1):
                content=item[paso5:]
                paso6=content.find('KG')
                if(paso6!=-1):
                    content=item[(paso5+6):(paso5+paso6)]
                    contents.append(float(str(content)))
            else:
                contents.append("")
            
            
            #sacar rates
            paso7=item.find('perT')
            if(paso7!=-1):
                rate=item[paso7:]
                paso8=rate.find('USDfor')
                if(paso6!=-1):
                    rate=item[paso7+4:(paso7+paso8)]
                    rate=(float(str(rate)))/1000
                    rates.append(rate)
            else:
                rates.append("")
            
            filenames.append(str(file_name))
            
        #Aumentar barra de progreso
            paso=paso+valor3
            bar['value'] = paso
            root.update_idletasks()
    #         print(paso)
    # print("LISTAS!!!!!!!!!")
    # print(len(modulos))
    # print(len(totales))
    # print(len(contents))
    # print(len(rates))
    # print(len(filenames))


    lista_de_listas=[modulos,totales,contents,rates,filenames]
        
    contador=0
    for lista in lista_de_listas:
        fila=4
        for elemento in lista:
            HC.cell(row=fila, column=int(lista_columnas[contador])).value=elemento
            if(int(lista_columnas[contador])==4):
                HC.cell(row=fila, column=int(lista_columnas[contador])+1).value=elemento
            
            fila=fila+1
        contador=contador+1
    
    book.save("PO_Summary.xlsx")
    
    #print("brake!!!!!!!!!!!")
    #print("PASO")
    valor=20/len(modulos)
    fila=4
    for elemento in modulos:
        HC.cell(row=fila, column=12).value=str("=ROUND((I"+str(fila)+"*"+"K"+str(fila)+"),2)")
        HC.cell(row=fila, column=13).value=str("=ROUND((L"+str(fila)+"+"+"H"+str(fila)+"),2)")
        HC.cell(row=fila, column=17).value=str("=ROUND((P"+str(fila)+"*$S$1),2)")
        HC.cell(row=fila, column=18).value=str("=ROUND((P"+str(fila)+"-(P"+str(fila)+"*$S$1)),2)")
        HC.cell(row=fila, column=20).value=str("=ROUND(S"+str(fila)+"*$K$"+str(fila)+",2)")
        HC.cell(row=fila, column=21).value=str("=ROUND(((P"+str(fila)+"-Q"+str(fila)+")+T"+str(fila)+"),2)")
        HC.cell(row=fila, column=22).value=str("=IF(U"+str(fila)+"=0,"",M"+str(fila)+"-U"+str(fila)+")")
        HC.cell(row=fila, column=23).value=str("=IF(U"+str(fila)+"=0,"",R"+str(fila)+"=H"+str(fila)+")")
        #HC.cell(row=fila, column=23).value=str("=IF(U4=0,'\"\"',R4=H4)")
        HC.cell(row=fila, column=24).value=str("=IF(U"+str(fila)+"=0,"",S"+str(fila)+"=I"+str(fila)+")")
        HC.cell(row=fila, column=25).value=str("=IF(U"+str(fila)+"=0,"",T"+str(fila)+"=L"+str(fila)+")")
        #HC.cell(row=fila, column=21).number_format="Currency
        fila=fila+1
        #Aumentar barra de progreso
        paso=paso+valor
        bar['value'] = paso
        root.update_idletasks()
        print(paso)
        
    variable=0                
    letras=list(string.ascii_uppercase)   

    for col in HC.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1
        HC.column_dimensions[str(letras[variable])].width = adjusted_width
        variable=variable+1
        
        
    pyautogui.alert(text='Extraction Report generated', title='Completed', button='OK')
    book.save("PO_Summary.xlsx")
    
def Button_PDF_onefile():
    bar['value'] = 0
    my_grayf = openpyxl.styles.colors.Color(rgb='00000080')
    my_fillgray = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_grayf)
    
    font7 = Font(name='Calibri',
    size=8,
    bold=True,
    italic=True,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FFFFFF')


    pyautogui.alert(text='Select PDF File', title='Select File', button='OK')
    directorio1 = filedialog.askopenfilename()
    directorio=os.path.split(directorio1)[0]
    names=os.path.split(directorio1)[1]
    os.chdir(directorio)
    file=str(names)
    paso25=file.find('Doc-')
    paso26=file.find('.pdf')
    file_name=file[paso25+4:paso26]
    #print(file_name)
    
    #dfs = tabula.read_pdf(file, pages=5)
    #print(dfs)
    
    # creating a pdf file object 
    pdfFileObj = open(file, 'rb') 
      
    # creating a pdf reader object 
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj) 
      
    # printing number of pages in pdf file 
    paginas=list(range(pdfReader.numPages ))
    #print(len(paginas))
    stringsito=""
      
    counter=0
    #creating a page object
    paso=10
    bar['value'] = paso
    valor=30/len(paginas)
    root.update_idletasks() 
    for pagina in paginas:
        pageObj = pdfReader.getPage(counter)
        #print(counter)
        #print(pageObj.extractText())

        str1=pageObj.extractText()
        #print(str1)
        #time.sleep(5)
        stringsito=stringsito+str1
        counter=counter+1
        
        paso=paso+valor
        bar['value'] = paso
        root.update_idletasks()
    
        
    # closing the pdf file object 
    pdfFileObj.close()
    #test="qqwewItem becomes:ddwqffeItem becomes:fweqdqf"
    #print(test.split('Item becomes:'))
    separado=stringsito.split('.00\nA')
    #for item in separado:
        #print("ITEM ITEM ITEM")
        #print(item)
        #time.sleep(5)
    separado=separado[1:]
    counter=0
    modulos=[]
    totales=[]
    contents=[]
    rates=[]
    

    

    valor=30/len(separado)
    paso=paso+valor
    for item in separado:

        #paso1=item.find('.00')
        paso2=item.find('----Z')
        modulo="A"+str(item[:paso2])
        #print(modulo)
        #time.sleep(1)
        paso3=item.find('per1Piece')
        #item2=item[paso3:]
        paso4=item.find('USD')
        total=item[paso3+9:paso4]
        #total = total[:-4]
        total = total.replace(",", ".")
        modulos.append(str(modulo))
        totales.append(float(total[1:]))
        
        #print(modulos)
        #print(totales)
        
        #sacar contents
        paso5=item.find('USDfor')
        if(paso5!=-1):
            content=item[paso5:]
            paso6=content.find('KG')
            if(paso6!=-1):
                content=item[(paso5+6):(paso5+paso6)]
                
                #print("CONTENT!!!!!!!!!")
                #print(len(modulos))
                #print(content)
                #print(float(str(content)))
                contents.append(float(str(content)))
                
                
        else:
            contents.append("")
        counter=counter+1
        
        #print(contents)
        
        #sacar rates
        paso7=item.find('perT')
        if(paso7!=-1):
            rate=item[paso7:]
            paso8=rate.find('USDfor')
            if(paso6!=-1):
                rate=item[paso7+4:(paso7+paso8)]
                rate=(float(str(rate)))/1000
                rates.append(rate)
        else:
            rates.append("")
        paso=paso+valor
        bar['value'] = paso
        root.update_idletasks()
        #print(rates)
        
        
        #print(paso5)
        #print(item)
        #print('---------------------------Separado-----------------------------------------')
        #print(rate)
        #time.sleep(0.5)
    # print("PASO!!!!!!!!")
    # print(paso)
    lista_de_listas=[modulos,totales,contents,rates]
    #Creacion de hoja excel--------------------------------------------------------------------------------------------------------------------------
    book = Workbook()
    book.save("PO_Summary_"+file_name+".xlsx")
    book.create_sheet('PO Summary')
    std=book.get_sheet_by_name('Sheet')
    book.remove_sheet(std)
    HC=book['PO Summary']
    labelrow=["File","Series","Module","Module","Family", "Cost","Cost","Cu Content (Kg)", "Cu Rate", "Current Cu Weight", "Cu Portion","Total Price","","", "Price w/o Cu","LTA","Price w/o Cu & LTA", "Cu Content", "Cu Portion", "Total Price"]
    lista_columnas=[4,8,9,10]
    columna=2
    fila=3
    for label in labelrow:
        HC.cell(row=fila, column=columna).value=str(label)
        if(len(label)>1):
            HC.cell(row=fila, column=columna).fill=my_fillgray
        HC.cell(row=fila, column=columna).font = font7
        columna=columna+1
        
    contador=0
    for lista in lista_de_listas:
        fila=4
        for elemento in lista:
            HC.cell(row=fila, column=int(lista_columnas[contador])).value=elemento
            if(int(lista_columnas[contador])==4):
                HC.cell(row=fila, column=int(lista_columnas[contador])+1).value=elemento
            
            fila=fila+1
        contador=contador+1
        
    fila=4
    step=30/len(modulos)
    print(step)
    for elemento in modulos:
        HC.cell(row=fila, column=2).value=str(file_name)
        HC.cell(row=fila, column=12).value=str("=ROUND((I"+str(fila)+"*"+"K"+str(fila)+"),2)")
        HC.cell(row=fila, column=13).value=str("=ROUND((L"+str(fila)+"+"+"H"+str(fila)+"),2)")
        HC.cell(row=fila, column=17).value=str("=ROUND((P"+str(fila)+"*$S$1),2)")
        HC.cell(row=fila, column=18).value=str("=ROUND((P"+str(fila)+"-(P"+str(fila)+"*$S$1)),2)")
        HC.cell(row=fila, column=20).value=str("=ROUND(S"+str(fila)+"*$K$"+str(fila)+",2)")
        HC.cell(row=fila, column=21).value=str("=ROUND(((P"+str(fila)+"-Q"+str(fila)+")+T"+str(fila)+"),2)")
        HC.cell(row=fila, column=22).value=str("=IF(U"+str(fila)+"=0,"",M"+str(fila)+"U"+str(fila)+")")
        HC.cell(row=fila, column=23).value=str("=IF(U"+str(fila)+"=0,"",R"+str(fila)+"=H"+str(fila)+")")
        #HC.cell(row=fila, column=23).value=str("=IF(U4=0,'\"\"',R4=H4)")
        HC.cell(row=fila, column=24).value=str("=IF(U"+str(fila)+"=0,"",S"+str(fila)+"=I"+str(fila)+")")
        HC.cell(row=fila, column=25).value=str("=IF(U"+str(fila)+"=0,"",T"+str(fila)+"=L"+str(fila)+")")
        #HC.cell(row=fila, column=21).number_format="Currency
        fila=fila+1
        print(fila)
        paso=paso+step
        bar['value'] = paso
        print(step)
        root.update_idletasks() 

        
    variable=0                
    letras=list(string.ascii_uppercase)   

    for col in HC.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1
        HC.column_dimensions[str(letras[variable])].width = adjusted_width
        variable=variable+1
    pyautogui.alert(text='Extraction Report generated', title='Completed', button='OK')
        
        
    
    book.save("PO_Summary_"+file_name+".xlsx")
    
def Button_Comparer():
    
    paso=0
    bar['value'] = paso
    
    paso=0
    print(paso)
    bar['value'] = paso
    root.update_idletasks()
    
    #Text an cell Style-------------------------------------------------------------------------------------------------------------------
    fontgreen = Font(name='Calibri',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='00003300')
    
    my_green = openpyxl.styles.colors.Color(rgb='a8f0bc')
    my_fillgreen = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_green)
    
    fontwhite = Font(name='Calibri',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='ffffff')
    
    my_darkgray = openpyxl.styles.colors.Color(rgb='333333')
    my_filldarkgray = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_darkgray)
    
    fontred = Font(name='Calibri',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='82000b')
    
    fontorange = Font(name='Calibri',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='543101')
    
    my_red = openpyxl.styles.colors.Color(rgb='f29da4')
    my_fillred = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
    
    my_orange = openpyxl.styles.colors.Color(rgb='ffac38')
    my_fillorange = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_orange)
    
    my_yellow = openpyxl.styles.colors.Color(rgb='fff3c2')
    my_fillyellow = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_yellow)
    
    
    pyautogui.alert(text='Select Old BOM', title='Select File', button='OK')
    directorio1 = filedialog.askopenfilename()
    directorio=os.path.split(directorio1)[0]
    names=os.path.split(directorio1)[1]
    os.chdir(directorio)
    
    pyautogui.alert(text='Select new BOM', title='Select File', button='OK')
    directorio1 = filedialog.askopenfilename()
    directorio=os.path.split(directorio1)[0]
    names2=os.path.split(directorio1)[1]

    alignment = Alignment(horizontal='center',
                          vertical='bottom',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=True,
                          indent=0)
    
    paso=10
    print(paso)
    bar['value'] = paso
    root.update_idletasks()
    
    #Leer old BOM y generar diccionario----------------------------------------------------------------------------------------------
    start_time = time.time()
    book = openpyxl.load_workbook(names, data_only=True)
    HL=book["Charted_BOM"]
    print("--- %s seconds ---" % (time.time() - start_time))
    pn_list_old=[]
    UOM_old=[]

    value=0
    rowinicial=10
    while(value!=None):
        cell_obj = HL.cell(row = rowinicial, column = 2)
        value=cell_obj.value
        pn_list_old.append(value)
        rowinicial=rowinicial+1
        
    value=0
    rowinicial=10
    while(value!=None):
        cell_obj = HL.cell(row = rowinicial, column = 11)
        value=cell_obj.value
        UOM_old.append(value)
        rowinicial=rowinicial+1
    
        
    pn_list_old=pn_list_old[:-1]
    UOM_old=UOM_old[:-1]
    
    conditional="A"
    columnainicial=13
    modules_old=[]
    lista_diccionarios=[]
    
    lista_diccionarios2=[]
    
    while(conditional=="A"):
        cell_obj = HL.cell(row = 9, column = columnainicial)
        value=str(cell_obj.value)
        row=10
        lista_pn=[]
        lista_quants=[]
        lista_uom=[]
        minidict={}
        minidict2={}
        modules_old.append(value)
        
        counter=0
        
        for item in pn_list_old:
            cell_obj2 = HL.cell(row = row, column = columnainicial)
            value2=cell_obj2.value
            if(value2!=0):
                lista_quants.append(float(value2))
                lista_pn.append(str(item))
                lista_uom.append(UOM_old[counter])
                
            row=row+1
            counter=counter+1

        conditional=value[0]
        columnainicial=columnainicial+1
        minidict = dict(zip(lista_pn, lista_quants))

        minidict2=dict(zip(lista_pn, zip(lista_quants, lista_uom)))

        

        
        lista_diccionarios.append(minidict)

        

        lista_diccionarios2.append(minidict2)



    modules_old=modules_old[:-1]
    lista_diccionarios=lista_diccionarios[:-1]
    lista_diccionarios2=lista_diccionarios2[:-1]
    


    olddict = dict(zip(modules_old, lista_diccionarios))
    olddict2= dict(zip(modules_old, lista_diccionarios2))
    
    print(len(olddict))
    print(len(olddict2))
    
    paso=35
    print(paso)
    bar['value'] = paso
    root.update_idletasks()
    
    
    
    #Leer new BOM y generar diccionario--------------------------------------------------------------------
    print("fin PRIMER PROCESO")
    print("--- %s seconds ---" % (time.time() - start_time))

    book = openpyxl.load_workbook(names2, data_only=True)
    HL=book["Charted_BOM"]
    
    pn_list_new=[]
    UOM_new=[]

    
    value=0
    rowinicial=10
    while(value!=None):
        cell_obj = HL.cell(row = rowinicial, column = 2)
        value=cell_obj.value
        pn_list_new.append(value)
        rowinicial=rowinicial+1
        
    value=0
    rowinicial=10
    while(value!=None):
        cell_obj = HL.cell(row = rowinicial, column = 11)
        value=cell_obj.value
        UOM_new.append(value)
        rowinicial=rowinicial+1
        
    pn_list_new=pn_list_new[:-1]
    UOM_new=UOM_new[:-1]
        
    conditional="A"
    columnainicial=13
    modules_new=[]
    lista_diccionarios=[]
    lista_diccionarios2=[]
    while(conditional=="A"):
        cell_obj = HL.cell(row = 9, column = columnainicial)
        value=str(cell_obj.value)
        row=10
        lista_pn=[]
        lista_quants=[]
        lista_uom=[]
        minidict={}
        modules_new.append(value)
        counter=0
        for item in pn_list_new:
            cell_obj2 = HL.cell(row = row, column = columnainicial)
            value2=cell_obj2.value
            if(value2!=0):
                lista_quants.append(float(value2))
                lista_pn.append(str(item))
                lista_uom.append(UOM_new[counter])
            row=row+1
            counter=counter+1
        conditional=value[0]
        columnainicial=columnainicial+1
        minidict = dict(zip(lista_pn, lista_quants))
        minidict2=dict(zip(lista_pn, zip(lista_quants, lista_uom)))

        lista_diccionarios.append(minidict)
        lista_diccionarios2.append(minidict2)
        
    modules_new=modules_new[:-1]
    lista_diccionarios=lista_diccionarios[:-1]
    lista_diccionarios2=lista_diccionarios2[:-1]
    newdict = dict(zip(modules_new, lista_diccionarios))
    newdict2= dict(zip(modules_new, lista_diccionarios2))
    
    print(len(newdict))
    print(len(newdict2))
    
    paso=60
    print(paso)
    bar['value'] = paso
    root.update_idletasks()
    
    
    added_modules=[]
    eliminated_modules=[]
    nc_modules=[]
    
    
    for modulo in modules_old:
        if(modulo in modules_new):
            nc_modules.append(modulo)
        else:
            eliminated_modules.append(modulo)
            
    for modulo in modules_new:
        if modulo not in modules_old:
            added_modules.append(modulo)
    
            
    book = Workbook()
    book.save("Summary_comparison.xlsx")
    book.create_sheet('Comparison')
    std=book.get_sheet_by_name('Sheet')
    book.remove_sheet(std)
    HC=book['Comparison']
    book.save("Summary_comparison.xlsx")
    
    rowinicial=1
    columnainicial=1
    
    paso=65
    print(paso)
    bar['value'] = paso
    root.update_idletasks()
    
    print("ADDED MODULES STARTO----------------------------------------------------------------")
    HC.cell(row=rowinicial, column=columnainicial).value=str("ADDEED MODULES")
    HC.merge_cells(start_row=rowinicial, start_column=columnainicial, end_row=rowinicial, end_column=columnainicial+18)
    HC.cell(row=rowinicial, column=columnainicial).font = fontwhite
    HC.cell(row=rowinicial, column=columnainicial).fill=my_filldarkgray
    HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
    
    rowinicial=3

    maxrow=[]
    rowmasgrande=0
    switchadded=1
    for modulo in added_modules:
        switchadded=2
        if(len(maxrow)==5):
            maxrow=[]
            rowinicial=rowmasgrande+1
            columnainicial=1
            
        HC.merge_cells(start_row=rowinicial, start_column=columnainicial, end_row=rowinicial, end_column=columnainicial+2)
        HC.cell(row=rowinicial, column=columnainicial).value=str(modulo)
        HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
        HC.cell(row=rowinicial, column=columnainicial).font = fontgreen
        HC.cell(row=rowinicial, column=columnainicial).fill=my_fillgreen
        rowinicial=rowinicial+1
        
        HC.cell(row=rowinicial, column=columnainicial).value=str("Part Numbers")
        HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
        HC.cell(row=rowinicial, column=columnainicial).font = fontwhite
        HC.cell(row=rowinicial, column=columnainicial).fill=my_filldarkgray
        columnainicial=columnainicial+1
        HC.cell(row=rowinicial, column=columnainicial).value=str("Quants")
        HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
        HC.cell(row=rowinicial, column=columnainicial).font = fontwhite
        HC.cell(row=rowinicial, column=columnainicial).fill=my_filldarkgray
        #columnainicial=columnainicial+1
        HC.cell(row=rowinicial, column=columnainicial+1).value=str("UOM")
        HC.cell(row=rowinicial, column=columnainicial+1).alignment = alignment
        HC.cell(row=rowinicial, column=columnainicial+1).font = fontwhite
        HC.cell(row=rowinicial, column=columnainicial+1).fill=my_filldarkgray
        
        rowinicial=rowinicial+1
        pnyquants=newdict2.get(modulo)
        
 
        print("PONER ATENCION PNYQUANTS!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        print(pnyquants)
        print(len(pnyquants))
        for key in pnyquants:
            columnainicial=columnainicial-1
            HC.cell(row=rowinicial, column=columnainicial).value=str(key)

            columnainicial=columnainicial+1
            valorsito=pnyquants.get(key)
            HC.cell(row=rowinicial, column=columnainicial).value=str(valorsito[0])
            HC.cell(row=rowinicial, column=columnainicial+1).value=str(valorsito[1])
            rowinicial=rowinicial+1
            
        finalrow=rowinicial
        maxrow.append(int(rowinicial))

        rowinicial=rowinicial-len(pnyquants)-2
        
        columnainicial=columnainicial+3
    
        rowmasgrande=max(maxrow)
    
    if switchadded == 2:
        rowinicial=finalrow
        rowinicial=rowinicial+1
    
    columnainicial=1
    book.save("Summary_comparison.xlsx")
    switcheliminated=0
    print("ELIMINATED MODULES!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
    HC.cell(row=rowinicial, column=columnainicial).value=str("ELIMINATED MODULES")
    HC.merge_cells(start_row=rowinicial, start_column=columnainicial, end_row=rowinicial, end_column=columnainicial+18)
    HC.cell(row=rowinicial, column=columnainicial).font = fontwhite
    HC.cell(row=rowinicial, column=columnainicial).fill=my_filldarkgray
    HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
    print(rowinicial)
    print(eliminated_modules)
    print(len(eliminated_modules))
    
    rowinicial=rowinicial+2
    maxrow=[]
    
    for modulo in eliminated_modules:
        switcheliminated=1
        if(len(maxrow)==5):
            maxrow=[]
            rowinicial=rowmasgrande+1
            columnainicial=1
            
        HC.merge_cells(start_row=rowinicial, start_column=columnainicial, end_row=rowinicial, end_column=columnainicial+2)
        HC.cell(row=rowinicial, column=columnainicial).value=str(modulo)
        HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
        HC.cell(row=rowinicial, column=columnainicial).font = fontred
        HC.cell(row=rowinicial, column=columnainicial).fill=my_fillred
        rowinicial=rowinicial+1
        
        HC.cell(row=rowinicial, column=columnainicial).value=str("Part Numbers")
        HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
        HC.cell(row=rowinicial, column=columnainicial).font = fontwhite
        HC.cell(row=rowinicial, column=columnainicial).fill=my_filldarkgray
        columnainicial=columnainicial+1
        HC.cell(row=rowinicial, column=columnainicial).value=str("Quants")
        HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
        HC.cell(row=rowinicial, column=columnainicial).font = fontwhite
        HC.cell(row=rowinicial, column=columnainicial).fill=my_filldarkgray
        
        HC.cell(row=rowinicial, column=columnainicial+1).value=str("UOM")
        HC.cell(row=rowinicial, column=columnainicial+1).alignment = alignment
        HC.cell(row=rowinicial, column=columnainicial+1).font = fontwhite
        HC.cell(row=rowinicial, column=columnainicial+1).fill=my_filldarkgray
        
        rowinicial=rowinicial+1
        pnyquants=olddict2.get(modulo)
        
        for key in pnyquants:
            columnainicial=columnainicial-1
            HC.cell(row=rowinicial, column=columnainicial).value=str(key)

            columnainicial=columnainicial+1
            valorsito=pnyquants.get(key)
            HC.cell(row=rowinicial, column=columnainicial).value=float(valorsito[0])
            HC.cell(row=rowinicial, column=columnainicial+1).value=str(valorsito[1])
            rowinicial=rowinicial+1
        
        maxrow.append(int(rowinicial))
        rowinicial=rowinicial-len(pnyquants)-2
        columnainicial=columnainicial+3
        rowmasgrande=max(maxrow)
        
    book.save("Summary_comparison.xlsx")

    if(switcheliminated==1):
        rowinicial=rowmasgrande+1
    
    columnainicial=1
    
    paso=90
    print(paso)
    bar['value'] = paso
    root.update_idletasks()

    HC.cell(row=rowinicial, column=columnainicial).value=str("CHANGED MODULES")
    HC.merge_cells(start_row=rowinicial, start_column=columnainicial, end_row=rowinicial, end_column=2)
    HC.cell(row=rowinicial, column=columnainicial).font = fontwhite
    HC.cell(row=rowinicial, column=columnainicial).fill=my_filldarkgray
    HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
    
    HC.cell(row=rowinicial, column=4).value=str("SUMMARY")
    HC.merge_cells(start_row=rowinicial, start_column=4, end_row=rowinicial, end_column=19)
    HC.cell(row=rowinicial, column=4).font = fontwhite
    HC.cell(row=rowinicial, column=4).fill=my_filldarkgray
    HC.cell(row=rowinicial, column=4).alignment = alignment
    
    summaryrow=rowinicial+2
    
    rowinicial=rowinicial+2
    
    maxrow=[]
    changed_modules=[]
    for modulo in nc_modules:
        
        oldvalue=olddict.get(modulo)
        newvalue=newdict.get(modulo)

        changes_component=[]
        changes_quant=[]
        add=[]
        remove=[]
        for diff in list(dictdiffer.diff(oldvalue, newvalue)):
            if(len(diff)>0):


                
                if (diff[0]=="change"):
                    changes_component.append(diff[1])
                    changes_quant.append(list(diff[2]))
                    
                if (diff[0]=="add"):
                    add.append(diff[2])
                    
                if (diff[0]=="remove"):
                    remove.append(diff[2])

        if(len(add)>0 or len(remove)>0 or len(changes_component)>0 or len(changes_quant)>0):
   
            HC.merge_cells(start_row=rowinicial, start_column=columnainicial, end_row=rowinicial, end_column=columnainicial+1)
            HC.cell(row=rowinicial, column=columnainicial).value=modulo
            HC.cell(row=rowinicial, column=columnainicial).value=str(modulo)
            changed_modules.append(str(modulo))
            HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
            HC.cell(row=rowinicial, column=columnainicial).font = fontorange
            HC.cell(row=rowinicial, column=columnainicial).fill=my_fillorange
            
            rowinicial=rowinicial+1
            
            HC.cell(row=rowinicial, column=columnainicial).value=str("Part Numbers")
            HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
            HC.cell(row=rowinicial, column=columnainicial).font = fontwhite
            HC.cell(row=rowinicial, column=columnainicial).fill=my_filldarkgray
            columnainicial=columnainicial+1
            HC.cell(row=rowinicial, column=columnainicial).value=str("Quants")
            HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
            HC.cell(row=rowinicial, column=columnainicial).font = fontwhite
            HC.cell(row=rowinicial, column=columnainicial).fill=my_filldarkgray
            
 
            columnainicial=columnainicial-1
            rowinicial=rowinicial+1

            
            if(len(add)>0):
                HC.merge_cells(start_row=rowinicial, start_column=columnainicial, end_row=rowinicial, end_column=columnainicial+1)
                HC.cell(row=rowinicial, column=columnainicial).value=modulo
                HC.cell(row=rowinicial, column=columnainicial).value="ADDED"
                HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
                HC.cell(row=rowinicial, column=columnainicial).font = fontgreen
                HC.cell(row=rowinicial, column=columnainicial).fill=my_fillgreen
                

                
                rowinicial=rowinicial+1
                
                for element in add:

                    for cosa in element:


                        HC.cell(row=rowinicial, column=columnainicial).value=cosa[0]
                        columnainicial=columnainicial+1

                        HC.cell(row=rowinicial, column=columnainicial).value=cosa[1]
                        columnainicial=columnainicial-1
                        rowinicial=rowinicial+1
                
            if(len(remove)>0):
                HC.merge_cells(start_row=rowinicial, start_column=columnainicial, end_row=rowinicial, end_column=columnainicial+1)
                HC.cell(row=rowinicial, column=columnainicial).value=modulo
                HC.cell(row=rowinicial, column=columnainicial).value="REMOVED"
                HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
                HC.cell(row=rowinicial, column=columnainicial).font = fontred
                HC.cell(row=rowinicial, column=columnainicial).fill=my_fillred
                
                rowinicial=rowinicial+1
                
                for element in remove:

                    for cosa in element:


                        HC.cell(row=rowinicial, column=columnainicial).value=cosa[0]
                        columnainicial=columnainicial+1
                        HC.cell(row=rowinicial, column=columnainicial).value=cosa[1]
                        columnainicial=columnainicial-1
                        rowinicial=rowinicial+1
            
                
            if(len(changes_component)>0):

                HC.merge_cells(start_row=rowinicial, start_column=columnainicial, end_row=rowinicial, end_column=columnainicial+1)
                HC.cell(row=rowinicial, column=columnainicial).value=modulo
                HC.cell(row=rowinicial, column=columnainicial).value="CHANGED"
                HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
                HC.cell(row=rowinicial, column=columnainicial).font = fontorange
                HC.cell(row=rowinicial, column=columnainicial).fill=my_fillyellow
                
                rowinicial=rowinicial+1
                
                rowinicialtemp=rowinicial

                for element in changes_component:
                    if (type(element)==list):
                        HC.cell(row=rowinicial, column=columnainicial).value=element[0]
                    else:
                       HC.cell(row=rowinicial, column=columnainicial).value=element 
                    rowinicial=rowinicial+1

                    
                for element in changes_quant:

                    HC.cell(row=rowinicialtemp, column=columnainicial+1).value=str(element[0])+"/"+str(element[1])
                    if element[0]>element[1]:
                        HC.cell(row=rowinicialtemp, column=columnainicial+1).font = fontred
                        HC.cell(row=rowinicialtemp, column=columnainicial+1).fill=my_fillred
                    if element[0]<element[1]:
                        HC.cell(row=rowinicialtemp, column=columnainicial+1).font = fontgreen
                        HC.cell(row=rowinicialtemp, column=columnainicial+1).fill=my_fillgreen
                    rowinicialtemp=rowinicialtemp+1
                        
                        
            rowinicial=rowinicial+1

            
            maxrow.append(int(rowinicial))
            book.save("Summary_comparison.xlsx")
    
    #add summary
    
    
    HC.cell(row=summaryrow, column=4).value="ADDED MODULES"
    HC.cell(row=summaryrow, column=4).alignment = alignment
    HC.cell(row=summaryrow, column=4).font = fontwhite
    HC.cell(row=summaryrow, column=4).fill=my_filldarkgray
            
    HC.cell(row=summaryrow, column=5).value=len(added_modules)
    
    iterrow=summaryrow+2
    for modulo in added_modules:
        HC.cell(row=iterrow, column=4).value=modulo
        iterrow=iterrow+1
        
    
    HC.cell(row=summaryrow, column=7).value="REMOVED MODULES"
    HC.cell(row=summaryrow, column=7).alignment = alignment
    HC.cell(row=summaryrow, column=7).font = fontwhite
    HC.cell(row=summaryrow, column=7).fill=my_filldarkgray
    
    HC.cell(row=summaryrow, column=8).value=len(eliminated_modules)
    
    iterrow2=summaryrow+2
    for modulo in eliminated_modules:
        HC.cell(row=iterrow2, column=7).value=modulo
        iterrow2=iterrow2+1
    
    HC.cell(row=summaryrow, column=10).value="CHANGED MODULES"
    HC.cell(row=summaryrow, column=10).alignment = alignment
    HC.cell(row=summaryrow, column=10).font = fontwhite
    HC.cell(row=summaryrow, column=10).fill=my_filldarkgray
    
    HC.cell(row=summaryrow, column=11).value=len(changed_modules)
    
    iterrow3=summaryrow+2
    for modulo in changed_modules:
        HC.cell(row=iterrow3, column=10).value=modulo
        iterrow3=iterrow3+1
    
 
   

    book.save("Summary_comparison.xlsx")

            
    print("TERMINAO")
    
    paso=100
    print(paso)
    bar['value'] = paso
    root.update_idletasks()
    
    pyautogui.alert(text='Proccess Completed', title='', button='OK')
    
def Button_Comparer_V2():
    
    paso=0
    bar['value'] = paso
    
    paso=0
    print(paso)
    bar['value'] = paso
    root.update_idletasks()
    
    #Text an cell Style-------------------------------------------------------------------------------------------------------------------
    fontgreen = Font(name='Calibri',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='00003300')
    
    fontbold = Font(name='Calibri',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='000000')
    
    fontboldbig = Font(name='Calibri',
    size=16,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='000000')
    
    fontbolnormal = Font(name='Calibri',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='000000')
    
    my_green = openpyxl.styles.colors.Color(rgb='a8f0bc')
    my_fillgreen = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_green)
    
    my_darkgreen = openpyxl.styles.colors.Color(rgb='00003300')
    my_filldarkgreen = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_darkgreen)
    
    fontwhite = Font(name='Calibri',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='ffffff')
    
    my_darkgray = openpyxl.styles.colors.Color(rgb='333333')
    my_filldarkgray = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_darkgray)
    
    fontred = Font(name='Calibri',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='82000b')
    

    fontorange = Font(name='Calibri',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='543101')
    
    my_red = openpyxl.styles.colors.Color(rgb='f29da4')
    my_fillred = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
    
    my_darkred = openpyxl.styles.colors.Color(rgb='82000b')
    my_filldarkred = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_darkred)
    
    my_orange = openpyxl.styles.colors.Color(rgb='ffac38')
    my_fillorange = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_orange)
    
    my_yellow = openpyxl.styles.colors.Color(rgb='fff3c2')
    my_fillyellow = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_yellow)
       
    pyautogui.alert(text='Select Old BOM', title='Select File', button='OK')
    directorio1 = filedialog.askopenfilename()
    directorio=os.path.split(directorio1)[0]
    names=os.path.split(directorio1)[1]
    os.chdir(directorio)
    
    pyautogui.alert(text='Select new BOM', title='Select File', button='OK')
    directorio1 = filedialog.askopenfilename()
    directorio=os.path.split(directorio1)[0]
    names2=os.path.split(directorio1)[1]

    alignment = Alignment(horizontal='center',
                          vertical='center',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=True,
                          indent=0)
    
    cell_border = Border(top = Side(border_style='thin', color='FF000000'),    
                                right = Side(border_style='thin', color='FF000000'), 
                                bottom = Side(border_style='thin', color='FF000000'),
                                left = Side(border_style='thin', color='FF000000'))
    
    def black_title_cell(value):
        HC.cell(row=rowinicial, column=columnainicial).value=str(value)
        HC.cell(row=rowinicial, column=columnainicial).font = fontwhite
        HC.cell(row=rowinicial, column=columnainicial).fill=my_filldarkgray
        HC.cell(row=rowinicial, column=columnainicial).alignment = alignment
    
    paso=10
    print(paso)
    bar['value'] = paso
    root.update_idletasks()
    
    #Leer old BOM y generar diccionario----------------------------------------------------------------------------------------------
    start_time = time.time()
    book = openpyxl.load_workbook(names, data_only=True)
    HL=book["Charted_BOM"]
    print("--- %s seconds ---" % (time.time() - start_time))
        
    #pn_list_old=[]    
    EPN_old=[]
    IPN_old=[]
    Type_old=[]
    Group_old=[]
    Mat_code_old=[]
    CPN_old=[]
    CUW_old=[]
    part_description_old=[]
    Klima=[]
    UOM_old=[]
        
    #Leer columnas inicioales---------------------------------------------------------------------------
    #columns_to_read=[2,3,4,5,6,7,8,9,10,11]
    lists_to_fill=[EPN_old,IPN_old,Type_old,
                   Group_old,Mat_code_old,CPN_old,
                   CUW_old,part_description_old,
                   Klima,UOM_old]
       
    counter=2
    for lista in lists_to_fill:
        border="DUMMY"
        rowinicial=10
        while(border!="medium"):
            cell_obj = HL.cell(row = rowinicial, column = counter)
            value=str(cell_obj.value)
            border=str(cell_obj.border.bottom.style)
            print(border)
            lista.append(value)
            rowinicial=rowinicial+1
        counter+=1
        
    rowklima=rowinicial
    #time.sleep(20)
    
    #Read Modules-------------------------------------------------------------------------------------   
    
    conditional="A"
    columnainicial=13
    modules_old=[]
    
    lista_valores_number_wires_klima_and_nonklima=[]
    
    lista_diccionarios=[]
    
    lista_diccionarios2=[]
    
    lista_diccionarios3=[]
    
    lista_diccionarios_family=[]
    
    
    while(str(value)!="Price"):
        
        time.sleep(0.2)
        number_wires_klimaandnonklima=[]
        
        cell_obj = HL.cell(row = 9, column = columnainicial)
        

        cell_obj_wires_klima = HL.cell(row =rowklima+4, column = columnainicial)
        cell_obj_wires_nonklima = HL.cell(row =rowklima+5, column = columnainicial)
        
        cell_obj_family = HL.cell(row = 7, column = columnainicial)
        
        value=str(cell_obj.value)
        
        
        if(value[0]=="A"):
            row=10
            
            lista_pn=[]
            lista_quants=[]        
            lista_uom=[]        
            EPN_old_temp=[]
            IPN_old_temp=[]
            Type_old_temp=[]
            Group_old_temp=[]
            Mat_code_old_temp=[]
            CPN_old_temp=[]
            CUW_old_temp=[]
            part_description_old_temp=[]
            Klima_temp=[]
            UOM_old_temp=[]
            
            minidict={}
            minidict2={}
                  
            if(value!="None"):
                modules_old.append(value)
                number_wires_klimaandnonklima.append(float(cell_obj_wires_klima.value))
                number_wires_klimaandnonklima.append(float(cell_obj_wires_nonklima.value))
                
            if(cell_obj_family.value!=None):
                value_family=str(cell_obj_family.value)
                    
            lista_diccionarios_family.append(value_family)
                          
            counter=0
                    
            conditional=value[0]
                    
            for item in EPN_old:
                cell_obj2 = HL.cell(row = row, column = columnainicial)
                value2=cell_obj2.value
                if(value2!=0):
                    
                    lista_pn.append(str(item))
                    lista_uom.append(UOM_old[counter])
                    
                    EPN_old_temp.append(str(item))
                    lista_quants.append(float(value2))                
                    IPN_old_temp.append(IPN_old[counter])
                    Type_old_temp.append(Type_old[counter])
                    Group_old_temp.append(Group_old[counter])
                    Mat_code_old_temp.append(Mat_code_old[counter])
                    CPN_old_temp.append(CPN_old[counter])
                    CUW_old_temp.append(CUW_old[counter])
                    part_description_old_temp.append(part_description_old[counter])
                    Klima_temp.append(Klima[counter])
                    UOM_old_temp.append(UOM_old[counter])
                    
                    
                row=row+1
                counter=counter+1
            
            conditional=value[0]
    
            
            minidict = dict(zip(lista_pn, lista_quants))
    
            minidict2=dict(zip(lista_pn, zip(lista_quants, lista_uom)))
            
            
            
            minidict3=dict((z[0], list(z[1:])) for z in zip(EPN_old_temp, lista_quants, IPN_old_temp,
                                                            Type_old_temp,Group_old_temp,Mat_code_old_temp,
                                                            CPN_old_temp,CUW_old_temp,part_description_old_temp,
                                                            Klima_temp,UOM_old_temp))
            
            
            if(value!="None"):
                
                lista_diccionarios.append(minidict)
                lista_diccionarios2.append(minidict2)
                lista_diccionarios3.append(minidict3)
    
                lista_valores_number_wires_klima_and_nonklima.append(number_wires_klimaandnonklima)
            
        columnainicial=columnainicial+1

    olddict = dict(zip(modules_old, lista_diccionarios))
    olddict3=dict(zip(modules_old,lista_diccionarios3))
    olddictnumberofwires=dict(zip(modules_old,lista_valores_number_wires_klima_and_nonklima))
    olddictfamily=dict(zip(modules_old, lista_diccionarios_family))
    root.update_idletasks()
        

    lista_diccionarios_family=[]
    print("FAMILY DICT")
    
    print(len(olddict))
    print(len(olddictfamily))

    paso=35
    bar['value'] = paso
    root.update_idletasks()
    
    #Read new BOM and make dictionary--------------------------------------------------------------------
    print("fin PRIMER PROCESO")
    print("--- %s seconds ---" % (time.time() - start_time))

    book = openpyxl.load_workbook(names2, data_only=True)
    HL=book["Charted_BOM"]
    
    print("--- %s seconds ---" % (time.time() - start_time))
    
    #pn_list_old=[]    
    EPN_new=[]
    IPN_new=[]
    Type_new=[]
    Group_new=[]
    Mat_code_new=[]
    CPN_new=[]
    CUW_new=[]
    part_description_new=[]
    Klima_new=[]
    UOM_new=[]
    
    
    #Read columns---------------------------------------------------------------------------
    #columns_to_read=[2,3,4,5,6,7,8,9,10,11]
    lists_to_fill=[EPN_new,IPN_new,Type_new,
                   Group_new,Mat_code_new,CPN_new,
                   CUW_new,part_description_new,
                   Klima_new,UOM_new]
    
    counter=2
    for lista in lists_to_fill:
        border="DUMMY"
        rowinicial=10
        while(border!="medium"):
            cell_obj = HL.cell(row = rowinicial, column = counter)
            value=cell_obj.value
            border=str(cell_obj.border.bottom.style)
            lista.append(value)
            rowinicial=rowinicial+1
        counter+=1
    
    rowklima=rowinicial
    
     #Read Modules-------------------------------------------------------------------------------------   
    
    conditional="A"
    columnainicial=13
    modules_new=[]
    lista_diccionarios=[]
    
    lista_diccionarios2=[]
    
    lista_diccionarios3=[]
    
    lista_valores_number_wires_klima_and_nonklima_new=[]
    
    #lista_diccionarios_family_new=[]
    
    while(str(value)!="Price"):
        root.update_idletasks()
        
        number_wires_klimaandnonklima=[]
        
        cell_obj = HL.cell(row = 9, column = columnainicial)
        
        cell_obj_wires_klima = HL.cell(row =rowklima+4, column = columnainicial)
        cell_obj_wires_nonklima = HL.cell(row =rowklima+5, column = columnainicial)
        
        cell_obj_family = HL.cell(row = 7, column = columnainicial)
        
        value=str(cell_obj.value)
        
        if(str(value[0])=="A"):
            row=10
            
            lista_pn=[]
            lista_quants=[]
            lista_uom=[]
            EPN_new_temp=[]
            IPN_new_temp=[]
            Type_new_temp=[]
            Group_new_temp=[]
            Mat_code_new_temp=[]
            CPN_new_temp=[]
            CUW_new_temp=[]
            part_description_new_temp=[]
            Klima_new_temp=[]
            UOM_new_temp=[]
            
            minidict={}
            minidict2={}
            
            if(value!="None"):
                modules_new.append(value)
                number_wires_klimaandnonklima.append(float(cell_obj_wires_klima.value))
                number_wires_klimaandnonklima.append(float(cell_obj_wires_nonklima.value))
                
            if(cell_obj_family.value!=None):
                value_family=str(cell_obj_family.value)
            lista_diccionarios_family.append(value_family)
            counter=0
            conditional=value[0]
        
            for item in EPN_new:
                cell_obj2 = HL.cell(row = row, column = columnainicial)
                value2=cell_obj2.value
                if(value2!=0):
                    
                    lista_pn.append(str(item))
                    print(len(UOM_old))
                    print(counter)
                    #lista_uom.append(UOM_old[counter])
                    
                    EPN_new_temp.append(str(item))
                    lista_quants.append(float(value2))
                    
                    IPN_new_temp.append(IPN_new[counter])
                    Type_new_temp.append(Type_new[counter])
                    Group_new_temp.append(Group_new[counter])
                    Mat_code_new_temp.append(Mat_code_new[counter])
                    CPN_new_temp.append(CPN_new[counter])
                    CUW_new_temp.append(CUW_new[counter])
                    part_description_new_temp.append(part_description_new[counter])
                    Klima_new_temp.append(Klima_new[counter])
                    UOM_new_temp.append(UOM_new[counter])
                        
                row=row+1
                counter=counter+1
            

    
            
            minidict = dict(zip(lista_pn, lista_quants))
    
            minidict2=dict(zip(lista_pn, zip(lista_quants, lista_uom)))
            
            minidict3=dict((z[0], list(z[1:])) for z in zip(EPN_new_temp, lista_quants, IPN_new_temp,
                                                            Type_new_temp,Group_new_temp,Mat_code_new_temp,
                                                            CPN_new_temp,CUW_new_temp,part_description_new_temp,
                                                            Klima_new_temp,UOM_new_temp))
            if(value!="None"):
                
                lista_diccionarios.append(minidict)
                lista_diccionarios2.append(minidict2)
                lista_diccionarios3.append(minidict3)
                lista_valores_number_wires_klima_and_nonklima_new.append(number_wires_klimaandnonklima)
        columnainicial=columnainicial+1

                  
    newdict3=dict(zip(modules_new,lista_diccionarios3))
    olddictnumberofwiresnew=dict(zip(modules_new,lista_valores_number_wires_klima_and_nonklima_new))
    newdictfamily=dict(zip(modules_new, lista_diccionarios_family))
    root.update_idletasks()
    paso=60
    #print(paso)
    bar['value'] = paso
    root.update_idletasks()
    
    EPN_new
    
    #Numeros de parte agregados
    #pn_added=list(set(EPN_new) - set(EPN_old))
    
    #Numeros de parte eliminados
    pn_removed=list(set(EPN_old) - set(EPN_new))
    
    master_pn=EPN_new+pn_removed

    
    print("--- %s seconds ---" % (time.time() - start_time))
    print("fin segundo proceso")
    
    
    book = Workbook()
    book.save("Summary_comparison.xlsx")
    book.create_sheet('Comparison')
    std=book.get_sheet_by_name('Sheet')
    book.remove_sheet(std)
    HC=book['Comparison']
    book.save("Summary_comparison.xlsx")
    
    
    #First table titles----------------------------------------------------------------------------------------------------------------
    
    titles_list=["EPN","IPN","Type","Group","Mat Code","CPN","CU Weight","Part Description"
                 ,"Klima","UOM"]
    
    #Group table titles-------------------------------------------------------------------------------------------------------------------
    
    grouptable_list=["Assembly","Bolt","Busbar","Cavity Plug","Cavity Seal","Channel","Clip"
                     ,"Connector","Connector FH","Connector Name","Cover","Eyelet","Fuse","Fuse Box"
                     ,"Grommet","Heatshrink","Housing","Label","Mating","Multicore Wire","Nut"
                     ,"Other","Relay","Resistor","Solder","Spacer","Splice","Strap","Strap Holder","Tape"
                     ,"Terminal","Tube","Washer","Washerhose","Wire"]
    
    klimawiretablelist=["Take Rate","Number of wires KLIMA [pcs.]", "Number of wires non-KLIMA [pcs.]","Total Number of wires [pcs.]"]
    

    rowwrite=7
    columnwrite=1
    
    for title in titles_list:
        HC.cell(row=rowwrite, column=columnwrite).value=title
        HC.cell(row=rowwrite, column=columnwrite).font = fontwhite
        HC.cell(row=rowwrite, column=columnwrite).fill=my_filldarkgray
        HC.cell(row=rowwrite, column=columnwrite).alignment = alignment
        columnwrite+=1
        
    
    
    #Creacion de diccionario para creacion de tabla------------------------------------------------------------------------------------
    master_new_pn_dict=dict((z[0], list(z[1:])) for z in zip(EPN_new, IPN_new,
                                                        Type_new,Group_new,Mat_code_new,
                                                        CPN_new,CUW_new,part_description_new,
                                                        Klima_new,UOM_new))
    
    
    master_old_pn_dict=dict((z[0], list(z[1:])) for z in zip(EPN_old, IPN_old,
                                                        Type_old,Group_old,Mat_code_old,
                                                        CPN_old,CUW_old,part_description_old,
                                                        Klima,UOM_old))
    
    for key in master_old_pn_dict:
        print("key")
        print(key)
        print(master_old_pn_dict.get(key))
    
    #Write master table-----------------------------------------------------------------------------------------------------------------

    rowwrite+=1
    columnwrite=1
    
    for dato in EPN_new:
        columnwrite=1
        pnrow=master_new_pn_dict.get(str(dato))
        
        print(dato)
        

        HC.cell(row=rowwrite, column=columnwrite).value=dato
        HC.cell(row=rowwrite, column=columnwrite).border=cell_border
        columnwrite=2
        
        print(type(pnrow))
        if(str(type(pnrow))!="<class 'NoneType'>"):
            for elemento in pnrow:
                HC.cell(row=rowwrite, column=columnwrite).value=elemento
                HC.cell(row=rowwrite, column=columnwrite).border=cell_border
                columnwrite+=1
            rowwrite+=1
        
    for dato in pn_removed:
        columnwrite=1
        pnrow=master_old_pn_dict.get(str(dato))
        
        print("dato")
        (print(dato))
        print(pnrow)

        HC.cell(row=rowwrite, column=columnwrite).value=dato
        HC.cell(row=rowwrite, column=columnwrite).border=cell_border
        columnwrite=2
        for elemento in pnrow:
            HC.cell(row=rowwrite, column=columnwrite).value=elemento
            HC.cell(row=rowwrite, column=columnwrite).border=cell_border
            columnwrite+=1
        rowwrite+=1
        
    mastertable_endrow=rowwrite+9
    mastertable_endrow_wiretotals=rowwrite+3
    mastertable_endrow_numberofwire=rowwrite+4
    mastertable_groupcolumn=10
    
    #write group table titles-------------------------------------------------------------------------------------------------------------
    HC.cell(row=mastertable_endrow, column=mastertable_groupcolumn-1).value="Material Quantities"
    HC.cell(row=mastertable_endrow, column=mastertable_groupcolumn-1).alignment=Alignment(vertical='center')
    HC.cell(row=mastertable_endrow, column=mastertable_groupcolumn-1).alignment = Alignment(textRotation=255)
    HC.cell(row=rowwrite, column=columnwrite).font = fontbold
    HC.cell(row=rowwrite, column=columnwrite).border = cell_border
    HC.merge_cells(start_row=mastertable_endrow, start_column=mastertable_groupcolumn-1, 
                   end_row=mastertable_endrow+34, end_column=mastertable_groupcolumn-1)
    
    #Write wire sum list title table---------------------------------------------------------------------------------------------------------
    for dato in klimawiretablelist:
        HC.cell(row=mastertable_endrow_wiretotals, column=mastertable_groupcolumn).value=dato
        HC.cell(row=mastertable_endrow_wiretotals, column=mastertable_groupcolumn).border=cell_border
        mastertable_endrow_wiretotals+=1
    
    #Write group table-----------------------------------------------------------------------------------------------------------------------
    for group in grouptable_list:
        HC.cell(row=mastertable_endrow, column=mastertable_groupcolumn).value=group
        HC.cell(row=mastertable_endrow, column=mastertable_groupcolumn).border=cell_border
        mastertable_endrow+=1

    mastertable_endrow+=2
    
    #Write vertical titlers-----------------------------------------------------------------------------------------------------------------------
    HC.cell(row=mastertable_endrow, column=mastertable_groupcolumn-1).value="Non Klima Qty"
    HC.cell(row=mastertable_endrow, column=mastertable_groupcolumn-1).alignment = Alignment(textRotation=255)
    HC.cell(row=rowwrite, column=columnwrite).font = fontbold
    HC.merge_cells(start_row=mastertable_endrow, start_column=mastertable_groupcolumn-1, 
                   end_row=mastertable_endrow+34, end_column=mastertable_groupcolumn-1)
    HC.cell(row=mastertable_endrow, column=mastertable_groupcolumn-1).border=cell_border
    
    
    HC.cell(row=mastertable_endrow+36, column=mastertable_groupcolumn).value="NON-KLIMA Content"
    HC.cell(row=mastertable_endrow+36, column=mastertable_groupcolumn).font = fontboldbig
    HC.cell(row=mastertable_endrow+36, column=mastertable_groupcolumn).border=cell_border
    
    
    
    for group in grouptable_list:
        HC.cell(row=mastertable_endrow, column=mastertable_groupcolumn).value=group
        HC.cell(row=mastertable_endrow, column=mastertable_groupcolumn).border=cell_border
        mastertable_endrow+=1
    

    modules_added=list(set(modules_new) - set(modules_old))

    modules_removed=list(set(modules_old) - set(modules_new))
    
    #Write modules and delta columns------------------------------------------------------------------------------------------------------
    rowwrite=6
    columnwrite=12
    
    number_of_wires_dict={}
    
    for modulos in modules_removed:
        modules_new.append(modulos)
        
    print("Lista maestra")
    print(modules_new)

    
    for modulo in modules_new:
        
        listanonklimaswitch=[]
        
        numberofwires=olddictnumberofwires.get(modulo)
        numberofwiresnew=olddictnumberofwiresnew.get(modulo)
        
        if(numberofwires!=None):

            numberwireklima=numberofwires[0]
            numberwirenonklima=numberofwires[1]
            
            HC.cell(row=mastertable_endrow_numberofwire, column=columnwrite+1).value=numberofwires[0]
            HC.cell(row=mastertable_endrow_numberofwire, column=columnwrite+1).border=cell_border
            
            HC.cell(row=mastertable_endrow_numberofwire+1, column=columnwrite+1).value=numberofwires[1]
            HC.cell(row=mastertable_endrow_numberofwire+1, column=columnwrite+1).border=cell_border
            
            sumawires=float(numberofwires[0]+numberofwires[1])
            
            HC.cell(row=mastertable_endrow_numberofwire+2, column=columnwrite+1).value=sumawires
            HC.cell(row=mastertable_endrow_numberofwire+2, column=columnwrite+1).border=cell_border
            
            
        else:
            
            numberwireklima=0
            numberwirenonklima=0
            sumawires=0
            
            HC.cell(row=mastertable_endrow_numberofwire, column=columnwrite+1).value=0
            HC.cell(row=mastertable_endrow_numberofwire, column=columnwrite+1).border=cell_border
            
            HC.cell(row=mastertable_endrow_numberofwire+1, column=columnwrite+1).value=0
            HC.cell(row=mastertable_endrow_numberofwire+1, column=columnwrite+1).border=cell_border
            
            
            HC.cell(row=mastertable_endrow_numberofwire+2, column=columnwrite+1).value=0
            HC.cell(row=mastertable_endrow_numberofwire+2, column=columnwrite+1).border=cell_border
            
        if(numberofwiresnew!=None):
            
            numberwireklimanew=numberofwiresnew[0]
            numberwirenonklimanew=numberofwiresnew[1]
            
            HC.cell(row=mastertable_endrow_numberofwire, column=columnwrite).value=numberofwiresnew[0]
            HC.cell(row=mastertable_endrow_numberofwire, column=columnwrite).border=cell_border
            
            HC.cell(row=mastertable_endrow_numberofwire+1, column=columnwrite).value=numberofwiresnew[1]
            HC.cell(row=mastertable_endrow_numberofwire+1, column=columnwrite).border=cell_border
            
            sumawiresnew=float(numberofwiresnew[0]+numberofwiresnew[1])
            
            HC.cell(row=mastertable_endrow_numberofwire+2, column=columnwrite).value=sumawiresnew
            HC.cell(row=mastertable_endrow_numberofwire+2, column=columnwrite).border=cell_border
            
        else:
            #print("numberviejo")
            
            numberwireklimanew=0
            numberwirenonklimanew=0
            sumawiresnew=0
            
            HC.cell(row=mastertable_endrow_numberofwire, column=columnwrite).value=0
            HC.cell(row=mastertable_endrow_numberofwire, column=columnwrite).border=cell_border
            
            HC.cell(row=mastertable_endrow_numberofwire+1, column=columnwrite).value=0
            HC.cell(row=mastertable_endrow_numberofwire+1, column=columnwrite).border=cell_border
            
            HC.cell(row=mastertable_endrow_numberofwire+2, column=columnwrite).value=0
            HC.cell(row=mastertable_endrow_numberofwire+2, column=columnwrite).border=cell_border
            
        
        HC.cell(row=mastertable_endrow_numberofwire, column=columnwrite+2).value=numberwireklimanew-numberwireklima
        HC.cell(row=mastertable_endrow_numberofwire, column=columnwrite+2).border=cell_border
        
        HC.cell(row=mastertable_endrow_numberofwire+1, column=columnwrite+2).value=numberwirenonklimanew-numberwirenonklima
        HC.cell(row=mastertable_endrow_numberofwire+1, column=columnwrite+2).border=cell_border
        
        HC.cell(row=mastertable_endrow_numberofwire+2, column=columnwrite+2).value=sumawiresnew-sumawires
        HC.cell(row=mastertable_endrow_numberofwire+2, column=columnwrite+2).border=cell_border
        
        if(sumawiresnew-sumawires!=0):
        
            number_of_wires_minidict={modulo:[numberwireklimanew-numberwireklima,numberwirenonklimanew-numberwirenonklima
                                          ,sumawiresnew-sumawires]}

            number_of_wires_dict.update(number_of_wires_minidict)
        
        if((sumawiresnew-sumawires)>0):
        
            HC.cell(row=mastertable_endrow_numberofwire+2, column=columnwrite+2).font = fontwhite
            HC.cell(row=mastertable_endrow_numberofwire+2, column=columnwrite+2).fill=my_filldarkgreen
            
        if((sumawiresnew-sumawires<0)):
        
            HC.cell(row=mastertable_endrow_numberofwire+2, column=columnwrite+2).font = fontwhite
            HC.cell(row=mastertable_endrow_numberofwire+2, column=columnwrite+2).fill=my_filldarkred
    
        #lists of groups---------------------------------------------------------------------------------------------------------------------
        Assembly_nk=[]
        Bolt_nk=[]
        Busbar_nk=[]
        Cavity_Plug_nk=[]
        Cavity_Seal_nk=[]
        Channel_nk=[]
        Clip_nk=[]
        Connector_nk=[]
        Connector_FH_nk=[]
        Connector_Name_nk=[]
        Cover_nk=[]
        Eyelet_nk=[]
        Fuse_nk=[]
        Fuse_Box_nk=[]
        Grommet_nk=[]
        Heatshrink_nk=[]
        Housing_nk=[]
        Label_nk=[]
        Mating_nk=[]
        Multicore_Wire_nk=[]
        Nut_nk=[]
        Other_nk=[]
        Relay_nk=[]
        Resistor_nk=[]
        SOLDER_nk=[]
        Spacer_nk=[]
        Splice_nk=[]
        Strap_nk=[]
        Strap_Holder_nk=[]
        Tape_nk=[]
        Terminal_nk=[]
        Tube_nk=[]
        Washer_nk=[]
        Washerhose_nk=[]
        Wire_nk=[]
        
        Assembly_k=[]
        Bolt_k=[]
        Busbar_k=[]
        Cavity_Plug_k=[]
        Cavity_Seal_k=[]
        Channel_k=[]
        Clip_k=[]
        Connector_k=[]
        Connector_FH_k=[]
        Connector_Name_k=[]
        Cover_k=[]
        Eyelet_k=[]
        Fuse_k=[]
        Fuse_Box_k=[]
        Grommet_k=[]
        Heatshrink_k=[]
        Housing_k=[]
        Label_k=[]
        Mating_k=[]
        Multicore_Wire_k=[]
        Nut_k=[]
        Other_k=[]
        Relay_k=[]
        Resistor_k=[]
        SOLDER_k=[]
        Spacer_k=[]
        Splice_k=[]
        Strap_k=[]
        Strap_Holder_k=[]
        Tape_k=[]
        Terminal_k=[]
        Tube_k=[]
        Washer_k=[]
        Washerhose_k=[]
        Wire_k=[]
        
        dictionary_of_groups={"Assembly":[Assembly_nk,Assembly_k],"Bolt":[Bolt_nk,Bolt_k],"Busbar":[Busbar_nk,Busbar_k]
                              ,"Cavity Plug":[Cavity_Plug_nk,Cavity_Plug_k],"Cavity Seal":[Cavity_Seal_nk,Cavity_Seal_k],"Channel":[Channel_nk,Channel_k]
                              ,"Clip":[Clip_nk,Clip_k],"Connector":[Connector_nk,Connector_k],"Connector FH":[Connector_FH_nk,Connector_FH_k]
                              ,"Connector Name":[Connector_Name_nk,Connector_Name_k],"Cover":[Cover_nk,Cover_k],"Eyelet":[Eyelet_nk,Eyelet_k]
                              ,"Fuse":[Fuse_nk,Fuse_k],"Fuse Box":[Fuse_Box_nk,Fuse_Box_k],"Grommet":[Grommet_nk,Grommet_k]
                              ,"Heatshrink":[Heatshrink_nk,Heatshrink_k],"Housing":[Housing_nk,Housing_k],"Label":[Label_nk,Label_k]
                              ,"Mating":[Mating_nk,Mating_k],"Multicore Wire":[Multicore_Wire_nk,Multicore_Wire_k],"Nut":[Nut_nk,Nut_k]
                              ,"Other":[Other_nk,Other_k],"Relay":[Relay_nk,Relay_k],"Resistor":[Resistor_nk,Resistor_k]
                              ,"Solder":[SOLDER_nk,SOLDER_k],"Spacer":[Spacer_nk,Spacer_k],"Splice":[Splice_nk,Splice_k]
                              ,"Strap":[Strap_nk,Strap_k],"Strap Holder":[Strap_Holder_nk,Strap_Holder_k],"Tape":[Tape_nk,Tape_k]
                              ,"Terminal":[Terminal_nk,Terminal_k],"Tube":[Tube_nk,Tube_k],"Washer":[Washer_nk,Washer_k]
                              ,"Washerhose":[Washerhose_nk,Washerhose_k],"Wire":[Wire_nk,Wire_k]}
        
    #lists of groups---------------------------------------------------------------------------------------------------------------------
        
        HC.cell(row=rowwrite-1, column=columnwrite).value=newdictfamily.get(modulo)
        HC.cell(row=rowwrite-1, column=columnwrite).font = fontwhite
        HC.cell(row=rowwrite-1, column=columnwrite).fill=my_filldarkgray
        HC.cell(row=rowwrite-1, column=columnwrite).alignment = alignment
        
        HC.cell(row=rowwrite, column=columnwrite).value=modulo
        HC.cell(row=rowwrite, column=columnwrite).font = fontwhite
        HC.cell(row=rowwrite, column=columnwrite).fill=my_filldarkgray
        HC.cell(row=rowwrite, column=columnwrite).alignment = alignment
        
        rowwrite+=1
        archivedrow=rowwrite+1
        
        HC.cell(row=rowwrite, column=columnwrite).value=modulo
        HC.cell(row=rowwrite, column=columnwrite).font = fontwhite
        HC.cell(row=rowwrite, column=columnwrite).fill=my_filldarkgray
        HC.cell(row=rowwrite, column=columnwrite).alignment = alignment
        
        rowwrite+=1
        
        matrix=newdict3.get(modulo)
        matrix2=olddict3.get(modulo)
        
        print("Analisisssssssssssssssssssssssssssssssssssssssssssssssssssssssssssss")
        print(matrix)
        print(matrix2)
        
        for element in master_pn:
            
            if(matrix!=None):
                check=matrix.get(element)
                print("Check and element-----------------")
                print(element)
                print(check)
                
            else:
                check="-"
            
            if(matrix2!=None):
                check2=matrix2.get(element)
                print("Check and element-----------------")
                print(element)
                print(check)
                
            else:
                check2="-"
            
            print("new value---------------------------------")
            #-new-value-new-value-new-value-new-value-new-value-new-value-new-value-new-value-new-value-new-value-new-value-new-value
            if(type(check)==list):
                print(check)
                HC.cell(row=rowwrite, column=columnwrite).value=check[0]
                HC.cell(row=rowwrite, column=columnwrite).border=cell_border
                value1=check[0]
                                
                if(str(check[8])=="KLIMA"):
                    dictionary_of_groups.get(str(check[3]))[1].append([float(check[0]),"New"])

                if(str(check[8])=="NON-KLIMA"):
                    dictionary_of_groups.get(str(check[3]))[0].append([float(check[0]),"New"])
        
            else:
                HC.cell(row=rowwrite, column=columnwrite).value=0
                HC.cell(row=rowwrite, column=columnwrite).border=cell_border
                value1=0
            
            columnwrite+=1
            
            #-old-value-old-value-old-value-old-value-old-value-old-value-old-value-old-value-old-value-old-value-old-value-old-value-old-value-          
            if(type(check2)==list):
                HC.cell(row=rowwrite, column=columnwrite).value=check2[0]
                HC.cell(row=rowwrite, column=columnwrite).border=cell_border
                value2=check2[0]
                
                if(str(check2[8])=="KLIMA"):
                    print("oli")
                    print(str(check2[3]))
                    print(dictionary_of_groups.get(str(check2[3])))
                    dictionary_of_groups.get(str(check2[3]))[1].append([float(check2[0]),"old"])
                if(str(check2[8])=="NON-KLIMA"):
                    dictionary_of_groups.get(str(check2[3]))[0].append([float(check2[0]),"old"])
                
            elif(type(check2)==str):
                HC.cell(row=rowwrite, column=columnwrite).value=check2
                HC.cell(row=rowwrite, column=columnwrite).border=cell_border
                value2=0
            else:
                HC.cell(row=rowwrite, column=columnwrite).value=0
                HC.cell(row=rowwrite, column=columnwrite).border=cell_border
                value2=0
                
            columnwrite+=1
            #-DELTA-value--DELTA-value-DELTA-value-DELTA-value-DELTA-value-DELTA-value-DELTA-value-DELTA-value-DELTA-value-DELTA-value-DELTA-value
            delta=value1-value2
            
            if(delta>0):
                HC.cell(row=rowwrite, column=columnwrite).value=value1-value2
                HC.cell(row=rowwrite, column=columnwrite).font = fontgreen
                HC.cell(row=rowwrite, column=columnwrite).fill=my_fillgreen
                HC.cell(row=rowwrite, column=columnwrite).border=cell_border
            elif(delta<0):
                HC.cell(row=rowwrite, column=columnwrite).value=value1-value2
                HC.cell(row=rowwrite, column=columnwrite).font = fontred
                HC.cell(row=rowwrite, column=columnwrite).fill=my_fillred
                HC.cell(row=rowwrite, column=columnwrite).border=cell_border
            else:
                HC.cell(row=rowwrite, column=columnwrite).value=value1-value2
                HC.cell(row=rowwrite, column=columnwrite).border=cell_border
            
            rowwrite+=1
            
            columnwrite-=2
            
            endprocessrow=rowwrite+9
            
        rowwrite=archivedrow-1
        columnwrite+=1
        
        HC.cell(row=rowwrite, column=columnwrite).value=modulo
        HC.cell(row=rowwrite, column=columnwrite).font = fontwhite
        HC.cell(row=rowwrite, column=columnwrite).fill=my_filldarkgray
        HC.cell(row=rowwrite, column=columnwrite).alignment = alignment
                
        rowwrite-=1
        
        HC.cell(row=rowwrite-1, column=columnwrite).value=olddictfamily.get(modulo)
        HC.cell(row=rowwrite-1, column=columnwrite).font = fontwhite
        HC.cell(row=rowwrite-1, column=columnwrite).fill=my_filldarkgray
        HC.cell(row=rowwrite-1, column=columnwrite).alignment = alignment
        
        HC.cell(row=rowwrite, column=columnwrite).value=modulo
        HC.cell(row=rowwrite, column=columnwrite).font = fontwhite
        HC.cell(row=rowwrite, column=columnwrite).fill=my_filldarkgray
        HC.cell(row=rowwrite, column=columnwrite).alignment = alignment
        
        columnwrite+=1
        
        HC.cell(row=rowwrite, column=columnwrite).value="Delta"
        HC.cell(row=rowwrite, column=columnwrite).font = fontwhite
        HC.cell(row=rowwrite, column=columnwrite).fill=my_filldarkgray
        HC.cell(row=rowwrite, column=columnwrite).alignment = alignment
        
        rowwrite+=1
        
        HC.cell(row=rowwrite, column=columnwrite).value=""
        HC.cell(row=rowwrite, column=columnwrite).font = fontwhite
        HC.cell(row=rowwrite, column=columnwrite).fill=my_filldarkgray
        HC.cell(row=rowwrite, column=columnwrite).alignment = alignment
        
        enprocesscolumn=columnwrite-2
        
###################################################################################################################
##################WRITE PART NUMBERS BY GROUP - WRITE PART NUMBER BY GROUP *####################################### 
###################################################################################################################
        for grupo in grouptable_list:

            iterator=dictionary_of_groups.get(grupo)
            klimaiterator=iterator[1]
            nonklimaitarator=iterator[0]

            old=[]
            new=[]
            for element in klimaiterator:
                #print(element)

                if(str(element[1])=="New"):
                    new.append(element[0])
                    
                if(str(element[1])=="old"):
                    old.append(element[0])
                    
            HC.cell(row=endprocessrow, column=enprocesscolumn).value=float(sum(new))
            HC.cell(row=endprocessrow, column=enprocesscolumn).border=cell_border
            
            HC.cell(row=endprocessrow, column=enprocesscolumn+1).value=float(sum(old))
            HC.cell(row=endprocessrow, column=enprocesscolumn+1).border=cell_border
            
            internal_delta=(float(sum(new)))-(float(sum(old)))
            
            
            if(internal_delta>0):
                HC.cell(row=endprocessrow, column=enprocesscolumn+2).value=internal_delta
                HC.cell(row=endprocessrow, column=enprocesscolumn+2).font = fontgreen
                HC.cell(row=endprocessrow, column=enprocesscolumn+2).fill=my_fillgreen
                HC.cell(row=endprocessrow, column=enprocesscolumn+2).border=cell_border
            elif(internal_delta<0):
                HC.cell(row=endprocessrow, column=enprocesscolumn+2).value=internal_delta
                HC.cell(row=endprocessrow, column=enprocesscolumn+2).font = fontred
                HC.cell(row=endprocessrow, column=enprocesscolumn+2).fill=my_fillred
                HC.cell(row=endprocessrow, column=enprocesscolumn+2).border=cell_border
            else:
                HC.cell(row=endprocessrow, column=enprocesscolumn+2).value=internal_delta
                HC.cell(row=endprocessrow, column=enprocesscolumn+2).border=cell_border
                
            
            old=[]
            new=[]
            
            
            for element in nonklimaitarator:
                #print(element)

                if(str(element[1])=="New"):

                    new.append(element[0])
                    
                if(str(element[1])=="old"):

                    old.append(element[0])
                    
            listanonklimaswitch.append(sum(new))
            listanonklimaswitch.append(sum(old))
            
            
            HC.cell(row=endprocessrow+37, column=enprocesscolumn).value=float(sum(new))
            HC.cell(row=endprocessrow+37, column=enprocesscolumn).border=cell_border
            
            HC.cell(row=endprocessrow+37, column=enprocesscolumn+1).value=float(sum(old))
            HC.cell(row=endprocessrow+37, column=enprocesscolumn+1).border=cell_border
            
            internal_delta=(float(sum(new)))-(float(sum(old)))
                        
            if(internal_delta>0):
                HC.cell(row=endprocessrow+37, column=enprocesscolumn+2).value=internal_delta
                HC.cell(row=endprocessrow+37, column=enprocesscolumn+2).border=cell_border
                
                HC.cell(row=endprocessrow+37, column=enprocesscolumn+2).font = fontgreen
                HC.cell(row=endprocessrow+37, column=enprocesscolumn+2).border=cell_border
                
                HC.cell(row=endprocessrow+37, column=enprocesscolumn+2).fill=my_fillgreen
                HC.cell(row=endprocessrow+37, column=enprocesscolumn+2).border=cell_border
                
            elif(internal_delta<0):
                HC.cell(row=endprocessrow+37, column=enprocesscolumn+2).value=internal_delta
                HC.cell(row=endprocessrow+37, column=enprocesscolumn+2).border=cell_border
                
                HC.cell(row=endprocessrow+37, column=enprocesscolumn+2).font = fontred
                HC.cell(row=endprocessrow+37, column=enprocesscolumn+2).border=cell_border
                
                HC.cell(row=endprocessrow+37, column=enprocesscolumn+2).fill=my_fillred
                HC.cell(row=endprocessrow+37, column=enprocesscolumn+2).border=cell_border
                  
            else:
                HC.cell(row=endprocessrow+37, column=enprocesscolumn+2).value=internal_delta
                HC.cell(row=endprocessrow+37, column=enprocesscolumn+2).border=cell_border
                
            endprocessrow+=1    
        
        if(sum(listanonklimaswitch)>0):
            HC.cell(row=endprocessrow+38, column=enprocesscolumn).value="YES"
            HC.merge_cells(start_row=endprocessrow+38, start_column=enprocesscolumn, 
            end_row=endprocessrow+38, end_column=enprocesscolumn+2)
            HC.cell(row=endprocessrow+38, column=enprocesscolumn).font = fontwhite
            HC.cell(row=endprocessrow+38, column=enprocesscolumn).fill=my_filldarkgray
            HC.cell(row=endprocessrow+38, column=enprocesscolumn).alignment = alignment
        else:
            HC.cell(row=endprocessrow+38, column=enprocesscolumn).value="NO"
            HC.merge_cells(start_row=endprocessrow+38, start_column=enprocesscolumn, 
            end_row=endprocessrow+38, end_column=enprocesscolumn+2)
            HC.cell(row=endprocessrow+38, column=enprocesscolumn).font = fontwhite
            HC.cell(row=endprocessrow+38, column=enprocesscolumn).fill=my_filldarkgray
            HC.cell(row=endprocessrow+38, column=enprocesscolumn).alignment = alignment
            
        
        rowwrite+=-1
        columnwrite+=2
        
    book.save("Summary_comparison.xlsx")
    
    book.create_sheet('Summary')
    HS=book['Summary']
    
    startrow=1
    startcolumn=2
    specialcolumn=2
    numberofwirescolumn=2
    specialrow=1
    
    Assembly_nk2=[]
    Bolt_nk2=[]
    Busbar_nk2=[]
    Cavity_Plug_nk2=[]
    Cavity_Seal_nk2=[]
    Channel_nk2=[]
    Clip_nk2=[]
    Connector_nk2=[]
    Connector_FH_nk2=[]
    Connector_Name_nk2=[]
    Cover_nk2=[]
    Eyelet_nk2=[]
    Fuse_nk2=[]
    Fuse_Box_nk2=[]
    Grommet_nk2=[]
    Heatshrink_nk2=[]
    Housing_nk2=[]
    Label_nk2=[]
    Mating_nk2=[]
    Multicore_Wire_nk2=[]
    Nut_nk2=[]
    Other_nk2=[]
    Relay_nk2=[]
    Resistor_nk2=[]
    SOLDER_nk2=[]
    Spacer_nk2=[]
    Splice_nk2=[]
    Strap_nk2=[]
    Strap_Holder_nk2=[]
    Tape_nk2=[]
    Terminal_nk2=[]
    Tube_nk2=[]
    Washer_nk2=[]
    Washerhose_nk2=[]
    Wire_nk2=[]
    
    Assembly_k2=[]
    Bolt_k2=[]
    Busbar_k2=[]
    Cavity_Plug_k2=[]
    Cavity_Seal_k2=[]
    Channel_k2=[]
    Clip_k2=[]
    Connector_k2=[]
    Connector_FH_k2=[]
    Connector_Name_k2=[]
    Cover_k2=[]
    Eyelet_k2=[]
    Fuse_k2=[]
    Fuse_Box_k2=[]
    Grommet_k2=[]
    Heatshrink_k2=[]
    Housing_k2=[]
    Label_k2=[]
    Mating_k2=[]
    Multicore_Wire_k2=[]
    Nut_k2=[]
    Other_k2=[]
    Relay_k2=[]
    Resistor_k2=[]
    SOLDER_k2=[]
    Spacer_k2=[]
    Splice_k2=[]
    Strap_k2=[]
    Strap_Holder_k2=[]
    Tape_k2=[]
    Terminal_k2=[]
    Tube_k2=[]
    Washer_k2=[]
    Washerhose_k2=[]
    Wire_k2=[]
    
    dictionary_of_groups2={"Assembly":[Assembly_nk2,Assembly_k2],"Bolt":[Bolt_nk2,Bolt_k2],"Busbar":[Busbar_nk2,Busbar_k2]
                          ,"Cavity Plug":[Cavity_Plug_nk2,Cavity_Plug_k2],"Cavity Seal":[Cavity_Seal_nk2,Cavity_Seal_k2],"Channel":[Channel_nk2,Channel_k2]
                          ,"Clip":[Clip_nk2,Clip_k2],"Connector":[Connector_nk2,Connector_k2],"Connector FH":[Connector_FH_nk2,Connector_FH_k2]
                          ,"Connector Name":[Connector_Name_nk2,Connector_Name_k2],"Cover":[Cover_nk2,Cover_k2],"Eyelet":[Eyelet_nk2,Eyelet_k2]
                          ,"Fuse":[Fuse_nk2,Fuse_k2],"Fuse Box":[Fuse_Box_nk2,Fuse_Box_k2],"Grommet":[Grommet_nk2,Grommet_k2]
                          ,"Heatshrink":[Heatshrink_nk2,Heatshrink_k2],"Housing":[Housing_nk2,Housing_k2],"Label":[Label_nk2,Label_k2]
                          ,"Mating":[Mating_nk2,Mating_k2],"Multicore Wire":[Multicore_Wire_nk2,Multicore_Wire_k2],"Nut":[Nut_nk2,Nut_k2]
                          ,"Other":[Other_nk2,Other_k2],"Relay":[Relay_nk2,Relay_k2],"Resistor":[Resistor_nk2,Resistor_k2]
                          ,"Solder":[SOLDER_nk2,SOLDER_k2],"Spacer":[Spacer_nk2,Spacer_k2],"Splice":[Splice_nk2,Splice_k2]
                          ,"Strap":[Strap_nk2,Strap_k2],"Strap Holder":[Strap_Holder_nk2,Strap_Holder_k2],"Tape":[Tape_nk2,Tape_k2]
                          ,"Terminal":[Terminal_nk2,Terminal_k2],"Tube":[Tube_nk2,Tube_k2],"Washer":[Washer_nk2,Washer_k2]
                          ,"Washerhose":[Washerhose_nk2,Washerhose_k2],"Wire":[Wire_nk2,Wire_k2]}
    
    
    modules_changed=[]
    #Calculate values for summary table (little one)--------------------------------------------------------------------------------
    for modulo in modules_new:
        specialrow2=3

        Assembly_nk=[]
        Bolt_nk=[]
        Busbar_nk=[]
        Cavity_Plug_nk=[]
        Cavity_Seal_nk=[]
        Channel_nk=[]
        Clip_nk=[]
        Connector_nk=[]
        Connector_FH_nk=[]
        Connector_Name_nk=[]
        Cover_nk=[]
        Eyelet_nk=[]
        Fuse_nk=[]
        Fuse_Box_nk=[]
        Grommet_nk=[]
        Heatshrink_nk=[]
        Housing_nk=[]
        Label_nk=[]
        Mating_nk=[]
        Multicore_Wire_nk=[]
        Nut_nk=[]
        Other_nk=[]
        Relay_nk=[]
        Resistor_nk=[]
        SOLDER_nk=[]
        Spacer_nk=[]
        Splice_nk=[]
        Strap_nk=[]
        Strap_Holder_nk=[]
        Tape_nk=[]
        Terminal_nk=[]
        Tube_nk=[]
        Washer_nk=[]
        Washerhose_nk=[]
        Wire_nk=[]
        
        Assembly_k=[]
        Bolt_k=[]
        Busbar_k=[]
        Cavity_Plug_k=[]
        Cavity_Seal_k=[]
        Channel_k=[]
        Clip_k=[]
        Connector_k=[]
        Connector_FH_k=[]
        Connector_Name_k=[]
        Cover_k=[]
        Eyelet_k=[]
        Fuse_k=[]
        Fuse_Box_k=[]
        Grommet_k=[]
        Heatshrink_k=[]
        Housing_k=[]
        Label_k=[]
        Mating_k=[]
        Multicore_Wire_k=[]
        Nut_k=[]
        Other_k=[]
        Relay_k=[]
        Resistor_k=[]
        SOLDER_k=[]
        Spacer_k=[]
        Splice_k=[]
        Strap_k=[]
        Strap_Holder_k=[]
        Tape_k=[]
        Terminal_k=[]
        Tube_k=[]
        Washer_k=[]
        Washerhose_k=[]
        Wire_k=[]
        
        dictionary_of_groups={"Assembly":[Assembly_nk,Assembly_k],"Bolt":[Bolt_nk,Bolt_k],"Busbar":[Busbar_nk,Busbar_k]
                              ,"Cavity Plug":[Cavity_Plug_nk,Cavity_Plug_k],"Cavity Seal":[Cavity_Seal_nk,Cavity_Seal_k],"Channel":[Channel_nk,Channel_k]
                              ,"Clip":[Clip_nk,Clip_k],"Connector":[Connector_nk,Connector_k],"Connector FH":[Connector_FH_nk,Connector_FH_k]
                              ,"Connector Name":[Connector_Name_nk,Connector_Name_k],"Cover":[Cover_nk,Cover_k],"Eyelet":[Eyelet_nk,Eyelet_k]
                              ,"Fuse":[Fuse_nk,Fuse_k],"Fuse Box":[Fuse_Box_nk,Fuse_Box_k],"Grommet":[Grommet_nk,Grommet_k]
                              ,"Heatshrink":[Heatshrink_nk,Heatshrink_k],"Housing":[Housing_nk,Housing_k],"Label":[Label_nk,Label_k]
                              ,"Mating":[Mating_nk,Mating_k],"Multicore Wire":[Multicore_Wire_nk,Multicore_Wire_k],"Nut":[Nut_nk,Nut_k]
                              ,"Other":[Other_nk,Other_k],"Relay":[Relay_nk,Relay_k],"Resistor":[Resistor_nk,Resistor_k]
                              ,"Solder":[SOLDER_nk,SOLDER_k],"Spacer":[Spacer_nk,Spacer_k],"Splice":[Splice_nk,Splice_k]
                              ,"Strap":[Strap_nk,Strap_k],"Strap Holder":[Strap_Holder_nk,Strap_Holder_k],"Tape":[Tape_nk,Tape_k]
                              ,"Terminal":[Terminal_nk,Terminal_k],"Tube":[Tube_nk,Tube_k],"Washer":[Washer_nk,Washer_k]
                              ,"Washerhose":[Washerhose_nk,Washerhose_k],"Wire":[Wire_nk,Wire_k]}
        
        nuevalista=newdict3.get(modulo)
        viejalista=olddict3.get(modulo)
        
        
        if(nuevalista!=None):
            for elemento in nuevalista: 
                dato=nuevalista.get(elemento)
                
                if(dato[8]=="KLIMA"):
                    dictionary_of_groups.get(dato[3])[1].append([dato[0],"NEW"])
                if(dato[8]=="NON-KLIMA"):
                    dictionary_of_groups.get(dato[3])[0].append([dato[0],"NEW"])
                   
        if(viejalista!=None):
            for elemento in viejalista: 
                dato=viejalista.get(elemento)
                
                if(dato[8]=="KLIMA"):
                    dictionary_of_groups.get(dato[3])[1].append([dato[0],"OLD"])
                if(dato[8]=="NON-KLIMA"):
                    dictionary_of_groups.get(dato[3])[0].append([dato[0],"OLD"])
        
        print("dictionaryyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyy")
        
        print(dictionary_of_groups)
        

        
        writedict=[]
        for key in dictionary_of_groups:
            print(key)
            #print(key)
            guilty=dictionary_of_groups.get(key)
            
            print("Guioltyyyyyyyyyyyyyyyyyyyyyyyyyyyyy")
            print(guilty)
            
            klimabox=guilty[1]
            nonklimabox=guilty[0]
            
            newbox=[]
            oldbox=[]
            
            for element in klimabox:
                variable=str(element[1])
                if(variable=="OLD"):
                    oldbox.append(element[0])
                else:
                    newbox.append(element[0])
            
            print("Klima-----------------------------------------------")
            print(newbox)
            print(oldbox)
            
            totalold=sum(oldbox)
            totalnew=sum(newbox)
            print("totals")
            print(totalold)
            print(totalnew)
            totalcomponent=totalnew-totalold
            
            print(totalcomponent)
            
           
            
            writedict.append(totalcomponent)

            newbox=[]
            oldbox=[]
            
            #print(nonklimabox)
            
            for element in nonklimabox:
                variable=str(element[1])
                if(variable=="OLD"):
                    oldbox.append(element[0])
                else:
                    newbox.append(element[0])
            
            print("nonklima-----------------------------------------------")
            print(newbox)
            print(oldbox)
            
            totalold=sum(oldbox)
            totalnew=sum(newbox)
            print("totals")
            print(totalold)
            print(totalnew)
                       
            totalcomponent2=totalnew-totalold
            print(totalcomponent2)
            
            
            writedict.append(totalcomponent2)
            
            
            print("WRITEDICT")
            print(writedict)
            print(len(writedict))
        
            
            
        print("WRITEDICT")
        print(writedict)
        print(len(writedict))
        
        

            

        paso=85
        print(paso)
        bar['value'] = paso
        root.update_idletasks()   
###################################################################################################################
##################ADD MODULES WITH CHGANGES ON COMPONENTS OR NUMBER OF WIRES####################################### 
###################################################################################################################
           
            
        if(number_of_wires_dict.get(modulo)!=None):
            writedict.append(sum(number_of_wires_dict.get(modulo)))
            
        if (sum(writedict)!=0):
            
            modules_changed.append(modulo)
            
            HS.cell(row=specialrow, column=specialcolumn).value = modulo
            HS.cell(row=specialrow, column=specialcolumn).border=cell_border
            HS.cell(row=specialrow, column=specialcolumn).alignment = Alignment(horizontal='center')
            HS.cell(row=specialrow, column=specialcolumn).alignment = Alignment(textRotation=180)
            
            if(olddict3.get(modulo)==None):
                HS.cell(row=specialrow, column=specialcolumn).fill=my_filldarkgreen
                HS.cell(row=specialrow, column=specialcolumn).font = fontwhite
                
            if(newdict3.get(modulo)==None):
                HS.cell(row=specialrow, column=specialcolumn).fill=my_filldarkred
                HS.cell(row=specialrow, column=specialcolumn).font = fontwhite
                
            if(newdict3.get(modulo)!=None and olddict3.get(modulo)!=None):
                HS.cell(row=specialrow, column=specialcolumn).fill=my_fillyellow
                HS.cell(row=specialrow, column=specialcolumn).font = fontorange
            
            counterwrite=0
            
            for key in dictionary_of_groups:
                
                print(key)
                print(counterwrite)
                dictionary_of_groups2.get(key)[1].append(writedict[counterwrite]) 
                
                HS.cell(row=specialrow2, column=specialcolumn).value = round(writedict[counterwrite], 2)
                HS.cell(row=specialrow2, column=specialcolumn).border = cell_border
                
                if(writedict[counterwrite]>0):
                    HS.cell(row=specialrow2, column=specialcolumn).font = fontgreen
                    HS.cell(row=specialrow2, column=specialcolumn).fill=my_fillgreen
                    HS.cell(row=specialrow2, column=specialcolumn).border = cell_border
                    
                if(writedict[counterwrite]<0):
                    HS.cell(row=specialrow2, column=specialcolumn).font = fontred
                    HS.cell(row=specialrow2, column=specialcolumn).fill=my_fillred
                    HS.cell(row=specialrow2, column=specialcolumn).border = cell_border
                    
                #"Second write-----------------------------------------------------------------------------------------"
                    
                HS.cell(row=specialrow2+36, column=specialcolumn).value = round(writedict[counterwrite+1], 2)
                HS.cell(row=specialrow2+36, column=specialcolumn).border = cell_border
            
                
                dictionary_of_groups2.get(key)[0].append(writedict[counterwrite+1]) 
                
                if(writedict[counterwrite+1]>0):
                    HS.cell(row=specialrow2+36, column=specialcolumn).font = fontgreen
                    HS.cell(row=specialrow2+36, column=specialcolumn).fill=my_fillgreen
                    HS.cell(row=specialrow2+36, column=specialcolumn).border = cell_border
                    
                if(writedict[counterwrite+1]<0):
                    HS.cell(row=specialrow2+36, column=specialcolumn).font = fontred
                    HS.cell(row=specialrow2+36, column=specialcolumn).fill=my_fillred
                    HS.cell(row=specialrow2+36, column=specialcolumn).border = cell_border
                    
                specialrow2+=1
                counterwrite+=2
                

            specialcolumn+=1
            
###################################################################################################################
##################WRITE TOTALS COLUMN K - WRITE TOTALS COLUMNK - WRITE TOTAL####################################### 
###################################################################################################################
    
    HS.cell(row=specialrow, column=specialcolumn).value="Total"
    HS.cell(row=specialrow, column=specialcolumn).font = fontwhite
    HS.cell(row=specialrow, column=specialcolumn).fill=my_filldarkgray
    HS.cell(row=specialrow, column=specialcolumn).alignment = alignment
    
    specialrow25=3
    
    for key in dictionary_of_groups2:
        print(key)
        llave=dictionary_of_groups2.get(key)
        print(llave)
        
        klima=llave[1]
        nonklima=llave[0]
        
        print(klima)
        print(nonklima)
        
        print(sum(klima))
        print(sum(nonklima))
        
        HS.cell(row=specialrow25, column=specialcolumn).value=sum(klima)
        HS.cell(row=specialrow25, column=specialcolumn).border = cell_border
            
        if(sum(klima)>0):
            HS.cell(row=specialrow25, column=specialcolumn).font = fontwhite
            HS.cell(row=specialrow25, column=specialcolumn).fill=my_filldarkgreen
            HS.cell(row=specialrow25, column=specialcolumn).border = cell_border
            
        if(sum(klima)<0):
            HS.cell(row=specialrow25, column=specialcolumn).font = fontwhite
            HS.cell(row=specialrow25, column=specialcolumn).fill=my_filldarkred
            HS.cell(row=specialrow25, column=specialcolumn).border = cell_border
            
        print(specialrow25)
        
        HS.cell(row=specialrow25+36, column=specialcolumn).value=sum(nonklima)
        HS.cell(row=specialrow25+36, column=specialcolumn).border = cell_border
        
        if(sum(nonklima)>0):
            HS.cell(row=specialrow25+36, column=specialcolumn).font = fontwhite
            HS.cell(row=specialrow25+36, column=specialcolumn).fill=my_filldarkgreen
            HS.cell(row=specialrow25+36, column=specialcolumn).border = cell_border
            
        if(sum(nonklima)<0):
            HS.cell(row=specialrow25+36, column=specialcolumn).font = fontwhite
            HS.cell(row=specialrow25+36, column=specialcolumn).fill=my_filldarkred
            HS.cell(row=specialrow25+36, column=specialcolumn).border = cell_border
            
        
        specialrow25=specialrow25+1
        
        numberofwirestartrow=specialrow25+37

    specialcolumn=specialcolumn+2

#################################################################################################################
##################################WRITE NUMBER OF WIRES MINI TABLE###############################################
################################################################################################################# 
   #now=number of wires
   
    now_total_klimabox=[]
    now_total_nonklimabox=[]
    now_totalbox=[]
    number_wires_totals=[now_total_klimabox,now_total_nonklimabox,now_totalbox]
    

    for modulo in modules_changed:
        counter=0
        
        if(number_of_wires_dict.get(modulo)!=None):
            wire_values=number_of_wires_dict.get(modulo)
        else:
            wire_values=[0,0,0]
        temprownow=numberofwirestartrow
        
        for setofvalues in wire_values:
            number_wires_totals[counter].append(setofvalues)
            HS.cell(row=temprownow, column=numberofwirescolumn).value=setofvalues
            if(setofvalues>0):
                HS.cell(row=temprownow, column=numberofwirescolumn).font = fontgreen
                HS.cell(row=temprownow, column=numberofwirescolumn).fill=my_fillgreen
            if(setofvalues<0):
                HS.cell(row=temprownow, column=numberofwirescolumn).font = fontred
                HS.cell(row=temprownow, column=numberofwirescolumn).fill=my_fillred
                  
            HS.cell(row=temprownow, column=numberofwirescolumn).border = cell_border
            temprownow+=1
            counter=counter+1
        numberofwirescolumn+=1
        
    for total in number_wires_totals:
        HS.cell(row=numberofwirestartrow, column=numberofwirescolumn).value=sum(total)
        HS.cell(row=numberofwirestartrow, column=numberofwirescolumn).border = cell_border
        if(sum(total)>0):
            HS.cell(row=numberofwirestartrow, column=numberofwirescolumn).font = fontwhite
            HS.cell(row=numberofwirestartrow, column=numberofwirescolumn).fill=my_filldarkgreen
        if(sum(total)<0):
            HS.cell(row=numberofwirestartrow, column=numberofwirescolumn).font = fontwhite
            HS.cell(row=numberofwirestartrow, column=numberofwirescolumn).fill=my_filldarkred
        numberofwirestartrow+=1
        

            
    specialrow50=1
    
    HS.cell(row=specialrow50, column=specialcolumn+1).value="Total"
    HS.cell(row=specialrow50, column=specialcolumn+1).font = fontwhite
    HS.cell(row=specialrow50, column=specialcolumn+1).fill=my_filldarkgray
    HS.cell(row=specialrow50, column=specialcolumn+1).alignment = alignment
    
    specialrow50=specialrow50+1
 
    HS.cell(row=specialrow50, column=specialcolumn).value="Global"
    HS.cell(row=specialrow50, column=specialcolumn).font = fontwhite
    HS.cell(row=specialrow50, column=specialcolumn).fill=my_filldarkgray
    HS.cell(row=specialrow50, column=specialcolumn).alignment = alignment
    specialrow50=specialrow50+1
    
    specialrow51=3
    
    for group in grouptable_list:
        HS.cell(row=specialrow50, column=specialcolumn).value=group
        HS.cell(row=specialrow50, column=specialcolumn).border=cell_border
        specialrow50+=1
        
    for key in dictionary_of_groups2:
        globala=dictionary_of_groups2.get(key)
        
        HS.cell(row=specialrow51, column=specialcolumn+1).value=sum(globala[0])+sum(globala[1])
        HS.cell(row=specialrow51, column=specialcolumn+1).border = cell_border
        
        if(sum(globala[0])+sum(globala[1])>0):
            HS.cell(row=specialrow51, column=specialcolumn+1).font = fontwhite
            HS.cell(row=specialrow51, column=specialcolumn+1).fill=my_filldarkgreen
            HS.cell(row=specialrow51, column=specialcolumn+1).border = cell_border
            
        if(sum(globala[0])+sum(globala[1])<0):
            HS.cell(row=specialrow51, column=specialcolumn+1).font = fontwhite
            HS.cell(row=specialrow51, column=specialcolumn+1).fill=my_filldarkred
            HS.cell(row=specialrow51, column=specialcolumn+1).border = cell_border
        specialrow51=specialrow51+1
        
#################################################################################################################
##################################MODULES SUMMARY TABLE-MODULES --###############################################
################################################################################################################# 

    Specialrow52=1
    specialcolumn3=specialcolumn+3
    HS.cell(row=Specialrow52, column=specialcolumn3).value="Modules Summary"
    HS.merge_cells(start_row=Specialrow52, start_column=specialcolumn3, end_row=Specialrow52, end_column=specialcolumn3+2)
    HS.cell(row=Specialrow52, column=specialcolumn3).font = fontwhite
    HS.cell(row=Specialrow52, column=specialcolumn3).fill=my_filldarkgray
    HS.cell(row=Specialrow52, column=specialcolumn3).alignment = alignment
    Specialrow52=Specialrow52+1
    
    HS.cell(row=Specialrow52, column=specialcolumn3).value="Modules ADDED: "+str(len(modules_added))
    HS.cell(row=Specialrow52, column=specialcolumn3).font = fontwhite
    HS.cell(row=Specialrow52, column=specialcolumn3).fill=my_filldarkgreen
    HS.cell(row=Specialrow52, column=specialcolumn3).alignment = alignment
    HS.cell(row=Specialrow52, column=specialcolumn3).border = cell_border
    
    added_row=Specialrow52+1
    for modulo in modules_added:
        HS.cell(row=added_row, column=specialcolumn3).value=str(modulo)
        HS.cell(row=added_row, column=specialcolumn3).border = cell_border
        added_row+=1
        
    
    HS.cell(row=Specialrow52, column=specialcolumn3+1).value="Modules REMOVED: "+str(len(modules_removed))
    HS.cell(row=Specialrow52, column=specialcolumn3+1).font = fontwhite  
    HS.cell(row=Specialrow52, column=specialcolumn3+1).fill=my_filldarkred
    HS.cell(row=Specialrow52, column=specialcolumn3+1).alignment = alignment
    HS.cell(row=Specialrow52, column=specialcolumn3+1).border = cell_border
    
    removed_row=Specialrow52+1
    for modulo in modules_removed:
        HS.cell(row=removed_row, column=specialcolumn3+1).value=str(modulo)
        HS.cell(row=removed_row, column=specialcolumn3+1).border = cell_border
        removed_row+=1
    
    changed_modules_for_real=list(set(modules_changed) - set((modules_added+modules_removed)))
    
    HS.cell(row=Specialrow52, column=specialcolumn3+2).value="Modules CHANGED: "+str(len(changed_modules_for_real))
    HS.cell(row=Specialrow52, column=specialcolumn3+2).font = fontorange
    HS.cell(row=Specialrow52, column=specialcolumn3+2).fill=my_fillyellow
    HS.cell(row=Specialrow52, column=specialcolumn3+2).alignment = alignment     
    HS.cell(row=Specialrow52, column=specialcolumn3+2).border = cell_border
    

    
    row_changed=Specialrow52+1
    for modulo in changed_modules_for_real:
        HS.cell(row=row_changed, column=specialcolumn3+2).value=str(modulo)
        HS.cell(row=row_changed, column=specialcolumn3+2).border = cell_border
        row_changed+=1

    #   Write first row of the table----------------------------------------------------------------------------------------
    HS.cell(row=2, column=1).value = "Klima"
    HS.cell(row=2, column=1).font = fontwhite
    HS.cell(row=2, column=1).fill=my_filldarkgray
        
    startcolumn+=1

    HS.cell(row=2, column=1).value = "Klima"
    HS.cell(row=2, column=1).font = fontwhite
    HS.cell(row=2, column=1).fill=my_filldarkgray
    
    startrow=3
    
    for group in grouptable_list:
        HS.cell(row=startrow, column=1).value = group
        HS.cell(row=startrow, column=1).border=cell_border
        startrow+=1
        
    HS.cell(row=startrow, column=1).value = "NON-Klima"
    HS.cell(row=startrow, column=1).font = fontwhite
    HS.cell(row=startrow, column=1).fill=my_filldarkgray
    startrow+=1
    
    for group in grouptable_list:
        HS.cell(row=startrow, column=1).value = group
        HS.cell(row=startrow, column=1).border=cell_border
        startrow+=1
        
    HS.cell(row=startrow, column=1).value = "Number of wires"
    HS.cell(row=startrow, column=1).font = fontwhite
    HS.cell(row=startrow, column=1).fill=my_filldarkgray
    startrow+=1
    
    names=["Klima","Non-Klima","Total"]
    
    for name in names:
        HS.cell(row=startrow, column=1).value = name
        HS.cell(row=startrow, column=1).border=cell_border
        startrow+=1

    book.save("Summary_comparison.xlsx")
    
    print("TERMINAO")
    
    print(number_of_wires_dict)
    
    print(number_wires_totals)
    
    print(modules_added)
    print(modules_removed)
    
    paso=100
    print(paso)
    bar['value'] = paso
    root.update_idletasks()
    
    pyautogui.alert(text='Proccess Completed', title='', button='OK')
    
    
def BOM_CUTSHEET_PREPARATION():
    
    my_grayf = openpyxl.styles.colors.Color(rgb='00000080')
    my_fillgray = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_grayf)
    
    font7 = Font(name='Calibri',
    size=8,
    bold=True,
    italic=True,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FFFFFF')
    number_format = 'Accounting'
    
    pyautogui.alert(text='Select PDF File', title='Select File', button='OK')
    
    path2 = askdirectory(title='Select Folder') # shows dialog box and return the path
    
    list_of_directories=os.listdir(path2)
    
    for directory in list_of_directories:
        
        path=str(path2)+'/'+str(directory)
    
        os.chdir(path)
        
        list_of_files=os.listdir(path)
        
        #Get full name of BOM file--------------------------------------------------------------------------------
        for file in list_of_files:
            if(str(file[-8:])=="BOM.xlsx"):
                print(file[-8:])
                print(file)
                BOM=file
        new_bom=str(BOM[:-5]+"-copy.xlsx")
       
        list_of_files=os.listdir(path)
                
        #Get full name of BOM file--------------------------------------------------------------------------------
        for file in list_of_files:
            if(str(file[-8:])=="eet.xlsx"):
                print(file[-8:])
                print(file)
                file=str(file)
                file_removed=str(file[:-5])
        
        print(file)
        
        file=str(file_removed)+".xlsx"
        print(file_removed)
        
        #Copy CS
        
        print("COPIAR")
        print(file)
        print(file_removed+"-copy.xlsx")
        shutil.copyfile(file, file_removed+"-copy.xlsx")
         #Copy CS
        shutil.copyfile(BOM, new_bom)
        
        book = openpyxl.load_workbook(file_removed+"-copy.xlsx", data_only=True)
        CS=book["Wire Sheet"]
        
        directorio=path
    
    #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    ###################################################################################################################@
    ##################READ BOM READ BOM READ BOM READ BOM READ BOM REA#################################################@ 
    ###################################################################################################################@   
    #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ 
    
        
        book_bom = openpyxl.load_workbook(new_bom, data_only=True)
        BS=book_bom["Detail"]
        
        ir_boom=11
        ic_boom=1
        
        number_of_rows_bom=0
        while True:
            cell_obj = BS.cell(row = ir_boom, column = ic_boom)
            border=str(cell_obj.border.bottom.style)
            print(border)
            ir_boom+=1
            number_of_rows_bom+=1
            if(border=="None"):
                break
        number_of_rows_bom=number_of_rows_bom-1
    
        
    ###################################################################################################################
    ################## Insert CW Weight column- Insert CW Weight column- Insert CW W###################################
    ###################################################################################################################
    
        ir_boom=9
        ic_boom=1
        
        
        while True:
            cell_obj = BS.cell(row = ir_boom, column = ic_boom)
            value=cell_obj.value
            print(value)
            
            if(str(value)=="CPN"):
                BS.insert_cols(ic_boom+1)
                book_bom.save(new_bom)
                column_cw_title=ic_boom+1
                break
            ic_boom=ic_boom+1
            
    #Insert CW Weight column title------------------------------
    
        BS.cell(row=ir_boom, column=column_cw_title).value="CW Weight"
        BS.cell(row=ir_boom, column=column_cw_title).border=cell_border
        BS.cell(row=ir_boom, column=column_cw_title).font = fontorange
        BS.cell(row=ir_boom, column=column_cw_title).fill=my_fillyellow
        BS.cell(row=ir_boom, column=column_cw_title).alignment = alignment
        
    #insert borders under CW Weight title---------------------------------------------
            
        ir_boom=11
        
        for row in range(number_of_rows_bom):
            BS.cell(row=row+11, column=column_cw_title).border=cell_border
        book_bom.save(new_bom)
        
    ###################################################################################################################
    ##################Rewrite modules PN-Rewrite modules PN-Rewrite modules PN-Rewr ###################################
    ###################################################################################################################
    
        ir_boom=9
        ic_boom=1
        
        list_of_modules=[]
        while True:
            cell_obj = BS.cell(row = ir_boom, column = ic_boom)
            value=cell_obj.value
            print(value)
            #time.sleep(1)
            if(str(value[0])=="A"):
                pn=str(value).split("-")
                BS.cell(row=ir_boom, column=ic_boom).value=pn[0]
                list_of_modules.append(pn[0])
                BS.cell(row=ir_boom, column=ic_boom).border=cell_border
                BS.cell(row=ir_boom, column=ic_boom).font = fontorange
                BS.cell(row=ir_boom, column=ic_boom).fill=my_fillyellow
                BS.cell(row=ir_boom, column=ic_boom).alignment = alignment
                book_bom.save(new_bom)
            if(str(value)=="TOTAL"):
                break
            ic_boom=ic_boom+1
        book_bom.save(new_bom)
        
    ###################################################################################################################
    ##################Rewrite UOM value- Rewrite UOM value- Rewrite UOM Value - Rewr###################################
    ###################################################################################################################
        ir_boom=9
        ic_boom=1
    
        
        while True:
            cell_obj = BS.cell(row = ir_boom, column = ic_boom)
            value=cell_obj.value
    
            
            if(str(value)=="Klima"):
                BS.cell(row=ir_boom, column=ic_boom+1).value="UOM"
                BS.cell(row=ir_boom, column=ic_boom+1).border=cell_border
                BS.cell(row=ir_boom, column=ic_boom+1).font = fontorange
                BS.cell(row=ir_boom, column=ic_boom+1).fill=my_fillyellow
                BS.cell(row=ir_boom, column=ic_boom+1).alignment = alignment
                
                uom_column=ic_boom+1
                
                book_bom.save(new_bom)
                break
            ic_boom=ic_boom+1
             
    ###################################################################################################################
    ##################Copy Detail sheet-Copy Detail sheet-Copy Detail sheet-Copy Det###################################
    ##################change name of worksheet-change name of worksheet-change name ###################################
    ###################################################################################################################
        book_bom.copy_worksheet(BS)
        book_bom.save(new_bom)
            
        conv_sheet = book_bom.get_sheet_by_name('Detail Copy')
        print(conv_sheet)
        conv_sheet.title  = 'Detail(Conv)'
        book_bom.save(new_bom)
        
    ###################################################################################################################
    ##################Insert formulas titles bottom module matrix table-###############################################
    ###################################################################################################################
        titles_to_insert=["Each","Per Length","Total","Validation","Total"]
        
        ir_boom=number_of_rows_bom+12
        for title in titles_to_insert:
            BS.cell(row=ir_boom, column=uom_column).value=title
            ir_boom=ir_boom+1
        book_bom.save(new_bom)
        
    #Each-Each-Each-Each-Each-Each-Each-----------------------------------------------------------
        start_column=uom_column+1
        startrow=number_of_rows_bom+13
        
        for module in list_of_modules:
            print(startrow)
            print(start_column)
            BS.cell(row=startrow, column=start_column).value="=SUMIFS("+str(get_column_letter(start_column))+"11:"+str(get_column_letter(start_column))+str(10+number_of_rows_bom)+",$I$11:$I$"+str(10+number_of_rows_bom)+",$I$"+str(13+number_of_rows_bom)+")"
            print("=SUMIFS("+str(get_column_letter(start_column))+"11:"+str(get_column_letter(start_column))+str(10+number_of_rows_bom)+",$I$11:$I$"+str(10+number_of_rows_bom)+",$I$"+str(12+number_of_rows_bom)+")")
            start_column=start_column+1
    
        book_bom.save(new_bom)
        
    #Total-Total-Total-Total-Total-Total-Total-Total-Total-Total-Total----------------------------------------------
        start_column=uom_column+1
        startrow=number_of_rows_bom+14
        
        for module in list_of_modules:
            print(startrow)
            print(start_column)
            BS.cell(row=startrow, column=start_column).value="=SUM("+str(get_column_letter(start_column))+"11:"+str(get_column_letter(start_column))+str(10+number_of_rows_bom)+")"
            start_column=start_column+1
        book_bom.save(new_bom)
    
    #Per Length-Per Length-Per Length-Per Length-Per Length---------------------------------------------
        start_column=uom_column+1
        startrow=number_of_rows_bom+12
        for module in list_of_modules:
            print(startrow)
            print(start_column)
            BS.cell(row=startrow, column=start_column).value="=SUMIFS("+str(get_column_letter(start_column))+"11:"+str(get_column_letter(start_column))+str(10+number_of_rows_bom)+",$I$11:$I$"+str(10+number_of_rows_bom)+",$I$"+str(12+number_of_rows_bom)+")"
            print("=SUMIFS("+str(get_column_letter(start_column))+"11:"+str(get_column_letter(start_column))+str(10+number_of_rows_bom)+",$I$11:$I$"+str(10+number_of_rows_bom)+",$I$"+str(12+number_of_rows_bom)+")")
            start_column=start_column+1
        book_bom.save(new_bom)
        
    #Validation-Validation-Validation-Validation-Validation-Validation-Validation-Validation-Validation-Validation-
        start_column=uom_column+1
        startrow=number_of_rows_bom+15
        for module in list_of_modules:
            print(startrow)
            print(start_column)
            BS.cell(row=startrow, column=start_column).value="=("+str(get_column_letter(start_column))+str(number_of_rows_bom+12)+"+"+str(get_column_letter(start_column))+str(number_of_rows_bom+13)+")-"+str(get_column_letter(start_column))+str(number_of_rows_bom+14)
            start_column=start_column+1
        book_bom.save(new_bom)
        
    #Total-Total-Total-Total-Total-Total-Total-Total-Total-Total-Total-Total-Total-Total-Total-Total-Total-Total-Total-Total
        start_column=uom_column+1
        startrow=number_of_rows_bom+16
        for module in list_of_modules:
            print(startrow)
            print(start_column)
            BS.cell(row=startrow, column=start_column).value="="+str(get_column_letter(start_column))+str(number_of_rows_bom+12)+"+("+str(get_column_letter(start_column))+str(number_of_rows_bom+13)+"/"+"1000)"
            start_column=start_column+1
        book_bom.save(new_bom)
        
    #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    ###################################################################################################################@
    ##################SHEET(CONV)-SHEET(CONV)-SHEET(CONV)-SHEET(CONV)-#################################################@ 
    ###################################################################################################################@   
    #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ 
        BSC=book_bom["Detail(Conv)"]
    
    ###################################################################################################################
    ##################Insert formulas titles bottom module matrix table-###############################################
    ###################################################################################################################
        titles_to_insert=["Total","Validation"]
        ir_boom=number_of_rows_bom+12
        for title in titles_to_insert:
            BSC.cell(row=ir_boom, column=uom_column).value=title
            ir_boom=ir_boom+1
        book_bom.save(new_bom)
        
    #Total-Total-Total-Total-Total-Total-Total-Total-Total-Total------------------------------------------------------------
    
        start_column=uom_column+1
        startrow=number_of_rows_bom+12
        for module in list_of_modules:
            BSC.cell(row=startrow, column=start_column).value="=SUM("+str(get_column_letter(start_column))+"11:"+str(get_column_letter(start_column))+str(10+number_of_rows_bom)+")"
            start_column=start_column+1
        book_bom.save(new_bom)
        
    #Validation-Validation-Validation-Validation-Validation-Validation-Validation-Validation-Validation-Validation-Validation-Validation-Validation-Validation
        start_column=uom_column+1
        startrow=number_of_rows_bom+13
        for module in list_of_modules:
            BSC.cell(row=startrow, column=start_column).value="="+str(get_column_letter(start_column))+str(12+number_of_rows_bom)+"-Detail!"+str(get_column_letter(start_column))+str(16+number_of_rows_bom)
            start_column=start_column+1
        book_bom.save(new_bom)
        
    ###################################################################################################################
    ############################ Multiply per lenght elements by 1/1000 ###############################################
    ###################################################################################################################
    
    
    
        # startrow=11
        
        # for row in range(number_of_rows_bom):
        #     cell_obj = BS.cell(row = startrow, column = uom_column)
        #     value=cell_obj.value
        #     start_column=uom_column+1
            
        #     if(str(value)=="Per Length"):
        #         for modulo in list_of_modules:
                    
        #             cell_obj = BS.cell(row = startrow, column = start_column)
        #             value=cell_obj.value
                    
        #             if(str(value)==""):
        #                 BSC.cell(row=startrow, column=start_column).value=float(0)
        #             else:
        #                 BSC.cell(row=startrow, column=start_column).value=float(value)/1000
        
        #             start_column=start_column+1
        #     startrow=startrow+1
    
        # book_bom.save(new_bom)
        
        
    #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    ###################################################################################################################@
    ################## CUT SHEET CUT SHEET CUT SHEET CUT SHEET CUT SHEET CUT SHEET#####################################@ 
    ###################################################################################################################@   
    #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ 
        
    
    
    ###################################################################################################################
    ################## GET NUMBER OF ROWS-GET NUMBER OF ROWS-GET NUMBER OF ROWS     ###################################
    ###################################################################################################################
    
        initialrow_get_number_of_rows=9
        initial_column=1
        
        number_of_rows=0
        while True:
            cell_obj = CS.cell(row = initialrow_get_number_of_rows, column = initial_column)
            border=str(cell_obj.border.bottom.style)
            print(border)
            initialrow_get_number_of_rows+=1
            number_of_rows+=1
            if(border=="None"):
                break
        
    ###################################################################################################################
    ################## GET WIRE TABLE DIMENSIONS- GET WIRE TABLE DIMENSION      ####################################### 
    ###################################################################################################################
    
        initial_column=1
        initialrow=9
        names_columns_to_readadd={
                                   'Term 1 Klima':'Term. 1 Cust',
                                   'Seal 1 Klima':'Seal 1 Cust',
                                   'Node 1 Klima':'Connector 1 Cust. PN',
                                   'Term 2 Klima':'Term. 2 Cust',
                                   'Seal 2 Klima':'Seal 2 Cust',
                                   'Node 2Klima':'Connector 2 Cust. PN'
                                   }
        
        list_of_modules=[]
        list_columns_to_apply_formula=[]
        list_of_nodes=[]
        list_number_column_node=[]
        limit=6
        column_counter=1                           
        while True:
            cell_obj = CS.cell(row = initialrow, column = column_counter)
            value=cell_obj.value
            border=str(cell_obj.border.bottom.style)
            print(column_counter)
            print(str(value))
            
            if(value!=None):
                print(value[:4])
                
            column_name=names_columns_to_readadd.get(str(value))
            print(column_name)
            if(column_name!=None):
                print(column_name)
                CS.insert_cols(column_counter+1)
                CS.cell(row=initialrow, column=column_counter+1).value=column_name
                CS.cell(row=initialrow, column=column_counter+1).border=cell_border
                CS.cell(row=initialrow, column=column_counter+1).font = fontorange
                CS.cell(row=initialrow, column=column_counter+1).fill=my_fillyellow
                CS.cell(row=initialrow, column=column_counter+1).alignment = alignment
                
                for numero in range (number_of_rows-2):
                    CS.cell(row=initialrow+numero+1, column=column_counter+1).value="=IFERROR(VLOOKUP($"+str(get_column_letter(column_counter-1))+str(initialrow+numero+1)+",'"+str(path)+"\Modified files"+"\["+str(new_bom)+"]Detail'!$A$11:$E$"+str(number_of_rows_bom+11)+",5,0),"")"
 
                    CS.cell(row=initialrow+numero+1, column=column_counter+1).border=cell_border
                book.save(file_removed+"-copy.xlsx")
                limit-=1

            #time.sleep(500)
            if(value=="Circuit Klima"):
                column_circuit_klima_1=column_counter
            
            if(value!=None):
                if(value[0]=="A"):
                    list_of_modules.append(value)
                    list_columns_to_apply_formula.append(column_counter)
            
            if(value!=None):
                if(value[:4]=="Node" and str(value[-1])!="a"):
                    print()
                    column_circuit_klima=column_counter
                    temp_row=initialrow+1
                    list_number_column_node.append(column_counter)
                    for numero in range (number_of_rows-2):
                        cell_obj = CS.cell(row = temp_row, column = column_counter)
                        tempvalue=cell_obj.value
                        list_of_nodes.append(tempvalue)
                        temp_row+=1
                    
            column_counter+=1
    
            if(border=="None"):
                break
            
        print("ya termine")
        print(column_circuit_klima)
        
        print(list_columns_to_apply_formula)
        print(list_of_modules)
        
    ###################################################################################################################
    ################## Minitable sum wires Klima & non Klima - Totals ######### ####################################### 
    ###################################################################################################################
    
    #Write mini list of titles
    
        titles=["Total","Klima","NON-KLIMA"]
        column_to_write_titles=int(list_columns_to_apply_formula[0])-1
        start_row_to_write_titles=number_of_rows+10
        
        print(start_row_to_write_titles)
        print(column_to_write_titles)
        temp_row=start_row_to_write_titles
        for title in titles:
            CS.cell(row=temp_row, column=column_to_write_titles).value=title
            temp_row+=1
             
        
    #Write formulas
       #Total
        temp_row=start_row_to_write_titles
        temp_column=column_to_write_titles+1
        
        validation_row_nodes=temp_row
        for number in list_columns_to_apply_formula:
            name_column=get_column_letter(number)
            CS.cell(row=temp_row, column=temp_column).value="=SUM("+name_column+"10"+":"+name_column+str(7+number_of_rows)+")"
            print("=SUM("+name_column+"10"+":"+name_column+str(10+number_of_rows)+")")
            temp_column+=1
            
        #Klima
        temp_row=start_row_to_write_titles+1
        temp_column=column_to_write_titles+1
        print(list_columns_to_apply_formula)
        for number in list_columns_to_apply_formula:
            name_column=get_column_letter(number)
            CS.cell(row=temp_row, column=temp_column).value="=SUMIFS("+name_column+"10"+":"+name_column+str(7+number_of_rows)+","+"$"+str(get_column_letter(column_circuit_klima_1))+"$"+"10:$"+str(get_column_letter(column_circuit_klima_1))+"$"+str(7+number_of_rows)+","+"$"+str(get_column_letter(column_to_write_titles))+"$"+str(start_row_to_write_titles+1)+")"
            print("=SUMIFS("+name_column+"10"+":"+name_column+str(7+number_of_rows)+","+"$"+str(get_column_letter(column_circuit_klima_1))+"$"+"10:$"+str(get_column_letter(column_circuit_klima))+"$"+str(7+number_of_rows)+","+"$"+str(get_column_letter(column_to_write_titles))+"$"+str(start_row_to_write_titles+1)+")")
           
            temp_column+=1
            
        #NON-KLIMA
        temp_row=start_row_to_write_titles+2
        temp_column=column_to_write_titles+1
        print(list_columns_to_apply_formula)
        for number in list_columns_to_apply_formula:
            name_column=get_column_letter(number)
            CS.cell(row=temp_row, column=temp_column).value="=SUMIFS("+name_column+"10"+":"+name_column+str(7+number_of_rows)+","+"$"+str(get_column_letter(column_circuit_klima_1))+"$"+"10:$"+str(get_column_letter(column_circuit_klima_1))+"$"+str(7+number_of_rows)+","+"$"+str(get_column_letter(column_to_write_titles))+"$"+str(start_row_to_write_titles+2)+")"
            print("=SUMIFS("+name_column+"10"+":"+name_column+str(7+number_of_rows)+","+"$"+str(get_column_letter(column_circuit_klima_1))+"$"+"10:$"+str(get_column_letter(column_circuit_klima))+"$"+str(7+number_of_rows)+","+"$"+str(get_column_letter(column_to_write_titles))+"$"+str(start_row_to_write_titles+2)+")")
           
            temp_column+=1
            
        
        list_of_nodes = list(dict.fromkeys(list_of_nodes))
        list_of_nodes = sorted(list_of_nodes)
            
    #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    ###################################################################################################################@
    ################## NODES SHEET NODES SHEET NODES SHEET NODES SHEET#################################################@ 
    ###################################################################################################################@   
    #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ 
        
        book.create_sheet('Nodes', index=2)
        
        initialrow=1
        initialcolumn=1
        
        
        CSN=book["Nodes"]
        
        CSN.cell(row=initialrow, column=initialcolumn).value="Nodes"
        CSN.cell(row=initialrow, column=initialcolumn).border=cell_border
        
        CSN.cell(row=initialrow, column=initialcolumn+1).value="KLIMA"
        CSN.cell(row=initialrow, column=initialcolumn+1).border=cell_border
        
        
        tempcolumn=initialcolumn+2
        for modulo in list_of_modules:
            CSN.cell(row=initialrow, column=tempcolumn).value=modulo
            CSN.cell(row=initialrow, column=tempcolumn).border=cell_border
            CSN.cell(row=initialrow, column=tempcolumn).font = fontorange
            CSN.cell(row=initialrow, column=tempcolumn).fill=my_fillyellow
            CSN.cell(row=initialrow, column=tempcolumn).alignment = alignment
            tempcolumn+=1
            
        initialrow=1
        initialcolumn=1
        
        list_number_of_rows_CSN=[]
        temprow=initialrow+1
        
        for node in list_of_nodes:
            CSN.cell(row=temprow, column=initialcolumn).value=node
            list_number_of_rows_CSN.append(temprow)
            temprow+=1
        
        
        list_rows_2nd_part=[]
        list_temp_values=["Total","validation"," ","TOTAL ckts","KLIMA","NON-KLIMA"]
        for valor in list_temp_values:
            CSN.cell(row=temprow, column=initialcolumn).value=valor
            list_rows_2nd_part.append(temprow)
            temprow+=1
            
        def make_formula(column_wire_sheet,number_row,list_number_column_node,number_of_rows):
            rows=str(7+number_of_rows)
            formula="=SUMPRODUCT(--(EXACT($A"+str(number_row)+",'Wire Sheet'!$"+str(get_column_letter(list_number_column_node[0]))+"$10"+":$"+str(get_column_letter(list_number_column_node[0]))+"$"+rows+")),'Wire Sheet'!"+str(get_column_letter(column_wire_sheet))+"$10"+":"+str(get_column_letter(column_wire_sheet))+"$"+rows+")+SUMPRODUCT(--(EXACT($A"+str(number_row)+",'Wire Sheet'!$"+str(get_column_letter(list_number_column_node[1]))+"$10"+":$"+str(get_column_letter(list_number_column_node[1]))+"$"+rows+")),'Wire Sheet'!"+str(get_column_letter(column_wire_sheet))+"$10"+":"+str(get_column_letter(column_wire_sheet))+"$"+rows+")"
            return(str(formula))
        
        initialrow=2
        initialcolumn=3
        
        for row in list_number_of_rows_CSN:
            initialcolumn=3
            for column in list_columns_to_apply_formula:
                formula=make_formula(column, row, list_number_column_node, number_of_rows)
                CSN.cell(row=initialrow, column=initialcolumn).value=formula
                initialcolumn+=1
            initialrow+=1
            
        print(list_rows_2nd_part)
        
        initial_column=3
        initialrow=list_rows_2nd_part[0]
        
    #Write formulas second part    
        #Total
    
        for modulo in list_of_modules:
            CSN.cell(row=initialrow, column=initial_column).value="=SUM("+str(get_column_letter(initial_column))+"2"+":"+str(get_column_letter(initial_column))+str(int(len(list_of_nodes))+1)+")"
            initial_column+=1
            
        #validation borderline
        
        initial_column=1
        for space in range(len(list_of_modules)+2):
            CSN.cell(row=list_rows_2nd_part[0], column=space+1).border = Border(top = Side(border_style='thin', color='FF000000'))
            
        #Validation
        
        initial_column=3
        initialrow=list_rows_2nd_part[1]
        
        for modulo in list_of_modules:
            CSN.cell(row=initialrow, column=initial_column).value="=("+str(get_column_letter(initial_column))+str(list_rows_2nd_part[0])+"/2)-'Wire Sheet'!"+str(get_column_letter(list_columns_to_apply_formula[initial_column-3]))+str(validation_row_nodes)
            initial_column+=1
                                    
            
         #TOTAL ckts
         
        initial_column=3
        initialrow=list_rows_2nd_part[3]
        
        for modulo in list_of_modules:
            CSN.cell(row=initialrow, column=initial_column).value="='Wire Sheet'!"+str(get_column_letter(list_columns_to_apply_formula[initial_column-3]))+str(validation_row_nodes)
            initial_column+=1
            
            #BORDERLINE
            
        initial_column=1
        for space in range(len(list_of_modules)+2):
            CSN.cell(row=list_rows_2nd_part[3], column=space+1).border = Border(bottom = Side(border_style='thin', color='FF000000'))
            
        #KLIMA
        
        initial_column=3
        initialrow=list_rows_2nd_part[4]
        
        for modulo in list_of_modules:
            CSN.cell(row=initialrow, column=initial_column).value="='Wire Sheet'!"+str(get_column_letter(list_columns_to_apply_formula[initial_column-3]))+str(validation_row_nodes+1)
            initial_column+=1
            
        #NON-KLIMA
        
        initial_column=3
        initialrow=list_rows_2nd_part[5]
        
        for modulo in list_of_modules:
            CSN.cell(row=initialrow, column=initial_column).value="='Wire Sheet'!"+str(get_column_letter(list_columns_to_apply_formula[initial_column-3]))+str(validation_row_nodes+2)
            initial_column+=1
         
        print(column_circuit_klima_1)
        print("done23")
        book.save(file_removed+"-copy.xlsx")
        
        print(list_of_files)
        
    # #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    # ###################################################################################################################@
    # ##################READ BOM READ BOM READ BOM READ BOM READ BOM REA#################################################@ 
    # ###################################################################################################################@   
    # #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ 
    
    
    #     #Get full name of BOM file--------------------------------------------------------------------------------
    #     for file in list_of_files:
    #         print(file[-8:])
    #         if(str(file[-8:])=="BOM.xlsx"):
    #             BOM=file
        
    #     print(BOM)
        
    #     book_bom = openpyxl.load_workbook(BOM, data_only=True)
    #     BS =book_bom["Detail"]
        
    ###################################################################################################################
    ############################ Create directory create directory crea ###############################################
    ###################################################################################################################
        
        
        path = os.path.join(directorio,"Modified files")
        os.makedirs(path) 
        
    
        shutil.move(str(directorio)+'/'+str(new_bom),str(path)+'/'+str(new_bom))
        shutil.move(str(directorio)+'/'+str(file_removed)+"-copy.xlsx",str(path)+'/'+str(file_removed)+"-copy.xlsx")

def ODM():
    
    #------------------------------------------------------------------------------#      
    #Select program folder and get paths of calculations and Technical Information #
    #------------------------------------------------------------------------------# 
    
    
    pyautogui.alert(text='Select program folder', title='Select File', button='OK')
    
    path2 = askdirectory(title='Select Folder') # shows dialog box and return the path
    
    #print(path2)
    
    list_of_directories=os.listdir(path2)
    
    #print(list_of_directories)
    
    for folder in list_of_directories:
        print(str(folder))
        if ("Calculation" in str(folder)):
            calculation_folder=str(folder)
            path_calculations_folder=path2+"/"+folder
            
        if ("Technical" in str(folder)):
            TI_folder=str(folder)
            path_TI=path2+"/"+folder
            
    ####################      
    #ADD ERROR HANDLER #
    ####################  
    
    #print("calculations path")
    #print(calculation_folder)
    #print(path_calculations_folder)
    #print("TI PATH")
    #print(TI_folder)
    #print(path_TI)
    
    #------------------------------------------------------------------------------#      
    #Read Tracking                                                                 #
    #------------------------------------------------------------------------------# 
    
    os.chdir(path_calculations_folder)
    list_of_directories=os.listdir(path_calculations_folder)
    
    #print(list_of_directories)
    
    for folder in list_of_directories:
        #print(str(folder))
        if ("Tracking" in str(folder)):
            tracking_folder=str(folder)
            path_tracking_folder=path_calculations_folder+"/"+folder
            
    #print(tracking_folder)
    #print(path_tracking_folder)
    
    
    #@------#working on tracking folder#------@#
    
    os.chdir(path_tracking_folder)
    list_of_directories=os.listdir(path_tracking_folder)
    
    #print(list_of_directories)
    
    for folder in list_of_directories:
        #print(str(folder))
        if (("Tracking" in str(folder)) and ((".xlsx") in str(folder)) and ("~$" not in str(folder) ) ):
            tracking_file=str(folder)
            
    #print(str(tracking_file))
    
    #Reading tracking file
    

    File=[]
    PartNumber=[]
    Family=[]
    
    book = openpyxl.load_workbook(tracking_file, data_only=True)
    HL=book["Tracking"]
    
    columns_to_read=[3,4,5]
    lists_to_fill=[File, PartNumber, Family]
    
    counter=0
    for lista in lists_to_fill:
        border="thin"
        rowinicial=6
        current_column=columns_to_read[counter]
        while(border=="thin"):
            cell_obj = HL.cell(row = rowinicial, column = current_column)
            value=cell_obj.value
            border=str(cell_obj.border.bottom.style)

            if(border=="thin"):
                lista.append(value)
            rowinicial=rowinicial+1
        counter=counter+1
            
    print(File)
    print(PartNumber)
    print(Family)
    
    #------------------------------------------------------------------------------#      
    #Create Folders, BOM & LMI DOCUMENTS                                           #
    #------------------------------------------------------------------------------# 
    
    os.chdir(path_calculations_folder)
    
    counter=0
    columns_to_read=[1,3,4,5,8]
    
    columns_to_read_ne=[1,2,3,4,5,6,7,8,9,10,
                        11,12,13,14,15,16,17,
                        18,19,20,21,22,23,24,
                        25,26,27]
    
    columns_to_read_oo=[1,2,3,4,5,6,7,8,9,10,
                        11,12,13,14,15,16,17,
                        18,19,20,21,22,23,24,
                        25,26,27,28,29,30,31,
                        32,33,34]

    for name in PartNumber:
        
        Number=[]
        Name=[]
        ComponentName=[]
        Description=[]
        Quantity=[]
        Control=[]
        
        lists_to_fill=[Number,Name,ComponentName,Description,Quantity,Control]

        
        os.chdir(path_calculations_folder)
        
        if(str(File[counter])!="None"):
            directory_name=str(name)+"_"+str(Family[counter])
            requested=str(name)
            os.chdir(path_calculations_folder)
            os.mkdir(directory_name)
            lmi_name="LMI_BLANK.xlsx"
            
            #------------------------------------------------------------------------------#      
            #Copy BOM To each folder                                                       #
            #------------------------------------------------------------------------------# 
            
            # Source path
        
            source =path_TI+"/"+str(File[counter])+".xlsx"
            Source_LMI='''C:/Users/LCardenasMontaz/Desktop/LMI/'''+str(lmi_name)
            
             
            # Destination path
            destination = path_calculations_folder+"/"+directory_name+"/"+str(File[counter])+".xlsx"
            destination2 = path_calculations_folder+"/"+directory_name+"/"+directory_name+".xlsx" 
            destination3= path_calculations_folder+"/"+directory_name+"/"+directory_name+"_LMI"+".xlsx"
            # Copy the content of
            # source to destination
             
            try:
                shutil.copy(source, destination)
                shutil.copy(source, destination2)
                shutil.copy(Source_LMI, destination3)
                print("File copied successfully.")
             
            # If source and destination are same
            except shutil.SameFileError:
                print("Source and destination represents the same file.")
             
            # If there is any permission issue
            except PermissionError:
                print("Permission denied.")
             
            # For other errors
            except:
                print("Error occurred while copying file.")
            
            os.chdir(path_calculations_folder+"/"+directory_name)
                 
    
            book_bom = openpyxl.load_workbook(directory_name+".xlsx")
            
            print(directory_name)
            
            #@------#Read old sheets#------@#
            BS=book_bom["Assembly Nav"]
            NEO=book_bom["Netlist Extraction"]
            OO=book_bom["Overstocks"]
            HCO=book_bom["Harness Calculation"]
            
            #@------#Create new sheets#------@#
            AN = book_bom.create_sheet("Assembly Nav costed",1)
            AN.title = "Assembly Nav costed"
            
            SCC = book_bom.create_sheet("Summary Component count",2)
            SCC.title = "Summary Component count"
            
            NEN = book_bom.create_sheet("Netlist Extraction costed",4)
            NEN.title = "Netlist Extraction costed"
            
            ON = book_bom.create_sheet("Overstocks costed",7)
            ON.title = "Overstocks costed"
            
            #print(directory_name+"/"+directory_name+".xlsx")
            number_of_rows=BS.max_row
            number_of_rows_neo=NEO.max_row
            number_of_rows_OO=OO.max_row
            
            #------------------------------------------------------------------------------#      
            #Read Assembly nav original                                                    #
            #------------------------------------------------------------------------------# 
            
            
            #@------#identify requested PN column number#------@#
            
            control="start"
            columnainicial=1
            while control != requested:
                cell_obj = BS.cell(row = 1, column = columnainicial)
                control=str(cell_obj.value)
                columnainicial=columnainicial+1
                control_number_column=columnainicial-1
                if(requested==value):
                    break
            
            #print("CONTROL NUMBER COLUMN!!!!!!!")
            #print(control_number_column)
            
            #@------#Read Columns#------@#
            
            counter2=0
            columns_to_read=[1,3,4,5,8,control_number_column]
            for lista in lists_to_fill:
                current_column=columns_to_read[counter2]
                rowinicial=2
                #print(current_column)
                for cell in range(number_of_rows-1):
                    cell_obj = BS.cell(row = rowinicial, column = current_column)
                    value=str(cell_obj.value)
                    lista.append(value)
                    rowinicial=rowinicial+1
                counter2+=1
                
            #------------------------------------------------------------------------------#      
            #Read and write Netlist extr original and new                                  #
            #------------------------------------------------------------------------------# 
            
            #@------#get list of harness number elements#------@#
            listaHNNEO=[]
            rowinicial=2
            for cell in range(number_of_rows_neo-1):
                cell_obj = NEO.cell(row = rowinicial, column = 18)
                value=str(cell_obj.value)
                listaHNNEO.append(value)
                rowinicial=rowinicial+1
            
            #@------#read row from original netlist and write row in new netlist if valor==requested#------@#
            rowinicial=2
            rowinicial2=2
            counterNEO=0
            list_from_to_conn=[]
            list_from_to_term=[]
            
            twist_names=[]
            wyre_type=[]
            Cable_id=[]
            Shield_Grp=[]
            wyre_length=[]
            
            for valor in listaHNNEO:
                if(str(valor)==requested):
                    
                    temp_container=[]
                    for columna in columns_to_read_ne:
                        cell_obj = NEO.cell(row = rowinicial, column = columna)
                        value=str(cell_obj.value)
                        temp_container.append(value)
                    
                    for columna in columns_to_read_ne:
                        
                        if(str(temp_container[columna-1])!="None"):
                            if(int(columna-1)==9 or int(columna-1)==11):
                                NEN.cell(row=rowinicial2, column=columna).value=float(temp_container[columna-1])
                            else:
                                NEN.cell(row=rowinicial2, column=columna).value=temp_container[columna-1]
                            
                            #@------#get information for shield and twist table#------@#
                            #if(columna==12):
                                #wyre_length.append(temp_container[columna-1])
                            if(columna==12):
                                wyre_length.append(temp_container[columna-1])
                            if(columna==14):
                                wyre_type.append(temp_container[columna-1])
                            if(columna==23):
                                Cable_id.append(temp_container[columna-1])
                            if(columna==24):
                                twist_names.append(temp_container[columna-1])
                            if(columna==25):
                                Shield_Grp.append(temp_container[columna-1])

    
                            
                            #@------#get from to, to conn data and store in list#------@#
                            if(columna==4 or columna==7 ):
                                list_from_to_conn.append(temp_container[columna-1])
                            #@------#get From Term, to Term data and store in list#------@#
                            if(columna==6 or columna==9 ):
                                list_from_to_term.append(temp_container[columna-1])
                        else:
                            if(columna==23):
                                Cable_id.append(temp_container[columna-1])
                            if(columna==24):
                                twist_names.append(temp_container[columna-1])
                            if(columna==25):
                                Shield_Grp.append(temp_container[columna-1])
                                
                    counterNEO=counterNEO+1
                    rowinicial2=rowinicial2+1 
                rowinicial=rowinicial+1
                    
            counterNEO=counterNEO+1
            


            
            #@------#compute unique twist pairs#------@#
            list_twist_information=[]
            list_twist_filtered=[]
            twist_names_list=[]
            twist_names_list_unique=[]
            dict_twist={}
            for data in range(len(wyre_length)):
                              temp_list=[]
                              temp_list.append(wyre_length[data])
                              temp_list.append(wyre_type[data])
                              temp_list.append(Cable_id[data])
                              temp_list.append(twist_names[data])
                              temp_list.append(Shield_Grp[data])
                              list_twist_information.append(temp_list)
            
            list_assembly_wire_type=["Airbag","A2B","Coax","Ethernet","USB","R2PP","LVDS"]
            
            for element in list_twist_information:

                assemblywire_test=any(wiretype.casefold() in str(element[1]).casefold() for wiretype in list_assembly_wire_type)

                if(assemblywire_test==False and str(element[2])=="None" and str(element[4])=="None" and str(element[3])!="None"):
                    list_twist_filtered.append(element)
                    twist_names_list.append(element[3])
            

            twist_names_list_unique=twist_names_list
            twist_names_list_unique = list(dict.fromkeys(twist_names_list_unique))
            
            for element in twist_names_list_unique:
                value=twist_names_list.count(element)
                dict_twist[element] = [value]
            #dict_twist = {i:twist_names_list.count(i) for i in twist_names_list}
            
        
            for key in dict_twist:

                list_temp=[]
                for element in list_twist_filtered:

                    if(str(key)==str(element[3])):
                        list_temp.append(float(element[0]))
                result = all(element == list_temp[0] for element in list_temp)
                if (result):
                    dict_twist[str(key)].append(list_temp[0])
            print(dict_twist)
                    
            
        

            #@------#remove duplicated values from list#------@#
            number_rows=len(list_from_to_conn)/2
            
            #list used to get number of splices
            list_to_use_splices=list_from_to_conn
            
            list_from_to_conn = list(dict.fromkeys(list_from_to_conn))
            list_from_to_conn.sort()
            
            list_from_to_term = list(dict.fromkeys(list_from_to_term))
            #list_from_to_term.remove("None")
            list_from_to_term.append("Sn")
            list_from_to_term.sort()


            
            #@------#write headers in new netlist#------@#
            list_of_headers=["Schematic Sub-System", "Schematic Wire Handle","Circuit #","From Conn",
                             "From Pin ID","From Term","To Conn","To Pin ID","To Term","Wire Guage",
                             "Wire Color","Wire Length","IVED Harness Name","Wire Type","W_spec",
                             "Circuit Type","RPO Code","Harness #","","","Wire O.D.","","Cable ID",
                             "Twist Grp","Shield Grp","",""]
            
            for columna in columns_to_read_ne:
                NEN.cell(row=1, column=columna).value=list_of_headers[columna-1]
                NEN.cell(row=1, column=columna).font = fontwhite
                NEN.cell(row=1, column=columna).fill=my_filldarkgray
            
            #@------#write headers of seconday table#------@#
            list_of_headers_secondary_table=["ID","from","to","Insertions/ends","Description",
                                              "ways","Conn Type","Seals","Plugs","Count","Heat shrink"]
            
            for columna in range(len(list_of_headers_secondary_table)):
                NEN.cell(row=counterNEO+4, column=columna+1).value=list_of_headers_secondary_table[columna]
                NEN.cell(row=counterNEO+4, column=columna+1).font = fontwhite
                NEN.cell(row=counterNEO+4, column=columna+1).fill=my_filldarkgray
                
            start_column_term=12
            for terminaltype in list_from_to_term:
                NEN.cell(row=counterNEO+4, column=start_column_term).value=str(terminaltype)+" Terminal count"
                NEN.cell(row=counterNEO+4, column=start_column_term).font = fontwhite
                NEN.cell(row=counterNEO+4, column=start_column_term).fill=my_filldarkgray

                
                row_term_iter=counterNEO+5
                for value in range(len(list_from_to_conn)):
                    if(str(terminaltype)=="Sn"):
                        terminaltype=""
                    NEN.cell(row=row_term_iter, column=start_column_term).value='''=IF(J'''+str(row_term_iter)+'''="x",COUNTIFS(D2:D'''+str(int(number_rows+1))+''',A'''+str(row_term_iter)+''',F2:F'''+str(int(number_rows+1))+''',"'''+str(terminaltype)+'''")+COUNTIFS(G2:G'''+str(int(number_rows+1))+''',A'''+str(row_term_iter)+''',I2:I'''+str(int(number_rows+1))+''',"'''+str(terminaltype)+'''"),0)'''
                    row_term_iter=row_term_iter+1
                
                start_column_term=start_column_term+1
                
            counterNEO_secondarytable_start=counterNEO+5
            counterNEO_secondarytable_start_record=counterNEO_secondarytable_start
            

                
            #@------#write contents of secondary table [non repeated connector]#------@#
            for value in range(len(list_from_to_conn)):
                current_row=counterNEO_secondarytable_start
                NEN.cell(row=current_row, column=1).value=list_from_to_conn[value]
                NEN.cell(row=current_row, column=2).value='=COUNTIF($D$2:$D$'+str(counterNEO)+',A'+str(current_row)+')'
                NEN.cell(row=current_row, column=3).value='=COUNTIF($G$2:$G$'+str(counterNEO)+',A'+str(current_row)+')'
                NEN.cell(row=current_row, column=4).value='=B'+str(current_row)+'+C'+str(current_row)
                
                if(list_from_to_conn[value][0:3]!="STD"):
                    NEN.cell(row=current_row, column=5).value="=VLOOKUP(A"+str(current_row)+",'Assembly Nav costed'!J3:K5000,2,0)"
                    NEN.cell(row=current_row, column=6).value='''=IF(J'''+str(current_row)+'''="x",MID(E'''+str(current_row)+''',SEARCHB(",",E'''+str(current_row)+''',(SEARCHB(",",E'''+str(current_row)+''',(SEARCHB(",",E'''+str(current_row)+''',1)+1))+1))+1,(SEARCHB(",",E'''+str(current_row)+''',(SEARCHB(",",E'''+str(current_row)+''',(SEARCHB(",",E'''+str(current_row)+''',(SEARCHB(",",E'''+str(current_row)+''',1)+1))+1))+1)) -SEARCHB(",",E'''+str(current_row)+''',(SEARCHB(",",E'''+str(current_row)+''',(SEARCHB(",",E'''+str(current_row)+''',1)+1))+1))-4)),"")'''
                    NEN.cell(row=current_row, column=7).value='''=IF(J'''+str(current_row)+'''="x",MID(E'''+str(current_row)+''',SEARCHB(",",E'''+str(current_row)+''',(SEARCHB(",",E'''+str(current_row)+''',1)+1))+1,(SEARCHB(",",E'''+str(current_row)+''',(SEARCHB(",",E'''+str(current_row)+''',(SEARCHB(",",E'''+str(current_row)+''',1)+1))+1))-SEARCHB(",",E'''+str(current_row)+''',(SEARCHB(",",E'''+str(current_row)+''',1)+1)))-1),"")'''
                    NEN.cell(row=current_row, column=8).value='=IF(AND(G'+str(current_row)+'="SLD",J'+str(current_row)+'="x"),D'+str(current_row)+',0)'
                    NEN.cell(row=current_row, column=9).value='=IF(AND(G'+str(current_row)+'="SLD",J'+str(current_row)+'="x"),(F'+str(current_row)+'-D'+str(current_row)+'),0)'
                    NEN.cell(row=current_row, column=10).value='x'
                
                if(list_from_to_conn[value][0:3]=="STD"):
                    NEN.cell(row=current_row, column=5).value='''=IF(G'''+str(current_row)+'''="GENERIC SPLICE SEALED",IF(D'''+str(current_row)+'''<5,"HTSHRNK TUBE,DUAL-WALL,50 MM LONG,BLK,BLK,110 C,XLPO,6 I.D.,774185P002",IF(AND(D'''+str(current_row)+'''>=5,D'''+str(current_row)+'''<=9),"HTSHRNK TUBE,DUAL-WALL,50 MM LONG,BLK,BLK,125 C,XLPO,11 I.D.,QSZH-125-NR3","")),"Splice Tape")'''
                    NEN.cell(row=current_row, column=7).value='''=_xlfn.XLOOKUP(A'''+str(current_row)+''','Assembly Nav costed'!$J$1:$J$5000,'Assembly Nav costed'!$I$1:$I$5000,"NOT FOUND",0)'''
                    NEN.cell(row=current_row, column=11).value='''=IF(AND(G'''+str(current_row)+'''<>"GENERIC SPLICE SEALED",G'''+str(current_row)+'''<>"GENERIC SPLICE UNSEALED"),"What is this?",IF(G'''+str(current_row)+'''="GENERIC SPLICE SEALED",IF(D'''+str(current_row)+'''<5,"3299362G2",IF(AND(D'''+str(current_row)+'''>=5,D'''+str(current_row)+'''<=9),"3299362M9","3299365JY")),"26-4412-BC"))'''
                counterNEO_secondarytable_start=counterNEO_secondarytable_start+1
                
            counterNEO_secondarytable_end=current_row
            
            
            #@------#Write seals, plugs and terminals table#------@#
            
            list_seals_plugs_terminals=[]
            
            #Headers#
            NEN.cell(row=counterNEO_secondarytable_end+2, column=1).value="Seals, Plugs & Terminals"
            NEN.cell(row=counterNEO_secondarytable_end+2, column=1).font = fontwhite
            NEN.cell(row=counterNEO_secondarytable_end+2, column=1).fill=my_filldarkgray
            
            NEN.cell(row=counterNEO_secondarytable_end+3, column=1).value="Cavity Plug_"+str(directory_name)
            NEN.cell(row=counterNEO_secondarytable_end+4, column=1).value="Wire Seal_"+str(directory_name)

            NEN.cell(row=counterNEO_secondarytable_end+2, column=2).value="QTY"
            NEN.cell(row=counterNEO_secondarytable_end+2, column=2).font = fontwhite
            NEN.cell(row=counterNEO_secondarytable_end+2, column=2).fill=my_filldarkgray
            
            list_seals_plugs_terminals.append("Cavity Plug_"+str(directory_name))
            list_seals_plugs_terminals.append("Wire Seal_"+str(directory_name))
            
            NEN.cell(row=counterNEO_secondarytable_end+3, column=2).value='=SUM(I'+str(counterNEO_secondarytable_start_record)+':I'+str(counterNEO_secondarytable_end)+')'
            NEN.cell(row=counterNEO_secondarytable_end+4, column=2).value='=SUM(H'+str(counterNEO_secondarytable_start_record)+':H'+str(counterNEO_secondarytable_end)+')'
            
            start_row_termtype=counterNEO_secondarytable_end+5
            start_column_termtype_formula=12
            for typeterminal in list_from_to_term:
                NEN.cell(row=start_row_termtype, column=1).value='Terminal_WP_'+str(typeterminal)+'_'+str(directory_name)
                NEN.cell(row=start_row_termtype+1, column=1).value='Terminal_NWP_'+str(typeterminal)+'_'+str(directory_name)
                
                list_seals_plugs_terminals.append('Terminal_WP_'+str(typeterminal)+'_'+str(directory_name))
                list_seals_plugs_terminals.append('Terminal_NWP_'+str(typeterminal)+'_'+str(directory_name))
                
                NEN.cell(row=start_row_termtype, column=2).value='''=SUMIFS('''+str(get_column_letter(start_column_termtype_formula))+str(counterNEO_secondarytable_start_record)+''':'''+str(get_column_letter(start_column_termtype_formula))+str(counterNEO_secondarytable_end)+''',G'''+str(counterNEO_secondarytable_start_record)+''':G'''+str(counterNEO_secondarytable_end)+''',"SLD")'''
                NEN.cell(row=start_row_termtype+1, column=2).value='''=SUMIFS('''+str(get_column_letter(start_column_termtype_formula))+str(counterNEO_secondarytable_start_record)+''':'''+str(get_column_letter(start_column_termtype_formula))+str(counterNEO_secondarytable_end)+''',G'''+str(counterNEO_secondarytable_start_record)+''':G'''+str(counterNEO_secondarytable_end)+''',"UNSLD")'''
                
                start_row_termtype=start_row_termtype+2
                start_column_termtype_formula=start_column_termtype_formula+1
            
            #@------#Write pure twist table#------@#
            start_row_twist=start_row_termtype+1
            
            if(len(twist_names_list)>0): 
                #Headers#
                NEN.cell(row=start_row_twist, column=1).value="Twist (solo)"
                NEN.cell(row=start_row_twist, column=1).font = fontwhite
                NEN.cell(row=start_row_twist, column=1).fill=my_filldarkgray
                
                NEN.cell(row=start_row_twist, column=2).value="Number of wires"
                NEN.cell(row=start_row_twist, column=2).font = fontwhite
                NEN.cell(row=start_row_twist, column=2).fill=my_filldarkgray
                
                NEN.cell(row=start_row_twist, column=3).value="Length"
                NEN.cell(row=start_row_twist, column=3).font = fontwhite
                NEN.cell(row=start_row_twist, column=3).fill=my_filldarkgray
                
                start_row_twist=start_row_twist+1
                
                for key in dict_twist:
                    NEN.cell(row=start_row_twist, column=1).value=str(key)
                    contents=dict_twist.get(key)
                    NEN.cell(row=start_row_twist, column=2).value=int(contents[0])
                    NEN.cell(row=start_row_twist, column=3).value=float(contents[1])
                    start_row_twist=start_row_twist+1
                    
                
        
            #------------------------------------------------------------------------------#      
            #Read Overstocks original                                                      #
            #------------------------------------------------------------------------------# 
            
            #@------#identify requested PN column number#------@#
            
            list_overstocks=[]
            
            control="start"
            columnainicial=1
            while control != requested:
                cell_obj = OO.cell(row=3, column = columnainicial)
                control=str(cell_obj.value)
                columnainicial=columnainicial+1
                control_number_column_OS=columnainicial-1
                if(requested==value):
                    break
                
            #print("CONTROL NUMBER COLUMN OVERSTOCKS!!!!!!!")
            #print(control_number_column_OS)
            
            control_list_oo=[]
            rowinicial=4
            for row in range(number_of_rows_OO-3):
                cell_obj = OO.cell(row = rowinicial, column=control_number_column_OS )
                value=str(cell_obj.value)
                control_list_oo.append(value)
                rowinicial=rowinicial+1
                
            #@------#read row from original overstocks and write row in new overstocks if valor==requested#------@#
            rowinicial=4
            for valor in control_list_oo:
                if(str(valor)=="X"):
                    temp_container=[]
                    for columna in columns_to_read_oo:
                        cell_obj = OO.cell(row = rowinicial, column = columna)
                        value=str(cell_obj.value)

                        temp_container.append(value)
                    
                    for columna in columns_to_read_oo:
                        if(str(temp_container[columna-1])!="None"):
                            if(columna==1):
                                list_overstocks.append(temp_container[columna-1])
                            ON.cell(row=rowinicial+1, column=columna).value=temp_container[columna-1]

                    rowinicial=rowinicial+1
                
            #@------#add columns on overstocks costed#------@#
            ON.insert_cols(3)
            ON.insert_cols(5)
            ON.insert_cols(27)    
            
            #@------#ADD headers to overstocks costed worksheet#------@#
            temp_container=[]
            for columna in columns_to_read_oo:
                cell_obj = OO.cell(row = 3, column = columna)
                if(columna==3):
                    temp_container.append("Labor")
                    
                if(columna==4):
                    temp_container.append("Labor2")
                    
                if(columna==25):
                    temp_container.append("OD Inches")
                    
                value=str(cell_obj.value)
                temp_container.append(value)
                
            
            
            for columna in range(len(columns_to_read_oo)+3):
                        if(str(temp_container[columna-1])!="None"):
                            ON.cell(row=4, column=columna+1).value=temp_container[columna]
                            ON.cell(row=4, column=columna+1).font = fontwhite
                            ON.cell(row=4, column=columna+1).fill=my_filldarkgray
                            
            #@------#ADD headers (secondary headers) to overstocks costed worksheet#------@#
            secondary_headers_overstocks=["Lineal  Length Tape (M)","Usage FT","Length Tube MT","PN Conduit/Sleeve/Tape",
                                          "Tube Spot Tape-Type","Tube Spot Tape-Qty","Tube Spot Tape-Usage","Cut tape with knife","Cut @ specific length"]
            
            column_start_vstck=len(columns_to_read_oo)+4
            
            for secheader in secondary_headers_overstocks:
                ON.cell(row=4, column=column_start_vstck).value=str(secheader)
                ON.cell(row=4, column=column_start_vstck).font = fontbolnormal
                ON.cell(row=4, column=column_start_vstck).fill=my_filllightgray
                column_start_vstck=column_start_vstck+1
                
            ON.cell(row=2, column=1).value="Overstock List"
            ON.cell(row=2, column=1).font = fontboldbig
            
            ON.cell(row=3, column=1).value="Note 1 - All lengths Rounded To The Nearest mm,   Note 2 - Length Does Not Include Additional Tape For Breakouts Or Additional Tape To Back Of Connector"
            ON.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)
            
            
            #@------#ADD headers (most right)#------@#
            ON.cell(row=3, column=43).value="charge by row [1 tube or sleeve][2 longitudinal tape]"
            ON.cell(row=3, column=44).value="[spot tape usage][longitudinal tape & tube]"
            ON.cell(row=3, column=45).value="Charge by row [1 by reverse tape]"
            ON.cell(row=3, column=46).value="Charge by row [1 by longitudinal tape]"
            
            #@------#ADD special apperance added columns#------@#
            ON.cell(row=4, column=3).font = fontbolnormal
            ON.cell(row=4, column=3).fill=my_filllightgray
            
            ON.cell(row=4, column=5).font = fontbolnormal
            ON.cell(row=4, column=5).fill=my_filllightgray
            
            ON.cell(row=4, column=27).font = fontbolnormal
            ON.cell(row=4, column=27).fill=my_filllightgray
            
            
            #------------------------------------------------------------------------------#      
            #Read Harness Calculations OLD                                                 #
            #------------------------------------------------------------------------------#
            
            #@------#identify row where requested appears#------@#
            
            control="start"
            rowhco=1
            while control != requested:
                cell_obj = HCO.cell(row=rowhco, column = 1)
                control=str(cell_obj.value)

                control_number_column_HC=rowhco
                rowhco=rowhco+1
                if(requested==value):
                    break
            #print("RWO de control HCO!!!!!!!!!!!!!!!!!!!!")    
            #print(control_number_column_HC)
            
            #@------#get number of rows down of requested and write formulas#------@#
            control="start"
            rowhco=control_number_column_HC+1
            rowhcoh=control_number_column_HC
            
            list_wires=[]
            list_rows=[]
            while(control!="None"):
                #wire name
                cell_obj = HCO.cell(row=rowhco, column = 2)
                control=str(cell_obj.value)
            
                if(str(control)!="None"):
                
                    cell_obj2 = HCO.cell(row=rowhco, column = 3)
                    gauge=float(cell_obj2.value)
                    
                    list_wires.append(control+"-"+str('%.2f' %gauge))
                    list_rows.append(rowhco)
                    
                    HCO.cell(row=rowhco, column=16).value='=B'+str(rowhco)+'&'+'"'+'-'+'"'+'&TEXT(C'+str(rowhco)+','+'"0.00")'
                    HCO.cell(row=rowhco, column=17).value='=F'+str(rowhco)+'/1000'
                    HCO.cell(row=rowhco, column=18).value='=CONVERT(F'+str(rowhco)+',"mm","ft")'
                    #Holder
                    HCO.cell(row=rowhco, column=20).value='=S'+str(rowhco)+'/1000'
                    HCO.cell(row=rowhco, column=21).value='=CONVERT(S'+str(rowhco)+',"mm","ft")'
                    HCO.cell(row=rowhco, column=22).value='=Q'+str(rowhco)+'-T'+str(rowhco)
                    HCO.cell(row=rowhco, column=23).value='=CONVERT(V'+str(rowhco)+',"m","ft")'
                    HCO.cell(row=rowhco, column=24).value='=Q'+str(rowhco)+'/T'+str(rowhco)
                    rowhco=rowhco+1
                if(control=="None"):
                    break
                
            #@------#Write headers#------@#
            
            HC_headers_list=["Wire Type","MT","FT","mm","MT","FT","REAL MT","REAL FT"]
            column_start_header_hc=16
            for header in HC_headers_list:
                HCO.cell(row=rowhcoh, column=column_start_header_hc).value=header
                HCO.cell(row=rowhcoh, column=column_start_header_hc).font = fontwhite
                HCO.cell(row=rowhcoh, column=column_start_header_hc).fill=my_filldarkgray
                column_start_header_hc=column_start_header_hc+1

            #------------------------------------------------------------------------------#      
            #Write Assembly nav costed                                                     #
            #------------------------------------------------------------------------------# 
            
            #@------#Write headers#------@#
            headers_to_write1=["Commodity 1","Commodity 2","Connector Classification"
                              ,"GPN","Engineering Feedback","Supplier Part Number",
                              "Description SEM","Type SEM"]
            
            headers_to_write2=["Family","Number","Name", "Component Name","Description","Quantity",
                               str(requested), "Typecode"]
            
            #headers_to_write1
            columns_to_write=[3,4,5,6,7,8,11,12]
            counter2=0
            for header in headers_to_write1:
                columnainicial=columns_to_write[counter2]
                rowinicial=1
                AN.cell(row=rowinicial, column=columnainicial).value=header
                AN.cell(row=rowinicial, column=columnainicial).font = fontwhite
                AN.cell(row=rowinicial, column=columnainicial).fill=my_filldarkgray
                counter2=counter2+1
            
            #headers_to_write2
            columns_to_write=[1,2,9,10,13,14,15,16]
            counter2=0
            for header in headers_to_write2:
                columnainicial=columns_to_write[counter2]
                rowinicial=2
                AN.cell(row=rowinicial, column=columnainicial).value=header
                AN.cell(row=rowinicial, column=columnainicial).font = fontbolnormal
                counter2=counter2+1
            

            for columna in range(16):
                AN.cell(row=2, column=1+columna).fill=my_filllightgray
                
            #@------#Write columns data from original Assembly Nav#------@#
            columns_to_write=[2,9,10,13,14,15]
            counter2=0
            for lista in lists_to_fill:
                columnainicial=columns_to_write[counter2]
                rowinicial=3
                counter3=0
                for cell in lista:
                    if(str(Control[counter3])!="None"):
                        
                        if(columnainicial==2 and str(cell)[0:3]=="SCP"):
                            AN.cell(row=rowinicial, column=1).value=directory_name
                            AN.cell(row=rowinicial, column=3).value='''=IF(I'''+str(rowinicial)+'''="GENERIC SPLICE SEALED","Heatshrink Splice",IF(I'''+str(rowinicial)+'''="GENERIC SPLICE UNSEALED","Splice Unsealed","NOT FOUND"))'''
                            AN.cell(row=rowinicial, column=columnainicial).value="=VLOOKUP(J"+str(rowinicial)+",'Netlist Extraction costed'!A"+str(counterNEO_secondarytable_start_record)+":K"+str(counterNEO_secondarytable_end)+",11,0)"
                            AN.cell(row=rowinicial, column=6).value='=B'+str(rowinicial)
                            AN.cell(row=rowinicial, column=11).value="=VLOOKUP(J"+str(rowinicial)+",'Netlist Extraction costed'!A"+str(counterNEO_secondarytable_start_record)+":K"+str(counterNEO_secondarytable_end)+",5,0)"
                            AN.cell(row=rowinicial, column=12).value='''=IF(I'''+str(rowinicial)+'''="GENERIC SPLICE SEALED","Tube/Non-Reflective/Heatshrink","")'''
                            AN.cell(row=rowinicial, column=14).value='''=IF(I'''+str(rowinicial)+'''="GENERIC SPLICE UNSEALED",0.2931,1)'''
                            AN.cell(row=rowinicial, column=15).value='''=IF(I'''+str(rowinicial)+'''="GENERIC SPLICE UNSEALED","FT","EA")'''
                            AN.cell(row=rowinicial, column=16).value='''=IF(I'''+str(rowinicial)+'''="GENERIC SPLICE SEALED","HSSA","")'''
                        else:
                            if(columnainicial!=15 and columnainicial!=14):
                                AN.cell(row=rowinicial, column=columnainicial).value=cell
                    
                        if(columnainicial==2 and str(cell)[0:3]!="SCP"):
                            AN.cell(row=rowinicial, column=1).value=directory_name
                            AN.cell(row=rowinicial, column=3).value="=VLOOKUP(B"+str(rowinicial)+","+"'"+str(path_TI)+"\[master.xlsx]master'!$A$1:$G$5000,7,0)"
                            AN.cell(row=rowinicial, column=5).value='''=IF(ISERROR(VLOOKUP(J'''+str(rowinicial)+''','Netlist Extraction costed'!$A$'''+str(counterNEO_secondarytable_start_record)+''':$G$'''+str(counterNEO_secondarytable_end)+''',7,0)=TRUE),"",IF(VLOOKUP(J'''+str(rowinicial)+''','Netlist Extraction costed'!$A$'''+str(counterNEO_secondarytable_start_record)+''':$G$'''+str(counterNEO_secondarytable_end)+''',7,0)="SLD","WP",IF(VLOOKUP(J'''+str(rowinicial)+''','Netlist Extraction costed'!$A$'''+str(counterNEO_secondarytable_start_record)+''':$G$'''+str(counterNEO_secondarytable_end)+''',7,0)="UNSLD","NWP","")))'''
                            AN.cell(row=rowinicial, column=6).value="=VLOOKUP(B"+str(rowinicial)+","+"'"+str(path_TI)+"\[master.xlsx]master'!$A$1:$F$356,2,0)"
                            AN.cell(row=rowinicial, column=8).value='''=IF(COUNT(FIND({0,1,2,3,4,5,6,7,8,9},MID(K'''+str(rowinicial)+''',MATCH(2,1/(MID(K'''+str(rowinicial)+''',_xlfn.SEQUENCE(LEN(K'''+str(rowinicial)+''')),1)=","))+1,LEN(K'''+str(rowinicial)+'''))))>0,MID(K'''+str(rowinicial)+''',MATCH(2,1/(MID(K'''+str(rowinicial)+''',_xlfn.SEQUENCE(LEN(K'''+str(rowinicial)+''')),1)=","))+1,LEN(K'''+str(rowinicial)+''')),"")'''
                            AN.cell(row=rowinicial, column=11).value="=VLOOKUP(B"+str(rowinicial)+","+"'"+str(path_TI)+"\[master.xlsx]master'!$A$1:$F$356,3,0)"
                            AN.cell(row=rowinicial, column=12).value="=VLOOKUP(B"+str(rowinicial)+","+"'"+str(path_TI)+"\[master.xlsx]master'!$A$1:$F$356,4,0)"
                            AN.cell(row=rowinicial, column=15).value="=VLOOKUP(B"+str(rowinicial)+","+"'"+str(path_TI)+"\[master.xlsx]master'!$A$1:$F$356,5,0)"
                            AN.cell(row=rowinicial, column=16).value="=VLOOKUP(B"+str(rowinicial)+","+"'"+str(path_TI)+"\[master.xlsx]master'!$A$1:$F$356,6,0)"
                            AN.cell(row=rowinicial, column=14).value=1
                        
                        rowinicial=rowinicial+1
                    counter3=counter3+1
                counter2=counter2+1
            #@------#Write columns data [TERMINALS, SEALS & PLUGS] from original Assembly Nav#------@#  
            for component in list_seals_plugs_terminals:
                AN.cell(row=rowinicial, column=1).value=directory_name
                AN.cell(row=rowinicial, column=2).value=component
                AN.cell(row=rowinicial, column=3).value=component.split("_", 1)[0]
                if(component.split("_", 1)[0]=="Terminal"):
                    AN.cell(row=rowinicial, column=4).value=component.split("_", 2)[1]
                AN.cell(row=rowinicial, column=6).value='''=B'''+str(rowinicial)
                AN.cell(row=rowinicial, column=9).value='''=B'''+str(rowinicial)
                AN.cell(row=rowinicial, column=14).value='''=VLOOKUP(B'''+str(rowinicial)+''','Netlist Extraction costed'!A'''+str(counterNEO_secondarytable_end+2)+''':B'''+str(start_row_termtype)+''',2,0)'''
                AN.cell(row=rowinicial, column=15).value='''EA'''
                
                rowinicial=rowinicial+1
                
            #@------#Write columns data [Wires] from original Assembly Nav#------@#  
            for wire in list_wires:
                AN.cell(row=rowinicial, column=1).value=directory_name
                AN.cell(row=rowinicial, column=2).value=str(wire)
                AN.cell(row=rowinicial, column=3).value="Wire"
                AN.cell(row=rowinicial, column=6).value=str(wire)
                AN.cell(row=rowinicial, column=9).value=str(wire)
                AN.cell(row=rowinicial, column=14).value='''=VLOOKUP(B'''+str(rowinicial)+''','Harness Calculation'!P'''+str(list_rows[0])+''':R'''+str(list_rows[-1])+''',3,0)'''
                AN.cell(row=rowinicial, column=15).value="FT"
    
                rowinicial=rowinicial+1
                
            #@------#Write columns data [Overstocks] from original Assembly Nav#------@#
            list_overstocks = list(dict.fromkeys(list_overstocks))
            list_overstocks.sort()
            for element in list_overstocks:
                element_name=element.split("_", 2)[0]
                element_name2=element.split("_", 2)[1]
                AN.cell(row=rowinicial, column=1).value=directory_name
                AN.cell(row=rowinicial, column=2).value=str(element_name)
                if(element_name2=="Conduit"):
                    AN.cell(row=rowinicial, column=3).value=str(element_name2)
                if(element_name2=="Spiral" or element_name2=="Solid"):
                    AN.cell(row=rowinicial, column=3).value="Tape"
                AN.cell(row=rowinicial, column=6).value=str(element_name)
                AN.cell(row=rowinicial, column=9).value=str(element_name)
                AN.cell(row=rowinicial, column=15).value="FT"
                rowinicial=rowinicial+1
                
            #------------------------------------------------------------------------------#      
            #Write summary components tab                                                  #
            #------------------------------------------------------------------------------# 
            
            #@------#WRITE HEADERS#------@#
            headers_SCC=["Commodity 1","Total","Commodity 2/Type Conn","Totals","Check","Final"]
            startrow_SCC=1
            start_column_SCC=1
            
            
            for header in headers_SCC:
                SCC.cell(row=startrow_SCC, column=start_column_SCC).value=header
                SCC.cell(row=startrow_SCC, column=start_column_SCC).font = fontwhite
                SCC.cell(row=startrow_SCC, column=start_column_SCC).fill=my_filldarkgray
                start_column_SCC=start_column_SCC+1
                
            #@------#Merge works#------@#
            #Clip clam
            SCC.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
            SCC.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)
            SCC.merge_cells(start_row=2, start_column=6, end_row=4, end_column=6)
            SCC.merge_cells(start_row=2, start_column=5, end_row=4, end_column=5)
            
            #Connector
            SCC.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)
            SCC.merge_cells(start_row=5, start_column=2, end_row=6, end_column=2)
            SCC.merge_cells(start_row=5, start_column=6, end_row=6, end_column=6)
            SCC.merge_cells(start_row=5, start_column=5, end_row=6, end_column=5)
            
            #Grommet
            SCC.merge_cells(start_row=7, start_column=1, end_row=8, end_column=1)
            SCC.merge_cells(start_row=7, start_column=2, end_row=8, end_column=2)
            SCC.merge_cells(start_row=7, start_column=6, end_row=8, end_column=6)
            SCC.merge_cells(start_row=7, start_column=5, end_row=8, end_column=5)
            
            #Terminal
            SCC.merge_cells(start_row=9, start_column=1, end_row=11, end_column=1)
            SCC.merge_cells(start_row=9, start_column=2, end_row=11, end_column=2)
            SCC.merge_cells(start_row=9, start_column=6, end_row=11, end_column=6)
            SCC.merge_cells(start_row=9, start_column=5, end_row=11, end_column=5)
            
            #Conduit
            SCC.merge_cells(start_row=12, start_column=1, end_row=13, end_column=1)
            SCC.merge_cells(start_row=12, start_column=2, end_row=13, end_column=2)
            SCC.merge_cells(start_row=12, start_column=6, end_row=13, end_column=6)
            SCC.merge_cells(start_row=12, start_column=5, end_row=13, end_column=5)
            
            #Rigid Tube
            SCC.merge_cells(start_row=14, start_column=1, end_row=15, end_column=1)
            SCC.merge_cells(start_row=14, start_column=2, end_row=15, end_column=2)
            SCC.merge_cells(start_row=14, start_column=6, end_row=15, end_column=6)
            SCC.merge_cells(start_row=14, start_column=5, end_row=15, end_column=5)
            
            #@------#Write Info#------@#
            groups=["Clip/Clamp","Connector","Grommet","Terminal","Conduit","Rigid Tube","Tie Strap",
                    "Eyelet","Cover","Protector (Shield/Channel)","Heatshrink Splice","Heatshrink Bluntcut","Wire Seal","Cavity Plug",
                    "Bolt","Nut","Heatshrink Eyelet","Sleeve","Foil Tape","Hank Tape","Label","Ribbon"]
            groups2=["Bracket","Tape on Clip","Snap on Clip","WP","NWP","Disc","Accordion","NWP","WP",
                     "Terminal Fuse","Slit Tube","Unslit Tube","Slit Tube","Unslit Tube"]
            rows_to_write=[2,5,7,9,12,14,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31]
            rows_to_write2=[2,3,4,5,6,7,8,9,10,11,12,13,14,15]
            
            #column 1
            for i in range(len(groups)):
                SCC.cell(row=rows_to_write[i], column=1).value=groups[i]
                SCC.cell(row=rows_to_write[i], column=2).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,A'''+str(rows_to_write[i])+''')'''
                SCC.cell(row=rows_to_write[i], column=6).value='''=B'''+str(rows_to_write[i])
                
            for i in range(len(groups2)):
                SCC.cell(row=rows_to_write2[i], column=3).value=groups2[i]
                SCC.cell(row=rows_to_write2[i], column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!D3:D5000,'Summary Component count'!C'''+str(rows_to_write2[i])+''')'''
                
            #write formulas Column D
            SCC.cell(row=2, column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,'Summary Component count'!A2,'Assembly Nav costed'!D3:D5000,C2)'''
            SCC.cell(row=3, column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,'Summary Component count'!A2,'Assembly Nav costed'!D3:D5000,C3)'''
            SCC.cell(row=4, column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,'Summary Component count'!A2,'Assembly Nav costed'!D3:D5000,C4)'''
            SCC.cell(row=5, column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,'Summary Component count'!A5:A5,'Assembly Nav costed'!E3:E5000,C5)'''
            SCC.cell(row=6, column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,'Summary Component count'!A5:A5,'Assembly Nav costed'!E3:E5000,C6)'''
            SCC.cell(row=7, column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,'Summary Component count'!A7,'Assembly Nav costed'!D3:D5000,C7)'''
            SCC.cell(row=8, column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,'Summary Component count'!A7,'Assembly Nav costed'!D3:D5000,C8)'''
            SCC.cell(row=9, column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,'Summary Component count'!A9,'Assembly Nav costed'!D3:D5000,C9)'''
            SCC.cell(row=10, column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,'Summary Component count'!A9,'Assembly Nav costed'!D3:D5000,C10)'''
            SCC.cell(row=11, column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,'Summary Component count'!A9,'Assembly Nav costed'!D3:D5000,C11)'''
            SCC.cell(row=12, column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,'Summary Component count'!A12,'Assembly Nav costed'!D3:D5000,C12)'''
            SCC.cell(row=13, column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,'Summary Component count'!A12,'Assembly Nav costed'!D3:D5000,C13)'''
            SCC.cell(row=14, column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,'Summary Component count'!A14,'Assembly Nav costed'!D3:D5000,C14)'''
            SCC.cell(row=15, column=4).value='''=SUMIFS('Assembly Nav costed'!N3:N5000,'Assembly Nav costed'!C3:C5000,'Summary Component count'!A14,'Assembly Nav costed'!D3:D5000,C15)'''
            
            #write formulas Column E
            SCC.cell(row=2, column=5).value='''=B2=SUM(D2:D4)'''
            SCC.cell(row=5, column=5).value='''=B5=SUM(D5:D6)'''
            SCC.cell(row=7, column=5).value='''=B7=SUM(D7:D8)'''
            SCC.cell(row=9, column=5).value='''=B9=SUM(D9:D11)'''
            SCC.cell(row=12, column=5).value='''=B12=SUM(D12:D13)'''
            SCC.cell(row=14, column=5).value='''=B14=SUM(D14:D15)'''
            
            #@------#Write Wires part of table#------@#
            SCC.merge_cells(start_row=32, start_column=1, end_row=len(list_wires)+31, end_column=1)
            SCC.cell(row=32, column=1).value="Wire"
            
            SCC.merge_cells(start_row=32, start_column=2, end_row=len(list_wires)+31, end_column=2)
            SCC.cell(row=32, column=2).value="""=SUMIF('Assembly Nav costed'!C3:C5000,"Wire",'Assembly Nav costed'!N3:N5000)"""
            
            SCC.merge_cells(start_row=32, start_column=5, end_row=len(list_wires)+31, end_column=5)
            SCC.cell(row=32, column=5).value='''=SUM(D32:D'''+str(len(list_wires)+31)+''')=B32'''
            
            SCC.merge_cells(start_row=32, start_column=6, end_row=len(list_wires)+31, end_column=6)
            
            start_row_wire=32
            for wire in list_wires:
                SCC.cell(row=start_row_wire, column=3).value=str(wire)
                SCC.cell(row=start_row_wire, column=4).value="""=SUMIF('Assembly Nav costed'!$B$3:$B$5000,'Summary Component count'!C"""+str(start_row_wire)+""",'Assembly Nav costed'!$N$3:$N$5000)"""
                start_row_wire=start_row_wire+1

            book_bom.save(directory_name+".xlsx")
                        
        counter=counter+1
        
        #------------------------------------------------------------------------------#      
        #Write summary components tab                                                  #
        #------------------------------------------------------------------------------# 
    
        os.chdir(path_calculations_folder+"/"+directory_name)
        
        #------------------------------------------------------------------------------#      
        #Write data LMI                                                                #
        #------------------------------------------------------------------------------# 
        
        book_LMI = openpyxl.load_workbook(directory_name+"_LMI"+".xlsx")
            
        #@------#Read old sheets#------@#
        LR=book_LMI["Labor Report"]
        SPL=book_LMI["Splicing"]
        CWWL=book_LMI["Cut wires with length"]
        SO=book_LMI["Special Operations"]
        TW=book_LMI["Twisting"]
        
        #@------#Write twist Info#------@#
        start_row_twist_sheet=10
        for key in dict_twist:
            values=dict_twist.get(key)
            TW.cell(row=start_row_twist_sheet, column=2).value=int(values[0])
            TW.cell(row=start_row_twist_sheet, column=3).value=float(values[1]/1000)
            start_row_twist_sheet=start_row_twist_sheet+1

        #@------#Get splices information#------@#
        list_splices=[]
        for data in list_to_use_splices:
            if(str(data)[0:3]=="STD"):
                list_splices.append(data)
        splices_dict=Counter(list_splices)
        list_splices2=list(splices_dict.values())
        splices_dict2=Counter(list_splices2)
    
        #@------#Write initial data#------@#
        LR.cell(row=6, column=8).value=int(counterNEO-1)
        LR.cell(row=9, column=11).value=directory_name
        
        #@------#Write splices sheet#------@#
        for key in splices_dict2:
            qty=splices_dict2.get(key)
            SPL.cell(row=int(key)+3, column=3).value=int(qty)
            
            
        #@------#Direct component Data#------@#
        #Clip/clamp Tape on Clip
        LR.cell(row=113, column=11).value="""='"""+str(path_calculations_folder)+"""/"""+str(directory_name)+"""\["""+str(directory_name)+""".xlsx]Summary Component count'!$D$3"""
        #Clip/clamp snap on clip
        LR.cell(row=94, column=11).value="""='"""+str(path_calculations_folder)+"""/"""+str(directory_name)+"""\["""+str(directory_name)+""".xlsx]Summary Component count'!$D$4"""
        #Connectors Final
        LR.cell(row=102, column=11).value="""='"""+str(path_calculations_folder)+"""/"""+str(directory_name)+"""\["""+str(directory_name)+""".xlsx]Summary Component count'!$D$5-SUM(K103:K104)"""
        LR.cell(row=106, column=11).value="""=SUM(K102:K104)"""
        #Terminal
        LR.cell(row=131, column=11).value="""='"""+str(path_calculations_folder)+"""/"""+str(directory_name)+"""\["""+str(directory_name)+""".xlsx]Summary Component count'!$D$9"""
        LR.cell(row=132, column=11).value="""='"""+str(path_calculations_folder)+"""/"""+str(directory_name)+"""\["""+str(directory_name)+""".xlsx]Summary Component count'!$D$10"""
        #Tie Strap
        LR.cell(row=110, column=11).value="""='"""+str(path_calculations_folder)+"""/"""+str(directory_name)+"""\["""+str(directory_name)+""".xlsx]Summary Component count'!$F$16"""
        #Heatshrink Splice
        LR.cell(row=50, column=11).value="""='"""+str(path_calculations_folder)+"""/"""+str(directory_name)+"""\["""+str(directory_name)+""".xlsx]Summary Component count'!$F$20"""
        #Plugs
        LR.cell(row=69, column=11).value="""='"""+str(path_calculations_folder)+"""/"""+str(directory_name)+"""\["""+str(directory_name)+""".xlsx]Summary Component count'!$F$23"""
        #Seals
        CWWL.cell(row=16, column=7).value="""='"""+str(path_calculations_folder)+"""/"""+str(directory_name)+"""\["""+str(directory_name)+""".xlsx]Summary Component count'!$F$22"""
        SO.cell(row=79, column=6).value="""='"""+str(path_calculations_folder)+"""/"""+str(directory_name)+"""\["""+str(directory_name)+""".xlsx]Summary Component count'!$F$22"""
        #Label
        LR.cell(row=135, column=11).value=1
        
        book_LMI.save(directory_name+"_LMI"+".xlsx")
    print("Terminado :D")
        
def ODM2():
    
    #------------------------------------------------------------------------------#      
    #Select program folder and get paths of calculations and Technical Information #
    #------------------------------------------------------------------------------# 
    pyautogui.alert(text='Select technical information files', title='Select File', button='OK')
    path2 = askdirectory(title='Select Folder') # shows dialog box and return the path
    
    #print(path2)
    list_of_directories=os.listdir(path2)
    os.chdir(path2)
    list_components=[]
    print(list_of_directories)
    for file in list_of_directories:
        book_bom = openpyxl.load_workbook(file)
        print(file)
        
        #@------#Read old sheets#------@#
        BS=book_bom["Assembly Nav"]

        #print(directory_name+"/"+directory_name+".xlsx")
        number_of_rows=BS.max_row
        print(number_of_rows)

        #------------------------------------------------------------------------------#      
        #Read Assembly nav original                                                    #
        #------------------------------------------------------------------------------# 
        
        
        #@------#identify requested PN column number#------@#
        for numberofrow in range (number_of_rows-2):
            cell_obj = BS.cell(row = numberofrow+3, column = 1)
            cell_obj = BS.cell(row = numberofrow+3, column = 1)
            control=str(cell_obj.value)

            list_components.append(control)
            
    list_components = list(dict.fromkeys(list_components))
    list_components.sort()
    print(list_components)

    mw = Workbook()
    mw.save("master.xlsx")
    mw.create_sheet('master')
    std=mw.get_sheet_by_name('Sheet')
    mw.remove_sheet(std)
    MS=mw['master']
    
    for element in range(len(list_components)):
        MS.cell(row=element+1, column=1).value=list_components[element]
        MS.cell(row=element+1, column=2).value="""=_xlfn.XLOOKUP(A"""+str(element+1)+""",'"""+str(path2)+"""\[slave1.xls]Innovator'!$D$2:$D$5000,'"""+str(path2)+"""\[slave1.xls]Innovator'!$A$2:$A$5000)"""
        MS.cell(row=element+1, column=3).value="""=VLOOKUP(B"""+str(element+1)+""",'"""+str(path2)+"""\[slave2.xls]Innovator'!$A$2:$P$5000,3,0)"""
        MS.cell(row=element+1, column=4).value="""=VLOOKUP(B"""+str(element+1)+""",'"""+str(path2)+"""\[slave2.xls]Innovator'!$A$2:$P$5000,4,0)"""
        MS.cell(row=element+1, column=5).value="""=VLOOKUP(B"""+str(element+1)+""",'"""+str(path2)+"""\[slave2.xls]Innovator'!$A$2:$P$5000,12,0)"""
        MS.cell(row=element+1, column=6).value="""=VLOOKUP(B"""+str(element+1)+""",'"""+str(path2)+"""\[slave2.xls]Innovator'!$A$2:$P$5000,16,0)"""
    mw.save("master.xlsx")
    
def OpenUrl():
    webbrowser.open_new("https://learcorporation-my.sharepoint.com/:f:/g/personal/lcardenasmontaz_lear_com/EmhvprrX705FoDRji8Od_IcBuStgZ4Ak21wpSC-t_WQUGg?e=nHMbKu")

#GUI-------------------------------------------------------------------------------------------------------------------------------------------------
root = tk.Tk()
root.title("CP Suite Tool")
root.geometry("530x160")

root.resizable(False, False)

var = IntVar()
#XML BUTTONS


Button_XML_BOM = tk.Button(root,
                   text="BOM Extraction",bg="white",relief="ridge",
                   command=Button_XML_BOM).grid(pady=5, padx=1,row=1,column=1,sticky=N+S+E+W,rowspan=1)

Button_XML_Circuitry = tk.Button(root,
                   text="Circuitry Extraction",bg="white",relief="ridge",
                   command=Button_XML_Circuitry).grid(pady=5, padx=1,row=2,column=1,sticky=N+S+E+W,rowspan=1)

Button_PDF_Multiple = tk.Button(root,
                   text="PDF BATCH",bg="white",relief="ridge",
                   command=Button_PDF_Multiple).grid(pady=5, padx=1,row=1,column=2,sticky=N+S+E+W,rowspan=1)

Button_PDF_onefile = tk.Button(root,
                   text="PDF Onefile",bg="white",relief="ridge",
                   command=Button_PDF_onefile).grid(pady=5, padx=1,row=2,column=2,sticky=N+S+E+W,rowspan=1)

Button_comparer = tk.Button(root,
                   text="Modules report",bg="white",relief="ridge",
                   command=Button_Comparer).grid(pady=5, padx=1,row=1,column=3,sticky=N+S+E+W,rowspan=1)

Button_comparer2 = tk.Button(root,
                   text="Global Report",bg="white",relief="ridge",
                   command=Button_Comparer_V2).grid(pady=5, padx=1,row=2,column=3,sticky=N+S+E+W,rowspan=1)


Button_BC_Prepare =tk.Button(root,
                   text="Select Folder",bg="white",relief="ridge",
                   command=BOM_CUTSHEET_PREPARATION).grid(pady=5, padx=1,row=1,column=4,sticky=N+S+E+W,rowspan=2)

Button_BC_Prepare =tk.Button(root,
                   text="Select Folder",bg="white",relief="ridge",
                   command=ODM).grid(pady=5, padx=1,row=2,column=5,sticky=N+S+E+W,rowspan=1)

Button_BC_Prepare =tk.Button(root,
                   text="Select Files",bg="white",relief="ridge",
                   command=ODM2).grid(pady=5, padx=1,row=1,column=5,sticky=N+S+E+W,rowspan=1)

Button_info = tk.Button(root,
                   text="How do i use these scripts?",bg="white",relief="ridge",
                   command=OpenUrl).grid(pady=5, padx=1,row=7,column=1,sticky=N+S+E+W,columnspan=5)



# THIS STYLE FOR THE PROGRESSBAR
style = ttk.Style(root)
style.layout('text.Horizontal.TProgressbar',
             [('Horizontal.Progressbar.trough',
               {'children': [('Horizontal.Progressbar.pbar',
                              {'side': 'left', 'sticky': 'ns'})],
                'sticky': 'nswe'}),
              ('Horizontal.Progressbar.label', {'sticky': ''})])      # ,lightcolor=None,bordercolo=None,darkcolor=None
style.configure('text.Horizontal.TProgressbar', text='0 %')

bar = Progressbar(root, length=150, style='black.Horizontal.TProgressbar')
bar.grid(column=1, row=15,columnspan=5, padx=1,sticky=N+S+E+W,rowspan=3)



Color_Labels="#424242"
#Labels
label_XML = Label(root,text="XML Extractor (TESLA)",bg=Color_Labels,fg="white").grid(row=0,column=1,sticky=W+E,padx=1)
label_XML2 = Label(root,text="PO PDF Extractor (Daimler)",bg=Color_Labels,fg="white").grid(row=0,column=2,sticky=W+E,padx=1)
label_XML3 = Label(root,text="Compare Tool",bg=Color_Labels,fg="white").grid(row=0,column=3,sticky=W+E,padx=1)
label_XML4 = Label(root,text="Prepare C&B",bg=Color_Labels,fg="white").grid(row=0,column=4,sticky=W+E,padx=1)
label_XML4 = Label(root,text="ODM",bg=Color_Labels,fg="white").grid(row=0,column=5,sticky=W+E,padx=1)
#label_XML3 = Label(root,text="PDF Extractor",bg=Color_Labels,fg="white").grid(pady=5,row=11,column=1,sticky=W+E,padx=1,columnspan=2)
root.mainloop()
